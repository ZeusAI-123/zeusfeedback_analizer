import streamlit as st
import pandas as pd
import re, os, sqlite3, uuid, time, io, hashlib, json, random
from datetime import datetime
from bs4 import BeautifulSoup
import requests
import nltk
import openai
from dotenv import load_dotenv
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import NMF
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import WordNetLemmatizer

# ── DB driver imports ────────────────────────────────────────────────────────
try:
    import mysql.connector
    MYSQL_AVAILABLE = True
except ImportError:
    MYSQL_AVAILABLE = False

try:
    import psycopg2
    POSTGRES_AVAILABLE = True
except ImportError:
    POSTGRES_AVAILABLE = False

try:
    import snowflake.connector
    SNOWFLAKE_AVAILABLE = True
except ImportError:
    SNOWFLAKE_AVAILABLE = False

try:
    import pymongo
    MONGO_AVAILABLE = True
except ImportError:
    MONGO_AVAILABLE = False

try:
    import pymssql
    MSSQL_AVAILABLE = True
except ImportError:
    MSSQL_AVAILABLE = False

try:
    from google.cloud import bigquery as bq_lib
    BIGQUERY_AVAILABLE = True
except ImportError:
    BIGQUERY_AVAILABLE = False

try:
    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.analytics.data_v1beta.types import RunReportRequest, DateRange, Metric, Dimension
    GA_AVAILABLE = True
except ImportError:
    GA_AVAILABLE = False

try:
    from databricks import sql as dbsql
    DATABRICKS_AVAILABLE = True
except ImportError:
    DATABRICKS_AVAILABLE = False

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from webdriver_manager.chrome import ChromeDriverManager
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")

st.set_page_config(
    page_title="ZEUS FEEDBACK ANALYZER",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@400;600;800&display=swap');
html,body,[class*="css"]{font-family:'Syne',sans-serif;background-color:#0d0d0d;color:#f0ece2;}
#MainMenu,footer{visibility:hidden;}
.stApp{background:linear-gradient(135deg,#0d0d0d 0%,#111827 100%);}
.hero-title{font-family:'Syne',sans-serif;font-weight:800;font-size:3.2rem;background:linear-gradient(90deg,#f9a825,#ff6f00,#e91e63);-webkit-background-clip:text;-webkit-text-fill-color:transparent;letter-spacing:-1px;line-height:1.1;margin-bottom:0.2rem;}
.hero-sub{font-family:'Space Mono',monospace;font-size:0.85rem;color:#888;letter-spacing:2px;text-transform:uppercase;margin-bottom:2rem;}
.card{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:16px;padding:1.5rem 2rem;margin-bottom:1.5rem;backdrop-filter:blur(10px);}
.card-title{font-family:'Space Mono',monospace;font-size:0.75rem;letter-spacing:3px;text-transform:uppercase;color:#f9a825;margin-bottom:1rem;}
.badge-0{background:rgba(249,168,37,0.15);border:1px solid #f9a825;color:#f9a825;}
.badge-1{background:rgba(233,30,99,0.15);border:1px solid #e91e63;color:#e91e63;}
.badge-2{background:rgba(0,188,212,0.15);border:1px solid #00bcd4;color:#00bcd4;}
.badge-3{background:rgba(76,175,80,0.15);border:1px solid #4caf50;color:#4caf50;}
.badge-4{background:rgba(156,39,176,0.15);border:1px solid #9c27b0;color:#9c27b0;}
.badge{border-radius:20px;padding:2px 10px;font-size:0.72rem;font-family:'Space Mono',monospace;font-weight:700;}
.source-tag{display:inline-block;background:rgba(249,168,37,0.1);border:1px solid rgba(249,168,37,0.4);border-radius:4px;font-family:'Space Mono',monospace;font-size:0.65rem;padding:2px 8px;color:#f9a825;margin-right:6px;}
.session-box{background:rgba(0,188,212,0.07);border:1px solid rgba(0,188,212,0.25);border-radius:10px;padding:0.8rem 1.2rem;font-family:'Space Mono',monospace;font-size:0.72rem;color:#00bcd4;letter-spacing:0.5px;margin-bottom:1rem;}
.feedback-id{display:inline-block;background:rgba(0,188,212,0.15);border:1px solid rgba(0,188,212,0.5);border-radius:6px;padding:2px 10px;font-family:'Space Mono',monospace;font-size:0.72rem;color:#00bcd4;margin-right:8px;font-weight:700;}
.db-badge-connected{display:inline-block;background:rgba(76,175,80,0.15);border:1px solid #4caf50;color:#4caf50;border-radius:6px;padding:2px 12px;font-family:'Space Mono',monospace;font-size:0.68rem;font-weight:700;}
.db-badge-avail{display:inline-block;background:rgba(249,168,37,0.1);border:1px solid rgba(249,168,37,0.4);color:#f9a825;border-radius:6px;padding:2px 12px;font-family:'Space Mono',monospace;font-size:0.68rem;font-weight:700;}
.db-badge-na{display:inline-block;background:rgba(120,120,120,0.1);border:1px solid #555;color:#666;border-radius:6px;padding:2px 12px;font-family:'Space Mono',monospace;font-size:0.68rem;}
.db-card{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.07);border-radius:12px;padding:1.2rem 1.5rem;margin-bottom:1rem;}
.stTabs [data-baseweb="tab-list"]{gap:8px;border-bottom:1px solid rgba(255,255,255,0.1);}
.stTabs [data-baseweb="tab"]{border-radius:8px 8px 0 0;font-family:'Space Mono',monospace;font-size:0.78rem;letter-spacing:1px;text-transform:uppercase;color:#888;padding:10px 20px;border:none;background:transparent;}
.stTabs [aria-selected="true"]{background:rgba(249,168,37,0.1)!important;color:#f9a825!important;border-bottom:2px solid #f9a825!important;}
.stButton>button{background:linear-gradient(135deg,#f9a825,#ff6f00);color:#0d0d0d;font-family:'Space Mono',monospace;font-weight:700;font-size:0.82rem;letter-spacing:2px;text-transform:uppercase;border:none;border-radius:8px;padding:0.6rem 2rem;width:100%;transition:all 0.2s;}
.stButton>button:hover{transform:translateY(-2px);box-shadow:0 8px 24px rgba(249,168,37,0.4);}
.stTextInput input,.stTextArea textarea{background:rgba(255,255,255,0.05)!important;border:1px solid rgba(255,255,255,0.12)!important;border-radius:8px!important;color:#f0ece2!important;font-family:'Space Mono',monospace!important;font-size:0.85rem!important;}
[data-testid="stSidebar"]{display:block !important;visibility:visible !important;background:rgba(17,24,39,0.98) !important;border-right:1px solid rgba(249,168,37,0.25) !important;min-width:260px !important;}
[data-testid="stSidebar"] > div{visibility:visible !important;display:block !important;}
section[data-testid="stSidebarContent"]{display:block !important;visibility:visible !important;}
header{visibility:visible !important;}
button[kind="header"]{visibility:visible !important;}
header .stAppHeader{visibility:hidden;}
[data-testid="stMetric"]{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:1rem;}
[data-testid="stMetricValue"]{color:#f9a825!important;font-family:'Space Mono',monospace!important;}
hr{border-color:rgba(255,255,255,0.08);}
.pulse{animation:pulse 2s infinite;}
@keyframes pulse{0%,100%{opacity:1;}50%{opacity:0.5;}}
.col-chip{display:inline-block;background:rgba(249,168,37,0.1);border:1px solid rgba(249,168,37,0.3);border-radius:20px;padding:2px 12px;font-family:'Space Mono',monospace;font-size:0.68rem;color:#f9a825;margin:2px 3px;}
.step-bar{display:flex;gap:0;margin-bottom:1.5rem;}
.step-item{flex:1;padding:0.6rem 1rem;font-family:'Space Mono',monospace;font-size:0.68rem;letter-spacing:1px;text-transform:uppercase;text-align:center;border:1px solid rgba(255,255,255,0.08);}
.step-item:first-child{border-radius:8px 0 0 8px;}
.step-item:last-child{border-radius:0 8px 8px 0;}
.step-active{background:rgba(249,168,37,0.15);border-color:#f9a825!important;color:#f9a825;font-weight:700;}
.step-done{background:rgba(76,175,80,0.1);border-color:#4caf50!important;color:#4caf50;}
.step-pending{background:rgba(255,255,255,0.02);color:#555;}
.table-chip{display:inline-block;background:rgba(0,188,212,0.08);border:1px solid rgba(0,188,212,0.3);border-radius:6px;padding:3px 12px;font-family:'Space Mono',monospace;font-size:0.7rem;color:#00bcd4;margin:3px 4px;cursor:pointer;}
.table-chip:hover{background:rgba(0,188,212,0.18);border-color:#00bcd4;}
[data-testid="stSidebar"] .nav-btn-active > button{background:linear-gradient(135deg,#f9a825,#ff6f00) !important;color:#0d0d0d !important;font-weight:700 !important;border:none !important;box-shadow:0 4px 14px rgba(249,168,37,0.35) !important;}
[data-testid="stSidebar"] .nav-btn-inactive > button{background:rgba(255,255,255,0.04) !important;color:#888 !important;border:1px solid rgba(255,255,255,0.08) !important;font-weight:400 !important;}
[data-testid="stSidebar"] .nav-btn-inactive > button:hover{background:rgba(249,168,37,0.08) !important;color:#f9a825 !important;border-color:rgba(249,168,37,0.3) !important;}
.stButton>button:disabled,.stButton>button[disabled]{background:rgba(255,255,255,0.06) !important;color:#555 !important;cursor:not-allowed !important;box-shadow:none !important;transform:none !important;border:1px solid rgba(255,255,255,0.1) !important;}
.dp-user{background:rgba(249,168,37,0.07);border-left:3px solid #f9a825;border-radius:6px;padding:0.6rem 1rem;margin:4px 0;font-size:0.83rem;color:#ddd;}
.dp-ai{background:rgba(0,188,212,0.07);border-left:3px solid #00bcd4;border-radius:6px;padding:0.6rem 1rem;margin:4px 0;font-size:0.83rem;color:#ddd;}
.conf-card{background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.08);border-radius:8px;padding:0.6rem;text-align:center;}
.conf-card-top{background:rgba(249,168,37,0.12);border:1px solid rgba(249,168,37,0.35);border-radius:8px;padding:0.6rem;text-align:center;}
.meta-row{display:flex;align-items:center;padding:5px 0;border-bottom:1px solid rgba(255,255,255,0.04);font-family:'Space Mono',monospace;font-size:0.7rem;}
.meta-col-name{color:#f9a825;min-width:160px;font-weight:700;}
.meta-col-type{color:#00bcd4;min-width:100px;}
.meta-col-null{color:#888;min-width:80px;}
.meta-col-sample{color:#ccc;flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
</style>
""", unsafe_allow_html=True)

@st.cache_resource
def download_nltk():
    for p in ['stopwords', 'punkt', 'wordnet', 'punkt_tab']:
        nltk.download(p, quiet=True)
download_nltk()

TOPIC_LABELS = {
    0: "⚡ Service Quality", 1: "📦 Product Issues",
    2: "💬 Communication",  3: "⏱️ Speed & Delays",
    4: "💡 Innovation & Features"
}
TOPIC_COLORS = {0:"badge-0",1:"badge-1",2:"badge-2",3:"badge-3",4:"badge-4"}

# ══════════════════════════════════════════════════════════════
# DB DRIVER REGISTRY
# ══════════════════════════════════════════════════════════════
DB_DRIVERS = {
    "MySQL / MariaDB":     ("mysql-connector-python",        MYSQL_AVAILABLE),
    "PostgreSQL":          ("psycopg2-binary",               POSTGRES_AVAILABLE),
    "MongoDB":             ("pymongo",                       MONGO_AVAILABLE),
    "SQLite (file)":       ("built-in",                      True),
    "SQL Server (MSSQL)":  ("pymssql",                       MSSQL_AVAILABLE),
    "Snowflake":           ("snowflake-connector-python",    SNOWFLAKE_AVAILABLE),
    "Google BigQuery":     ("google-cloud-bigquery",         BIGQUERY_AVAILABLE),
    "Google Analytics 4":  ("google-analytics-data",         GA_AVAILABLE),
    "Databricks SQL":      ("databricks-sql-connector",      DATABRICKS_AVAILABLE),
}
ALL_DB_TYPES = list(DB_DRIVERS.keys())

# ══════════════════════════════════════════════════════════════
# SMART COLUMN DETECTION DICTIONARIES
# ══════════════════════════════════════════════════════════════
FEEDBACK_COL_EXACT = [
    "feedback","review","comment","text","body","description","message","opinion",
    "response","remark","note","testimonial","content","reply","remarks","review_text",
    "comment_text","feedback_text","user_comment","customer_feedback","customer_review",
    "review_body","comment_body","post_text","tweet","post","answer","complaint",
    "suggestion","verbatim","open_ended","narrative","user_feedback","product_review",
    "service_review","user_review","review_description","feedback_description",
    "customer_comment","client_feedback","patient_feedback","employee_feedback",
    "survey_response","survey_text","survey_answer","nps_comment","nps_feedback",
    "csat_comment","open_text","free_text","freetext","qualitative","written_feedback",
    "written_review","text_response","long_text","details","additional_comments",
    "additional_feedback","comments","reviews","feedbacks","responses",
]
FEEDBACK_COL_PARTIAL = [
    "review","feedback","comment","opinion","remark","suggest","complain",
    "response","testimonial","verbatim","narrative","text","body","description",
    "message","reply","content","note","answer","complaint","survey",
]
NAME_COL_EXACT = [
    "name","author","reviewer","username","user","customer","user_name","author_name",
    "reviewer_name","customer_name","full_name","display_name","posted_by","written_by",
    "reviewer_id","customer_id","client_name","patient_name","employee_name",
    "respondent","submitter","first_name","screen_name","handle",
]
DATE_COL_EXACT = [
    "date","created_at","posted_at","timestamp","review_date","feedback_date",
    "created_date","submission_date","updated_at","time","when","datetime",
    "post_date","review_time","submitted_at","recorded_at","entry_date",
    "response_date","date_created","date_posted","date_submitted","date_reviewed",
]

# ══════════════════════════════════════════════════════════════
# SENTIMENT ENGINE
# ══════════════════════════════════════════════════════════════
_POS = re.compile(
    r'\b(excellent|amazing|outstanding|wonderful|fantastic|great|good|love|loved|brilliant|superb|perfect|'
    r'delicious|awesome|fabulous|impressive|enjoy|enjoyed|pleasant|satisfied|satisfaction|happy|pleased|'
    r'best|highly recommend|recommend|worth|value for money|value for price|tasty|fresh|'
    r'friendly|helpful|attentive|professional|polite|courteous|efficient|prompt|quick|fast|'
    r'clean|hygienic|comfortable|cozy|ambience|nice place|will visit again|visit again|'
    r'must visit|must try|5 star|five star|10/10|loved it|exceeded expectations|top notch|'
    r'world class|phenomenal|spectacular|mouth.?watering|finger.?licking|well cooked|'
    r'perfectly cooked|good service|great service|good food|great food|amazing food|'
    r'yummy|scrumptious|lip.?smacking)\b', re.I)

_NEG = re.compile(
    r'\b(terrible|horrible|awful|worst|bad|poor|disappointing|disappointed|disgusting|pathetic|'
    r'rude|arrogant|unprofessional|impolite|hostile|aggressive|misbehav|attitude problem|'
    r'tasteless|bland|overcooked|undercooked|half.?cooked|raw|stale|rubbery|dry|cold food|'
    r'slow|delay|wait|waited|too long|overcrowded|understaffed|dirty|unclean|unhygienic|'
    r'overpriced|expensive|rip.?off|not worth|waste of money|never again|avoid|'
    r'do not recommend|not recommend|waste|pest|cockroach|insect|filth|stink|bad smell|'
    r'missing|wrong order|incorrect order|food poisoning|sick|fell ill)\b', re.I)

_MIXED_IND = re.compile(
    r'\bbut\b|\bhowever\b|\bthough\b|\balthough\b|\bexcept\b|\bdespite\b|\byet\b|'
    r'\bon the other hand\b|\bpartially\b|\bnot bad\b|\bok but\b', re.I)

def rule_based_sentiment(text):
    if not text or len(text.strip()) < 10: return 'Neutral'
    t = text.lower().strip()
    pos = len(_POS.findall(t)); neg = len(_NEG.findall(t))
    if pos >= 1 and neg >= 1 and _MIXED_IND.search(t): return 'Mixed'
    if pos > neg * 1.5: return 'Positive'
    if neg > pos * 1.5: return 'Negative'
    if pos > 0 and neg > 0: return 'Mixed'
    if pos > 0: return 'Positive'
    if neg > 0: return 'Negative'
    mild_pos = re.search(r'\b(nice|okay|ok|fine|decent|average|alright|not bad|acceptable)\b', t)
    mild_neg = re.search(r'\b(not great|could be better|needs improvement|lacking|mediocre|underwhelming)\b', t)
    if mild_pos and not mild_neg: return 'Positive'
    if mild_neg and not mild_pos: return 'Negative'
    if mild_pos and mild_neg: return 'Mixed'
    return 'Neutral'

def get_sentiment_batch_ai(feedback_list):
    if not openai.api_key or not feedback_list:
        return [rule_based_sentiment(fb) for fb in feedback_list]
    try:
        prompt = ("Classify sentiment of each feedback. Reply ONLY with numbered lines:\n"
                  "1. Positive\n2. Negative\n...\nValid: Positive, Negative, Neutral, Mixed\n\nFeedbacks:\n")
        for i, fb in enumerate(feedback_list, 1):
            prompt += f"{i}. {str(fb)[:300]}\n"
        resp = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[{"role":"system","content":"Precise sentiment classifier. Output only numbered labels."},
                      {"role":"user","content":prompt}],
            max_tokens=len(feedback_list)*8+50, temperature=0.0)
        raw = resp["choices"][0]["message"]["content"].strip()
        results = {}
        for line in raw.split('\n'):
            m = re.match(r'^(\d+)\.\s*(Positive|Negative|Neutral|Mixed)', line.strip(), re.I)
            if m: results[int(m.group(1))-1] = m.group(2).capitalize()
        return [results.get(i, rule_based_sentiment(fb)) for i, fb in enumerate(feedback_list)]
    except Exception:
        return [rule_based_sentiment(fb) for fb in feedback_list]

def analyze_sentiments_all(feedback_list):
    rule_results = [rule_based_sentiment(fb) for fb in feedback_list]
    if not openai.api_key: return rule_results
    uncertain_idx = [i for i, s in enumerate(rule_results) if s == 'Neutral']
    if not uncertain_idx: return rule_results
    uncertain_texts = [feedback_list[i] for i in uncertain_idx]
    ai_results = []
    for i in range(0, len(uncertain_texts), 20):
        ai_results.extend(get_sentiment_batch_ai(uncertain_texts[i:i+20]))
    for orig_idx, ai_sent in zip(uncertain_idx, ai_results):
        rule_results[orig_idx] = ai_sent
    return rule_results

# ══════════════════════════════════════════════════════════════
# FALLBACK SUGGESTIONS
# ══════════════════════════════════════════════════════════════
_KW_ACTIONS = [
    (r'made.*wait|waited.*long|took.*too long|long.*queue|slow.*service|delay.*order|waited.*hour|no.*refill',
     "Audit the service timeline at peak hours, assign a dedicated expediter role during rush periods, and set a 10-minute maximum wait target with manager alerts when breached."),
    (r'rude|behaviour|attitude|unprofessional|hostile|impolite|arrogant|disrespectful|misbehav',
     "Launch a monthly frontline staff empathy workshop, introduce peer-review scorecards, and tie guest-satisfaction ratings directly to service team performance incentives."),
    (r'tasteless|no taste|bland|overcooked|undercooked|half.?cooked|raw.*meat|cold.*food|stale|rubbery|insipid|dry.*chicken',
     "Establish a kitchen quality-control checkpoint: assign a dedicated quality officer per shift to inspect every dish before it leaves the kitchen, with a discard protocol for sub-standard items."),
    (r'overpriced|not worth|rip.?off|too expensive|poor.*value|price.*hike|price.*increase|not.*worth.*money',
     "Introduce a transparent value communication strategy — highlight premium ingredients on menus, publish cost-per-portion data for high-value items, and launch a loyalty tier with exclusive repeat-visitor pricing."),
    (r'dirty|unhygienic|unclean|pest|cockroach|insect|filth|stink|smell.*bad|bad.*smell|not.*clean',
     "Schedule bi-weekly hygiene audits with a third-party inspector, post the latest inspection score at the entrance, and create an anonymous staff reporting channel for cleanliness violations."),
]
_TOPIC_ACTIONS = {
    0:["Introduce a real-time floor-monitoring dashboard for managers.",
       "Deploy a post-meal digital feedback kiosk at exit points.",
       "Establish a mystery dining programme for quarterly audits.",
       "Create a 'Service Champion' recognition system for floor staff.",
       "Map the guest journey and identify the three highest-friction touchpoints."],
    1:["Implement a live product quality log for the kitchen.",
       "Introduce a guest-facing freshness timer display at each buffet station.",
       "Set up a weekly tasting panel to score each dish against a benchmark.",
       "Conduct a root-cause analysis on the top 3 most-complained-about dishes.",
       "Create a feedback card specifically rating food quality."],
    2:["Train all floor staff on a standard 3-step proactive communication script.",
       "Introduce a pre-visit automated WhatsApp message confirming reservations.",
       "Set up a post-visit email sequence with feedback survey within 24 hours.",
       "Create a laminated 'Guest FAQ' on each table covering top 10 common questions.",
       "Implement a structured complaint-handling protocol."],
    3:["Introduce staggered entry slots for large group reservations.",
       "Install a kitchen display system linking front-of-house seating data.",
       "Set a maximum 8-minute starter replenishment SLA during peak hours.",
       "Analyse historical reservation data to identify the busiest 2-hour window.",
       "Pilot a 'Fast Lane' table category for guests with 60-minute dining windows."],
    4:["Launch a 'Guest Chef' quarterly event.",
       "Introduce a digital loyalty passport.",
       "Create a seasonal limited-edition menu released every 90 days.",
       "Set up a live innovation lab at one flagship outlet.",
       "Partner with a local culinary school for a rotating feature."],
}
_topic_ctr = {k: 0 for k in _TOPIC_ACTIONS}

def generate_fallback_suggestion(feedback_text, topic_id):
    t = (feedback_text or '').lower()
    for pattern, suggestion in _KW_ACTIONS:
        if re.search(pattern, t): return suggestion
    bucket = int(hashlib.md5(t[:300].encode()).hexdigest()[:8], 16)
    templates = _TOPIC_ACTIONS.get(topic_id, _TOPIC_ACTIONS[0])
    idx = bucket % len(templates)
    prev = _topic_ctr.get(topic_id, 0)
    if idx == prev and len(templates) > 1: idx = (idx+1) % len(templates)
    _topic_ctr[topic_id] = idx
    return templates[idx]

# ══════════════════════════════════════════════════════════════
# SQLITE APP DATABASE
# ══════════════════════════════════════════════════════════════
DB_PATH = "zeus_feedback.db"
SOURCE_PREFIX_MAP = {
    'quora':'Q','reddit':'RE','trustpilot':'TR','yelp':'YL','g2':'G2','capterra':'CP',
    'amazon':'AM','amazon reviews':'AM','tripadvisor':'TA','glassdoor':'GL','indeed':'ID',
    'producthunt':'PH','zomato':'ZO','swiggy':'SG','booking':'BK','booking.com':'BK',
    'csv upload':'CS','csv':'CS','mysql':'MY','postgresql':'PG','snowflake':'SF',
    'mongodb':'MG','sql server':'MS','sqlite':'SQ','bigquery':'BQ',
    'google analytics':'GA','databricks':'DB',
}
MONTH_MAP = {
    'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,
    'oct':10,'nov':11,'dec':12,'january':1,'february':2,'march':3,'april':4,
    'june':6,'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
}

def get_prefix(src):
    sl = src.lower().strip()
    for k, p in SOURCE_PREFIX_MAP.items():
        if k in sl: return p
    c = re.sub(r'[^a-zA-Z]', '', src).upper()
    return c[:2] if len(c) >= 2 else 'FB'

def assign_ids(df):
    ctr = {}
    try:
        conn = sqlite3.connect(DB_PATH)
        ex = pd.read_sql_query("SELECT feedback_id FROM feedback_entries", conn); conn.close()
        for fid in ex['feedback_id'].dropna():
            m = re.match(r'^([A-Z]+)(\d+)$', str(fid).strip())
            if m:
                p, n = m.group(1), int(m.group(2))
                if ctr.get(p, 0) <= n: ctr[p] = n+1
    except: pass
    ids, local = [], {}
    for _, row in df.iterrows():
        p = get_prefix(row['Source'])
        if p not in local: local[p] = ctr.get(p, 1)
        n = local[p]
        ids.append(f"{p}{str(n).zfill(2) if n < 100 else n}")
        local[p] = n+1
    df = df.copy(); df.insert(0, 'Feedback_ID', ids)
    return df

def extract_exact_date(text):
    if not text or not isinstance(text, str): return ''
    m = re.search(r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})', text)
    if m:
        try:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if 1990 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
                return datetime(y, mo, d).strftime("%b %d, %Y")
        except: pass
    m = re.search(r'(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\.?\s+(\d{1,2}),?\s+(\d{4})', text, re.I)
    if m:
        try:
            mo = MONTH_MAP[m.group(1).lower()[:3]]; d, y = int(m.group(2)), int(m.group(3))
            return datetime(y, mo, d).strftime("%b %d, %Y")
        except: pass
    m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', text)
    if m:
        try:
            a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            mo, d = (a, b) if a <= 12 else (b, a)
            if 1 <= mo <= 12 and 1 <= d <= 31: return datetime(y, mo, d).strftime("%b %d, %Y")
        except: pass
    return ''

def _normalize_for_hash(text):
    if not text or not isinstance(text, str): return ''
    t = re.sub(r'http\S+|www\S+', '', text.lower())
    t = re.sub(r'[^\w\s]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()[:500]

def make_content_hash(fb_text):
    return hashlib.sha256(_normalize_for_hash(fb_text).encode('utf-8')).hexdigest()

def init_db():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS sessions(
        session_id TEXT PRIMARY KEY, user_name TEXT, short_uuid TEXT,
        created_date TEXT, created_time TEXT, created_at TEXT,
        source_type TEXT, total_entries INTEGER, notes TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS feedback_entries(
        id INTEGER PRIMARY KEY AUTOINCREMENT, session_id TEXT, entry_uuid TEXT UNIQUE,
        feedback_id TEXT, source TEXT, reviewer_name TEXT, feedback_date TEXT,
        feedback TEXT, topic TEXT, sentiment TEXT, suggestion TEXT, analyzed_at TEXT,
        content_hash TEXT,
        FOREIGN KEY(session_id) REFERENCES sessions(session_id))""")
    for col in ['feedback_id','reviewer_name','feedback_date','sentiment','content_hash']:
        try: c.execute(f"ALTER TABLE feedback_entries ADD COLUMN {col} TEXT")
        except: pass
    try: c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_content_hash ON feedback_entries(content_hash)")
    except: pass
    try:
        missing = conn.execute("SELECT id, feedback FROM feedback_entries WHERE content_hash IS NULL OR content_hash = ''").fetchall()
        for row_id, fb_text in missing:
            c.execute("UPDATE feedback_entries SET content_hash=? WHERE id=?", (make_content_hash(fb_text or ''), row_id))
        if missing: conn.commit()
    except: pass
    conn.commit(); conn.close()

_HASH_CACHE: set = set()
_HASH_CACHE_LOADED: bool = False

def _ensure_cache_loaded():
    global _HASH_CACHE, _HASH_CACHE_LOADED
    if _HASH_CACHE_LOADED: return
    try:
        conn = sqlite3.connect(DB_PATH)
        rows = conn.execute("SELECT content_hash FROM feedback_entries WHERE content_hash IS NOT NULL AND content_hash != ''").fetchall()
        conn.close()
        _HASH_CACHE.update(r[0] for r in rows)
    except: pass
    _HASH_CACHE_LOADED = True

def is_dup(fb_text):
    h = make_content_hash(fb_text)
    _ensure_cache_loaded()
    if h in _HASH_CACHE: return True
    try:
        conn = sqlite3.connect(DB_PATH)
        row = conn.execute("SELECT 1 FROM feedback_entries WHERE content_hash=? LIMIT 1", (h,)).fetchone()
        conn.close()
        if row: _HASH_CACHE.add(h); return True
    except: pass
    return False

def save_session(sid, uname, uid8, src, total, notes=""):
    now = datetime.now(); conn = sqlite3.connect(DB_PATH)
    conn.execute("INSERT OR REPLACE INTO sessions(session_id,user_name,short_uuid,created_date,created_time,created_at,source_type,total_entries,notes)VALUES(?,?,?,?,?,?,?,?,?)",
        (sid, uname, uid8, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), now.strftime("%Y-%m-%d %H:%M:%S"), src, total, notes))
    conn.commit(); conn.close()

def save_entries(session_id, df):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    at = datetime.now().strftime("%Y-%m-%d %H:%M:%S"); saved = 0; skipped = 0
    _ensure_cache_loaded()
    for _, row in df.iterrows():
        fb = row.get('Feedback', ''); h = make_content_hash(fb)
        if h in _HASH_CACHE: skipped += 1; continue
        try:
            c.execute("INSERT INTO feedback_entries(session_id,entry_uuid,feedback_id,source,reviewer_name,feedback_date,feedback,topic,sentiment,suggestion,analyzed_at,content_hash)VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (session_id, str(uuid.uuid4()), row.get('Feedback_ID',''), row.get('Source',''), row.get('Reviewer_Name',''),
                 row.get('Feedback_Date',''), fb, row.get('Topic',''), row.get('Sentiment','Neutral'),
                 row.get('Suggestion',''), at, h))
            _HASH_CACHE.add(h); saved += 1
        except sqlite3.IntegrityError: _HASH_CACHE.add(h); skipped += 1
    conn.commit(); conn.close()
    return saved, skipped

def get_all_sessions():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT session_id,user_name,created_date,created_time,source_type,total_entries,notes FROM sessions ORDER BY created_at DESC", conn)
    conn.close(); return df

def get_session_entries(sid):
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT feedback_id,source,reviewer_name,feedback_date,feedback,topic,sentiment,suggestion,analyzed_at FROM feedback_entries WHERE session_id=? ORDER BY id ASC", conn, params=(sid,))
    conn.close(); return df

def get_full_db_export():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("""SELECT f.feedback_id,f.source,f.reviewer_name,f.feedback_date,
        f.feedback,f.topic,f.sentiment,f.suggestion,f.analyzed_at,s.user_name
        FROM sessions s JOIN feedback_entries f ON s.session_id=f.session_id
        ORDER BY s.created_at DESC, f.id ASC""", conn)
    conn.close(); return df

def delete_session(sid):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM feedback_entries WHERE session_id=?", (sid,))
    conn.execute("DELETE FROM sessions WHERE session_id=?", (sid,))
    conn.commit(); conn.close()

init_db()

def gen_sid(uname):
    now = datetime.now(); safe = re.sub(r'[^a-zA-Z0-9]', '', uname.upper())[:12] or "USER"
    uid8 = str(uuid.uuid4()).replace('-', '')[:8].upper()
    return f"{safe}-{now.strftime('%Y%m%d')}-{now.strftime('%H%M%S')}-{uid8}", uid8

# ══════════════════════════════════════════════════════════════
# MULTI-DATABASE CONNECTORS
# ══════════════════════════════════════════════════════════════
def _fetch_sqlite(db_path, query):
    conn = sqlite3.connect(db_path)
    try: return pd.read_sql_query(query, conn)
    finally: conn.close()

def _list_sqlite_tables(db_path):
    conn = sqlite3.connect(db_path)
    tbls = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)["name"].tolist()
    conn.close(); return tbls

def _fetch_mysql(host, port, user, password, database, query):
    if not MYSQL_AVAILABLE: raise ImportError("pip install mysql-connector-python")
    conn = mysql.connector.connect(host=host, port=int(port), user=user, password=password, database=database)
    try: return pd.read_sql_query(query, conn)
    finally: conn.close()

def _list_mysql_tables(host, port, user, password, database):
    if not MYSQL_AVAILABLE: return []
    conn = mysql.connector.connect(host=host, port=int(port), user=user, password=password, database=database)
    cur = conn.cursor(); cur.execute("SHOW TABLES")
    tbls = [r[0] for r in cur.fetchall()]; conn.close(); return tbls

def _fetch_postgres(host, port, user, password, database, query):
    if not POSTGRES_AVAILABLE: raise ImportError("pip install psycopg2-binary")
    conn = psycopg2.connect(host=host, port=int(port), user=user, password=password, dbname=database)
    try: return pd.read_sql_query(query, conn)
    finally: conn.close()

def _list_postgres_tables(host, port, user, password, database):
    if not POSTGRES_AVAILABLE: return []
    conn = psycopg2.connect(host=host, port=int(port), user=user, password=password, dbname=database)
    df = pd.read_sql_query("SELECT table_name FROM information_schema.tables WHERE table_schema='public'", conn)
    conn.close(); return df['table_name'].tolist()

def _fetch_snowflake(account, user, password, warehouse, database, schema, query):
    if not SNOWFLAKE_AVAILABLE: raise ImportError("pip install snowflake-connector-python")
    conn = snowflake.connector.connect(account=account, user=user, password=password, warehouse=warehouse, database=database, schema=schema)
    try: return pd.read_sql_query(query, conn)
    finally: conn.close()

def _list_snowflake_tables(account, user, password, warehouse, database, schema):
    if not SNOWFLAKE_AVAILABLE: return []
    df = _fetch_snowflake(account, user, password, warehouse, database, schema, f"SHOW TABLES IN SCHEMA {database}.{schema}")
    return df["name"].tolist() if "name" in df.columns else []

def _sanitize_mongo_uri(uri):
    from urllib.parse import quote_plus
    if not uri or not uri.strip(): return uri
    uri = uri.strip()
    if MONGO_AVAILABLE:
        try:
            from pymongo import uri_parser
            uri_parser.parse_uri(uri); return uri
        except: pass
    if uri.startswith("mongodb+srv://"): scheme = "mongodb+srv"; rest = uri[len("mongodb+srv://"):]
    elif uri.startswith("mongodb://"): scheme = "mongodb"; rest = uri[len("mongodb://"):]
    else: return uri
    at_pos = next((i for i, ch in enumerate(rest) if ch == "@"), None)
    if at_pos is None: return uri
    creds_part = rest[:at_pos]; host_part = rest[at_pos+1:]
    if ":" in creds_part: username, password = creds_part.split(":", 1)
    else: username, password = creds_part, ""
    if "%" not in username: username = quote_plus(username)
    if "%" not in password: password = quote_plus(password)
    return f"{scheme}://{username}:{password}@{host_part}"

def _fetch_mongodb(uri, database, collection, fields, limit=0):
    if not MONGO_AVAILABLE: raise ImportError("pip install pymongo")
    client = pymongo.MongoClient(_sanitize_mongo_uri(uri))
    try:
        proj = {f: 1 for f in fields} if fields else {}
        proj["_id"] = 0
        cursor = client[database][collection].find({}, proj)
        if limit > 0: cursor = cursor.limit(limit)
        return pd.DataFrame(list(cursor))
    finally: client.close()

def _list_mongodb_collections(uri, database):
    if not MONGO_AVAILABLE: return []
    client = pymongo.MongoClient(_sanitize_mongo_uri(uri))
    try: return client[database].list_collection_names()
    finally: client.close()

def _fetch_mssql(host, port, user, password, database, query):
    if not MSSQL_AVAILABLE: raise ImportError("pip install pymssql")
    conn = pymssql.connect(server=host, port=int(port), user=user, password=password, database=database)
    try: return pd.read_sql_query(query, conn)
    finally: conn.close()

def _list_mssql_tables(host, port, user, password, database):
    if not MSSQL_AVAILABLE: return []
    df = _fetch_mssql(host, port, user, password, database,
        "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE'")
    return df["TABLE_NAME"].tolist() if "TABLE_NAME" in df.columns else []

def _fetch_bigquery(project, query, credentials_json=None):
    if not BIGQUERY_AVAILABLE: raise ImportError("pip install google-cloud-bigquery db-dtypes")
    if credentials_json:
        import google.oauth2.service_account as sa
        creds = sa.Credentials.from_service_account_info(json.loads(credentials_json))
        client = bq_lib.Client(project=project, credentials=creds)
    else: client = bq_lib.Client(project=project)
    return client.query(query).to_dataframe()

def _list_bigquery_datasets(project, credentials_json=None):
    if not BIGQUERY_AVAILABLE: return []
    if credentials_json:
        import google.oauth2.service_account as sa
        creds = sa.Credentials.from_service_account_info(json.loads(credentials_json))
        client = bq_lib.Client(project=project, credentials=creds)
    else: client = bq_lib.Client(project=project)
    return [d.dataset_id for d in client.list_datasets()]

def _fetch_ga4(property_id, credentials_json, start_date="30daysAgo", end_date="today"):
    if not GA_AVAILABLE: raise ImportError("pip install google-analytics-data")
    import google.oauth2.service_account as sa
    creds = sa.Credentials.from_service_account_info(
        json.loads(credentials_json),
        scopes=["https://www.googleapis.com/auth/analytics.readonly"])
    client = BetaAnalyticsDataClient(credentials=creds)
    req = RunReportRequest(
        property=f"properties/{property_id}",
        date_ranges=[DateRange(start_date=start_date, end_date=end_date)],
        dimensions=[Dimension(name="pageTitle"), Dimension(name="pagePath"), Dimension(name="country")],
        metrics=[Metric(name="sessions"), Metric(name="bounceRate"), Metric(name="averageSessionDuration")],
        limit=500)
    resp = client.run_report(req)
    rows = []
    for row in resp.rows:
        rows.append({d.name: row.dimension_values[i].value for i, d in enumerate(resp.dimension_headers)}
                  | {m.name: row.metric_values[j].value for j, m in enumerate(resp.metric_headers)})
    return pd.DataFrame(rows)

def _fetch_databricks(host, http_path, token, query):
    if not DATABRICKS_AVAILABLE: raise ImportError("pip install databricks-sql-connector")
    with dbsql.connect(server_hostname=host, http_path=http_path, access_token=token) as conn:
        with conn.cursor() as cur:
            cur.execute(query)
            cols = [d[0] for d in cur.description]
            return pd.DataFrame(cur.fetchall(), columns=cols)

def _list_databricks_tables(host, http_path, token, catalog="main", schema="default"):
    if not DATABRICKS_AVAILABLE: return []
    df = _fetch_databricks(host, http_path, token, f"SHOW TABLES IN {catalog}.{schema}")
    return df["tableName"].tolist() if "tableName" in df.columns else []

# ══════════════════════════════════════════════════════════════
# FETCH ALL ROWS  — unified fetcher (no LIMIT)
# ══════════════════════════════════════════════════════════════
def fetch_all_rows(db_type, conn_params, table_name, tmp_path=""):
    """Fetch ALL rows from any DB type, no limit."""
    query = f"SELECT * FROM {table_name}"

    if db_type == "SQLite (file)":
        return _fetch_sqlite(tmp_path, query)
    elif db_type == "MySQL / MariaDB":
        return _fetch_mysql(conn_params['host'], conn_params['port'],
                            conn_params['user'], conn_params['password'],
                            conn_params['database'], query)
    elif db_type == "PostgreSQL":
        return _fetch_postgres(conn_params['host'], conn_params['port'],
                               conn_params['user'], conn_params['password'],
                               conn_params['database'], query)
    elif db_type == "SQL Server (MSSQL)":
        return _fetch_mssql(conn_params['host'], conn_params['port'],
                            conn_params['user'], conn_params['password'],
                            conn_params['database'], query)
    elif db_type == "Snowflake":
        return _fetch_snowflake(conn_params['account'], conn_params['user'],
                                conn_params['password'], conn_params['warehouse'],
                                conn_params['database'], conn_params['schema'], query)
    elif db_type == "MongoDB":
        client = pymongo.MongoClient(_sanitize_mongo_uri(conn_params['uri']))
        try:
            docs = list(client[conn_params['database']][table_name].find({}, {"_id": 0}))
            return pd.DataFrame(docs)
        finally: client.close()
    elif db_type == "Google BigQuery":
        return _fetch_bigquery(conn_params['project'], query,
                               conn_params.get('credentials_json'))
    elif db_type == "Google Analytics 4":
        return _fetch_ga4(conn_params['property_id'],
                          conn_params.get('credentials_json', ''),
                          conn_params.get('start_date', '30daysAgo'),
                          conn_params.get('end_date', 'today'))
    elif db_type == "Databricks SQL":
        return _fetch_databricks(conn_params['host'], conn_params['http_path'],
                                 conn_params['token'], query)
    return pd.DataFrame()

# ══════════════════════════════════════════════════════════════
# METADATA EXTRACTOR  — column-level stats for any DB type
# ══════════════════════════════════════════════════════════════
def get_table_metadata(db_type, conn_params, table_name, tmp_path="", raw_df=None):
    """
    Returns a list of dicts:
    { col_name, data_type, null_count, null_pct, unique_count,
      avg_len, min_len, max_len, sample_values }
    Works from raw_df if already fetched, otherwise queries DB for schema first.
    """
    meta = []

    # Use already-fetched df if available
    if raw_df is not None and not raw_df.empty:
        total = len(raw_df)
        for col in raw_df.columns:
            series = raw_df[col]
            null_count  = int(series.isna().sum())
            null_pct    = round(null_count / total * 100, 1) if total else 0
            unique_count = int(series.nunique())
            is_text = series.dtype == object
            lengths = series.dropna().astype(str).str.len() if is_text else None
            avg_len  = round(float(lengths.mean()), 1)  if is_text and len(lengths) else 0
            min_len  = int(lengths.min())               if is_text and len(lengths) else 0
            max_len  = int(lengths.max())               if is_text and len(lengths) else 0
            samples  = series.dropna().astype(str).head(3).tolist()
            meta.append({
                "col_name":     col,
                "data_type":    str(series.dtype),
                "null_count":   null_count,
                "null_pct":     null_pct,
                "unique_count": unique_count,
                "avg_len":      avg_len,
                "min_len":      min_len,
                "max_len":      max_len,
                "sample_values": samples,
            })
        return meta

    # Fallback: schema-only query (before full fetch)
    try:
        if db_type == "SQLite (file)":
            conn = sqlite3.connect(tmp_path)
            info = pd.read_sql_query(f"PRAGMA table_info({table_name})", conn)
            row_count = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
            conn.close()
            for _, r in info.iterrows():
                meta.append({"col_name": r["name"], "data_type": r["type"],
                             "null_count": "?", "null_pct": "?",
                             "unique_count": "?", "avg_len": "?",
                             "min_len": "?", "max_len": "?",
                             "sample_values": [], "_row_count": row_count})

        elif db_type == "MySQL / MariaDB":
            info = _fetch_mysql(conn_params['host'], conn_params['port'],
                                conn_params['user'], conn_params['password'],
                                conn_params['database'], f"DESCRIBE {table_name}")
            for _, r in info.iterrows():
                meta.append({"col_name": r["Field"], "data_type": r["Type"],
                             "null_count":"?","null_pct":"?","unique_count":"?",
                             "avg_len":"?","min_len":"?","max_len":"?","sample_values":[]})

        elif db_type == "PostgreSQL":
            info = _fetch_postgres(conn_params['host'], conn_params['port'],
                                   conn_params['user'], conn_params['password'],
                                   conn_params['database'],
                                   f"SELECT column_name,data_type FROM information_schema.columns WHERE table_name='{table_name}'")
            for _, r in info.iterrows():
                meta.append({"col_name": r["column_name"], "data_type": r["data_type"],
                             "null_count":"?","null_pct":"?","unique_count":"?",
                             "avg_len":"?","min_len":"?","max_len":"?","sample_values":[]})

        elif db_type == "SQL Server (MSSQL)":
            info = _fetch_mssql(conn_params['host'], conn_params['port'],
                                conn_params['user'], conn_params['password'],
                                conn_params['database'],
                                f"SELECT COLUMN_NAME,DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='{table_name}'")
            for _, r in info.iterrows():
                meta.append({"col_name": r["COLUMN_NAME"], "data_type": r["DATA_TYPE"],
                             "null_count":"?","null_pct":"?","unique_count":"?",
                             "avg_len":"?","min_len":"?","max_len":"?","sample_values":[]})

        elif db_type == "Snowflake":
            info = _fetch_snowflake(conn_params['account'], conn_params['user'],
                                    conn_params['password'], conn_params['warehouse'],
                                    conn_params['database'], conn_params['schema'],
                                    f"DESCRIBE TABLE {table_name}")
            for _, r in info.iterrows():
                meta.append({"col_name": r.get("name","?"), "data_type": r.get("type","?"),
                             "null_count":"?","null_pct":"?","unique_count":"?",
                             "avg_len":"?","min_len":"?","max_len":"?","sample_values":[]})

        elif db_type == "Databricks SQL":
            info = _fetch_databricks(conn_params['host'], conn_params['http_path'],
                                     conn_params['token'], f"DESCRIBE {table_name}")
            for _, r in info.iterrows():
                meta.append({"col_name": r.get("col_name","?"), "data_type": r.get("data_type","?"),
                             "null_count":"?","null_pct":"?","unique_count":"?",
                             "avg_len":"?","min_len":"?","max_len":"?","sample_values":[]})

        elif db_type == "MongoDB":
            # Sample 1 doc to infer field types
            client = pymongo.MongoClient(_sanitize_mongo_uri(conn_params['uri']))
            try:
                sample = client[conn_params['database']][table_name].find_one({}, {"_id": 0})
                if sample:
                    for k, v in sample.items():
                        meta.append({"col_name": k, "data_type": type(v).__name__,
                                     "null_count":"?","null_pct":"?","unique_count":"?",
                                     "avg_len":"?","min_len":"?","max_len":"?","sample_values":[str(v)]})
            finally: client.close()

    except Exception as ex:
        meta = [{"col_name": f"Error: {ex}", "data_type":"","null_count":"","null_pct":"",
                 "unique_count":"","avg_len":"","min_len":"","max_len":"","sample_values":[]}]

    return meta


def render_metadata_panel(meta, detected_fb_col=""):
    """Render a styled metadata table with column-level stats."""
    if not meta:
        st.info("No metadata available.")
        return

    st.markdown(
        '<div style="font-family:Space Mono,monospace;font-size:0.7rem;color:#888;'
        'letter-spacing:2px;text-transform:uppercase;margin-bottom:0.6rem;">'
        '🗂️ Table Metadata — Column Inspector</div>',
        unsafe_allow_html=True
    )

    header_html = (
        '<div class="meta-row" style="border-bottom:2px solid rgba(249,168,37,0.4);'
        'padding-bottom:6px;margin-bottom:4px;">'
        '<span class="meta-col-name">Column</span>'
        '<span class="meta-col-type">Type</span>'
        '<span class="meta-col-null">Nulls</span>'
        '<span style="min-width:70px;font-family:Space Mono,monospace;font-size:0.68rem;color:#f9a825;">Unique</span>'
        '<span style="min-width:80px;font-family:Space Mono,monospace;font-size:0.68rem;color:#f9a825;">Avg Len</span>'
        '<span class="meta-col-sample">Sample Values</span>'
        '</div>'
    )

    rows_html = ""
    for m in meta:
        is_fb = m["col_name"] == detected_fb_col
        highlight = "background:rgba(249,168,37,0.07);border-radius:4px;" if is_fb else ""
        fb_badge  = ' <span style="background:#f9a825;color:#000;border-radius:3px;padding:0 5px;font-size:0.6rem;font-weight:700;">FEEDBACK</span>' if is_fb else ""
        null_pct  = f"{m['null_pct']}%" if isinstance(m['null_pct'], (int, float)) else str(m['null_pct'])
        avg_len   = f"{m['avg_len']}" if m['avg_len'] != "?" else "?"
        samples   = " · ".join([str(s)[:40] for s in m['sample_values'][:2]]) if m['sample_values'] else "—"

        rows_html += (
            f'<div class="meta-row" style="{highlight}">'
            f'<span class="meta-col-name">{m["col_name"]}{fb_badge}</span>'
            f'<span class="meta-col-type">{m["data_type"]}</span>'
            f'<span class="meta-col-null">{null_pct}</span>'
            f'<span style="min-width:70px;font-family:Space Mono,monospace;font-size:0.68rem;color:#aaa;">{m["unique_count"]}</span>'
            f'<span style="min-width:80px;font-family:Space Mono,monospace;font-size:0.68rem;color:#aaa;">{avg_len}</span>'
            f'<span class="meta-col-sample">{samples}</span>'
            f'</div>'
        )

    st.markdown(
        f'<div style="background:rgba(0,0,0,0.2);border:1px solid rgba(255,255,255,0.07);'
        f'border-radius:10px;padding:0.8rem 1rem;">'
        f'{header_html}{rows_html}</div>',
        unsafe_allow_html=True
    )


# ══════════════════════════════════════════════════════════════
# SMART COLUMN DETECTION  — 4-layer system
# ══════════════════════════════════════════════════════════════
def smart_detect_columns(df):
    cols_lower = {c.lower().strip(): c for c in df.columns}

    def match_exact(name_list):
        return next((cols_lower[n] for n in name_list if n in cols_lower), None)

    def match_partial(keyword_list):
        return next((orig for lower, orig in cols_lower.items()
                     if any(kw in lower for kw in keyword_list)), None)

    def match_longest_text():
        skip_re = re.compile(
            r'\b(id|uuid|key|hash|token|flag|count|score|rating|status|type|'
            r'category|label|tag|code|email|phone|url|link|image|photo|avatar|ip)\b', re.I)
        text_cols = [c for c in df.select_dtypes(include='object').columns
                     if not skip_re.search(c)]
        if not text_cols: return None
        return max(text_cols,
                   key=lambda c: df[c].dropna().astype(str).str.len().mean())

    # ── Feedback column ───────────────────────────────────────
    layer_used = ""
    fb_col = match_exact(FEEDBACK_COL_EXACT)
    if fb_col: layer_used = "Layer 1 — Exact name match"

    if not fb_col:
        fb_col = match_partial(FEEDBACK_COL_PARTIAL)
        if fb_col: layer_used = "Layer 2 — Partial name match"

    if not fb_col:
        fb_col = match_longest_text()
        if fb_col: layer_used = "Layer 3 — Longest text column (data-driven)"

    if not fb_col:
        try:
            ai_res = ai_detect_feedback_column(df)
            fb_col = ai_res.get("feedback_col")
            layer_used = "Layer 4 — AI fallback"
        except: pass

    # ── Name column ───────────────────────────────────────────
    nm_col = match_exact(NAME_COL_EXACT)
    if not nm_col:
        nm_col = match_partial(["name", "author", "reviewer", "user", "customer"])

    # ── Date column ───────────────────────────────────────────
    dt_col = match_exact(DATE_COL_EXACT)
    if not dt_col:
        dt_col = match_partial(["date", "time", "created", "posted", "submitted"])

    # ── Per-column confidence scores ──────────────────────────
    col_scores = {}
    for lower, orig in cols_lower.items():
        if orig == fb_col:
            score = 95
        elif lower in FEEDBACK_COL_EXACT:
            score = 80
        elif any(kw in lower for kw in FEEDBACK_COL_PARTIAL):
            score = 50
        else:
            avg = df[orig].dropna().astype(str).str.len().mean() if df[orig].dtype == object else 0
            score = min(40, int(avg / 10))
        col_scores[orig] = score

    return {
        "feedback_col":     fb_col,
        "name_col":         nm_col,
        "date_col":         dt_col,
        "layer_used":       layer_used,
        "confidence_score": col_scores.get(fb_col, 0) if fb_col else 0,
        "column_scores":    col_scores,
        "reason":           f"Detected via {layer_used}" if layer_used else "No feedback column found",
    }


# ── Legacy AI detect (kept as Layer 4 fallback) ──────────────
_STRICT_DETECT_SYSTEM = (
    "You are a strict and highly accurate data schema analyst. "
    "Your job is to analyze a dataset schema and identify which column contains CUSTOMER FEEDBACK. "
    "Return ONLY valid JSON — no markdown, no explanation."
)
_FEEDBACK_DEFINITION = """\
A feedback column contains human-written natural language opinions, complaints, reviews, suggestions.
NOT feedback: IDs, ratings, status fields, emails, URLs, categories.
OUTPUT FORMAT (return ONLY this JSON):
{"feedback_col":"<col or null>","name_col":"<col or null>","date_col":"<col or null>",
 "confidence_score":<0-100>,"reason":"<one sentence>","column_scores":{"<col>":<0-100>}}"""

def ai_detect_feedback_column(df, source_label=""):
    hfb, hnm, hdt = _heuristic_detect(df)
    if not openai.api_key:
        return {"feedback_col": hfb, "name_col": hnm, "date_col": hdt,
                "confidence_score": 50, "reason": "No API key.", "column_scores": {}}
    schema_lines = []
    for col in df.columns:
        sample  = df[col].dropna().astype(str).head(3).tolist()
        avg_len = int(df[col].dropna().astype(str).str.len().mean()) if len(df[col].dropna()) else 0
        schema_lines.append(f'  - "{col}" | dtype:{df[col].dtype} | avg_len:{avg_len} | samples:{sample}')
    try:
        resp = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[{"role":"system","content":_STRICT_DETECT_SYSTEM},
                      {"role":"user","content":f"{_FEEDBACK_DEFINITION}\n\n" + "\n".join(schema_lines)}],
            max_tokens=500, temperature=0.0)
        raw = re.sub(r"^```(?:json)?|```$", "", resp["choices"][0]["message"]["content"].strip(), flags=re.M).strip()
        result = json.loads(raw)
        result.setdefault("feedback_col", hfb)
        result.setdefault("name_col", hnm)
        result.setdefault("date_col", hdt)
        result.setdefault("confidence_score", 0)
        result.setdefault("reason", "")
        result.setdefault("column_scores", {})
        return result
    except:
        return {"feedback_col": hfb, "name_col": hnm, "date_col": hdt,
                "confidence_score": 50, "reason": "AI error; heuristic used.", "column_scores": {}}

def _heuristic_detect(df):
    text_cols  = [c for c in df.columns if df[c].dtype == object]
    skip_re    = re.compile(r'\b(id|uuid|key|hash|token|flag|count|score|rating|status|type|category|label|tag|code|email|phone|url|link|image|photo|avatar)\b', re.I)
    candidates = [c for c in text_cols if not skip_re.search(c)]
    fb_col     = next((c for c in candidates if any(k in c.lower() for k in ['feedback','review','comment','text','body','description','message','opinion','remark','note'])), None)
    if not fb_col and candidates:
        fb_col = max(candidates, key=lambda c: df[c].dropna().astype(str).str.len().mean())
    if not fb_col and text_cols: fb_col = text_cols[0]
    nm_col = next((c for c in df.columns if any(k in c.lower() for k in ['name','reviewer','author','user','customer','username'])), None)
    dt_col = next((c for c in df.columns if any(k in c.lower() for k in ['date','time','created','posted','timestamp','when','at'])), None)
    return fb_col, nm_col, dt_col


# ══════════════════════════════════════════════════════════════
# NLP
# ══════════════════════════════════════════════════════════════
def preprocess(text):
    if not text or not isinstance(text, str): return ''
    text = text.lower(); text = re.sub(r'http\S+|www\S+', '', text)
    text = re.sub(r'[^a-z0-9\s]', '', text)
    try: words = word_tokenize(text)
    except: words = text.split()
    sw = set(stopwords.words('english')); words = [w for w in words if w not in sw and len(w) > 2]
    try: lem = WordNetLemmatizer(); words = [lem.lemmatize(w) for w in words]
    except: pass
    return ' '.join(words) if len(words) >= 2 else ''

def topic_model(texts, n=5):
    if len(texts) < 2: return [0]*len(texts)
    n = min(n, len(texts))
    try:
        vec = TfidfVectorizer(stop_words='english', max_features=500, min_df=1)
        X = vec.fit_transform(texts)
        W = NMF(n_components=n, random_state=42, max_iter=300).fit_transform(X)
        return W.argmax(axis=1).tolist()
    except: return [0]*len(texts)

# ══════════════════════════════════════════════════════════════
# AI PROMPTS
# ══════════════════════════════════════════════════════════════
URL_SYSTEM = """You are a CCX Officer analyzing real-time scraped reviews.
For each review: 1) Extract reviewer name (or Anonymous) 2) Write ONE precise operational recommendation.
Rules: Reference what this reviewer experienced. Name the team/process. Max 2 sentences. Vary action verbs."""

CSV_SYSTEM = """You are a senior CX analyst. For each customer feedback, write ONE specific operational action recommendation.
Max 2 sentences. Be concrete — name the team, process, or metric. Vary language across items."""

def get_suggestions_openai(feedback_list):
    if not openai.api_key: return [""] * len(feedback_list)
    try:
        prompt = "For each feedback, prescribe ONE specific operational action. Max 2 sentences. Be unique per item.\n\nFeedbacks:\n"
        for i, fb in enumerate(feedback_list, 1): prompt += f"{i}. {str(fb)[:400]}\n"
        resp = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[{"role":"system","content":CSV_SYSTEM},{"role":"user","content":prompt}],
            max_tokens=3000, temperature=0.85)
        raw = resp["choices"][0]["message"]["content"].strip()
        segments = re.split(r'(?m)^(\d+)[\.):]\s+', raw)
        suggestions_map = {}
        i = 1
        while i < len(segments) - 1:
            try:
                num = int(segments[i]); text = re.sub(r'\s*\n\s*', ' ', segments[i+1].strip()).strip()
                if len(text) > 10: suggestions_map[num] = text
            except: pass
            i += 2
        return [suggestions_map.get(idx, "") for idx in range(1, len(feedback_list)+1)]
    except Exception as e:
        st.warning(f"⚠️ AI suggestion error: {e}"); return [""] * len(feedback_list)

def get_ai_names_and_suggestions(feedback_list, existing_names=None):
    if existing_names is None: existing_names = [''] * len(feedback_list)
    try:
        prompt = "For each review, extract reviewer name and write ONE operational action.\nFormat: <n>. [NAME: Full Name] Recommendation.\n\nReviews:\n"
        for i, (fb, nm) in enumerate(zip(feedback_list, existing_names), 1):
            hint = f" [Reviewer: {nm}]" if nm and nm.strip() else ""
            prompt += f"{i}. {fb}{hint}\n"
        resp = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[{"role":"system","content":URL_SYSTEM},{"role":"user","content":prompt}],
            max_tokens=2500, temperature=0.85)
        text = resp["choices"][0]["message"]["content"]
        parsed = {}
        for line in text.strip().split('\n'):
            line = line.strip()
            if not line: continue
            num_m = re.match(r'^(\d+)[\.\)]\s*', line)
            if not num_m: continue
            idx = int(num_m.group(1)) - 1; line = line[num_m.end():]
            name = ""; nm_m = re.search(r'\[NAME:\s*([^\]]+)\]', line, re.I)
            if nm_m:
                name = nm_m.group(1).strip()
                if name.lower() in ("anonymous","unknown","n/a"): name = ""
                line = line.replace(nm_m.group(0), "").strip()
            if idx not in parsed: parsed[idx] = (name, line.strip())
        return [parsed.get(i, ("", "")) for i in range(len(feedback_list))]
    except: return [("", "")] * len(feedback_list)

def run_ai_analysis(fb_df, mode='full'):
    ai_names, sugs, sents = [], [], []
    existing = [fb_df['Reviewer_Name'].iloc[i] if 'Reviewer_Name' in fb_df.columns and i < len(fb_df) else '' for i in range(len(fb_df))]
    bs = 10; prog = st.progress(0); stat = st.empty(); tb = max(1, (len(fb_df)+bs-1)//bs)
    if openai.api_key:
        for i in range(0, len(fb_df), bs):
            batch = fb_df['Feedback'].iloc[i:i+bs].tolist(); bn = i//bs+1
            stat.markdown(f'<p class="pulse" style="color:#f9a825;font-family:Space Mono,monospace;font-size:0.8rem;">🤖 Processing batch {bn}/{tb}...</p>', unsafe_allow_html=True)
            if mode == 'full':
                for nm, sg in get_ai_names_and_suggestions(batch, existing[i:i+bs]):
                    ai_names.append(nm); sugs.append(sg)
            else:
                for sg in get_suggestions_openai(batch):
                    ai_names.append(''); sugs.append(sg)
            prog.progress(min(0.7, (i+bs)/len(fb_df))); time.sleep(0.3)
    else:
        ai_names = [''] * len(fb_df); sugs = [''] * len(fb_df)
    stat.markdown('<p class="pulse" style="color:#00bcd4;font-family:Space Mono,monospace;font-size:0.8rem;">🎯 Analyzing sentiment...</p>', unsafe_allow_html=True)
    sents = analyze_sentiments_all(fb_df['Feedback'].tolist())
    prog.progress(1.0); stat.empty(); prog.empty()
    return ai_names, sugs, sents

# ══════════════════════════════════════════════════════════════
# JUNK DETECTION
# ══════════════════════════════════════════════════════════════
_JUNK_RE = re.compile(
    r'^(upvote|share\s|log\s?in|sign\s?in|sign\s?up|related|sponsored'
    r'|privacy\s+policy|terms\s+of\s+service|\u00a9|copyright|page\s+not\s+found'
    r'|404|security\s+service|please\s+enable\s+javascript|cloudflare|access\s+denied'
    r'|you\s+have\s+been\s+blocked|cookies?\s+policy|all\s+rights\s+reserved'
    r'|follow\s+us\s|subscribe\s|newsletter|download\s+the\s+app|write\s+a\s+review'
    r'|add\s+a\s+review|loading\.\.\.|please\s+wait|redirecting)', re.I)
_JUNK_PATTERNS = [
    re.compile(r'privacy\s*[\u00b7·]\s*terms', re.I),
    re.compile(r'©\s*\w+.*\d{4}', re.I),
    re.compile(r'security\s+service\s+to\s+protect', re.I),
    re.compile(r'enable\s+javascript', re.I),
    re.compile(r'you\s+have\s+been\s+blocked', re.I),
    re.compile(r'^[\s\u00b7·\-\|,]+$'),
]

def is_junk(text):
    if not text: return True
    t = text.strip().lower()
    if len(t) < 50: return True
    if _JUNK_RE.search(t): return True
    for pat in _JUNK_PATTERNS:
        if pat.search(text): return True
    alpha = sum(1 for c in t if c.isalpha())
    return len(t) > 0 and alpha / len(t) < 0.4

_TRUNC_RE = re.compile(
    r'(\.{2,}|…|\u2026)\s*$|\(more\)\s*$|\[more\]\s*$|'
    r'(?:read more|see more|show more|view more|continue reading|show full review)\s*$', re.I)

def is_truncated(text):
    if not text: return False
    return bool(_TRUNC_RE.search(text.strip()))

def clean_text(text):
    if not text or not isinstance(text, str): return ''
    text = re.sub(
        r'\s*[\u2026\.]{2,}\s*$|\s*\(more[\.\u2026]?\)\s*$|\s*\[more\]\s*$|'
        r'\s*(?:read|see|show|view)\s+more\s*$|\s*continue\s+reading\s*$|\s*show\s+full\s+review\s*$',
        '', text.strip(), flags=re.I).strip()
    text = re.sub(r'[\.\u2026]+$', '', text).strip()
    return re.sub(r'\s+', ' ', text).strip()

# ══════════════════════════════════════════════════════════════
# EXPORTS
# ══════════════════════════════════════════════════════════════
def make_summary_txt(df, session_id, user_name):
    tc = df['Topic'].value_counts(); total = len(df)
    lines = ["╔═══════════════════════════════════════════════╗",
             "║      ZEUS FEEDBACK ANALYZER  —  REPORT        ║",
             "╚═══════════════════════════════════════════════╝",
             f"  Session ID   : {session_id}", f"  Analyst      : {user_name}",
             f"  Generated    : {datetime.now().strftime('%d %B %Y, %H:%M:%S')}",
             f"  Total Entries: {total}",
             "─────────────────────────────────────────────────","","TOPIC DISTRIBUTION","──────────────────"]
    for topic, count in tc.items():
        bar = "█" * int(count/total*30)
        lines.append(f"  {topic:<35} {count:>4} ({round(count/total*100):>2}%)  {bar}")
    if 'Sentiment' in df.columns:
        sc = df['Sentiment'].value_counts()
        lines += ["","SENTIMENT BREAKDOWN","───────────────────"]
        for sent, cnt in sc.items():
            lines.append(f"  {sent:<20} {cnt:>4} ({round(cnt/total*100):>2}%)")
    lines += ["","ALL ENTRIES","───────────"]
    for i, row in df.iterrows():
        fid = row.get('Feedback_ID', f"#{i+1}"); name = row.get('Reviewer_Name','')
        date = row.get('Feedback_Date',''); sent = row.get('Sentiment','')
        lines += [f"\n  [{fid}] [{row['Source']}]  {row['Topic']}  [{sent}]",
                  f"  Reviewer   : {name if name else 'Anonymous'}",
                  f"  Date       : {date if date else 'Unknown'}",
                  f"  Feedback   : {row['Feedback']}",
                  f"  Suggestion : {row['Suggestion']}", "  "+"·"*65]
    return '\n'.join(lines)

def make_feedback_excel(df, session_id, user_name):
    try:
        from openpyxl import Workbook as WB
        from openpyxl.styles import PatternFill as PF, Font as FT, Alignment as AL, Border as BD, Side as SD
        from openpyxl.utils import get_column_letter as gcl
    except ImportError: return None
    DARK='1A1A2E'; GOLD='F9A825'; WHITE='FFFFFF'; LGT='F5F5F5'; GRY='E0E0E0'; DGR='666666'
    TC={'⚡ Service Quality':('F9A825','1A1A2E'),'📦 Product Issues':('E53935','FFFFFF'),
        '💬 Communication':('0288D1','FFFFFF'),'⏱️ Speed & Delays':('F57C00','FFFFFF'),
        '💡 Innovation & Features':('7B1FA2','FFFFFF')}
    SC_EXCEL={'Positive':('2E7D32','FFFFFF'),'Negative':('C62828','FFFFFF'),'Neutral':('455A64','FFFFFF'),'Mixed':('E65100','FFFFFF')}
    def ft(sz=10,b=False,co='111111'): return FT(name='Arial',size=sz,bold=b,color=co)
    def fl(c): return PF('solid',fgColor=c)
    def bd():
        s = SD(style='thin',color=GRY); return BD(left=s,right=s,top=s,bottom=s)
    def al(h='left',v='center',w=False): return AL(horizontal=h,vertical=v,wrap_text=w)
    def gc(cols): return next((c for c in cols if c in df.columns),'')
    idc=gc(['Feedback_ID','feedback_id']); rvc=gc(['Reviewer_Name','reviewer_name'])
    dtc=gc(['Feedback_Date','feedback_date']); fbc=gc(['Feedback','feedback'])
    tpc=gc(['Topic','topic']); sgc=gc(['Suggestion','suggestion']); stc=gc(['Sentiment','sentiment'])
    src=gc(['Source','source'])
    df2 = df.copy()
    src_label = ', '.join(df2[src].unique()[:3]) if src else 'N/A'
    wb = WB(); ws1 = wb.active; ws1.title = 'Feedback Data'; ws1.sheet_view.showGridLines = False
    ws1.merge_cells(start_row=1,start_column=1,end_row=1,end_column=8)
    c = ws1.cell(row=1,column=1,value=f'  ⚡ ZEUS — {src_label}  |  {len(df2)} Entries  |  {datetime.now().strftime("%d %b %Y")}')
    c.font=ft(13,True,GOLD); c.fill=fl(DARK); c.alignment=al('left','center'); ws1.row_dimensions[1].height=32
    HDRS = ['S.No','ID','Name of Feedbacker','Date','Feedback','Type','Sentiment','AI Suggestion']
    for i,w in enumerate([6,9,20,13,55,22,14,55],1): ws1.column_dimensions[gcl(i)].width = w
    for ci,lbl in enumerate(HDRS,1):
        c = ws1.cell(row=4,column=ci,value=lbl)
        c.font=ft(10,True,WHITE); c.fill=fl(DARK); c.border=bd(); c.alignment=al('center','center')
    ws1.row_dimensions[4].height = 22
    for ri,(_,row) in enumerate(df2.iterrows(),start=5):
        alt=(ri%2==0); rbg=LGT if alt else WHITE
        tp=str(row[tpc] if tpc else ''); tbg,tfg=TC.get(tp,('888888','FFFFFF'))
        sent=str(row[stc] if stc else 'Neutral'); sbg,sfg=SC_EXCEL.get(sent,('455A64','FFFFFF'))
        sug=str(row[sgc] if sgc else '')
        vals=[ri-4,row[idc] if idc else '',row[rvc] if rvc else '',row[dtc] if dtc else '',row[fbc] if fbc else '',tp,sent,sug]
        styles=[(rbg,DGR,9,False,'center'),(rbg,'0D47A1',9,True,'center'),(rbg,'1A237E',9,True,'center'),
                (rbg,DGR,9,False,'center'),(rbg,'212121',9,False,'left'),(tbg,tfg,8,True,'center'),
                (sbg,sfg,8,True,'center'),('FFFDE7','4E342E',9,False,'left')]
        for ci,(v,(bg,fg,sz,b,h)) in enumerate(zip(vals,styles),1):
            c = ws1.cell(row=ri,column=ci,value=str(v) if v is not None else '')
            c.fill=fl(bg); c.font=ft(sz,b,fg); c.border=bd(); c.alignment=al(h,'top' if h=='left' else 'center',w=(h=='left'))
        fb_len = len(str(row[fbc])) if fbc and row[fbc] else 0
        ws1.row_dimensions[ri].height = max(45, min(409, 45 + (fb_len//80)*14))
    ws1.freeze_panes = 'A5'; ws1.auto_filter.ref = f'A4:{gcl(8)}4'
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ══════════════════════════════════════════════════════════════
# BUILD RESULTS
# ══════════════════════════════════════════════════════════════
def build_results(fbs, src, names=None, dates=None, mode='full'):
    if names is None: names = [''] * len(fbs)
    if dates is None: dates = [''] * len(fbs)
    seen_hashes = set(); valid = []
    for i, (fb, nm, dt) in enumerate(zip(fbs, names, dates)):
        if not fb or not str(fb).strip(): continue
        fb_str = str(fb).strip()
        if is_junk(fb_str): continue
        h = make_content_hash(fb_str)
        if h in seen_hashes: continue
        seen_hashes.add(h); valid.append((i, fb_str, nm, dt))
    if not valid: return pd.DataFrame()
    positions, fbs2, names2, dates2 = zip(*valid)
    df = pd.DataFrame({'Feedback': list(fbs2), 'Source': src,
                       'Reviewer_Name': list(names2), 'Feedback_Date': list(dates2),
                       '_orig_pos': list(positions)})
    df['_c'] = df['Feedback'].apply(preprocess)
    df = df[df['_c'].str.strip() != ''].reset_index(drop=True)
    if df.empty: return pd.DataFrame()
    df['TopicID'] = topic_model(df['_c'].tolist())
    df['Topic']   = df['TopicID'].map(TOPIC_LABELS)
    ai_names, sugs, sents = run_ai_analysis(df, mode=mode)
    final_names = []
    for scraped, ai_nm in zip(df['Reviewer_Name'].tolist(), ai_names):
        if scraped and str(scraped).strip() and scraped not in ('nan','None',''):
            final_names.append(str(scraped).strip())
        elif ai_nm and str(ai_nm).strip():
            final_names.append(str(ai_nm).strip())
        else:
            final_names.append('')
    df['Reviewer_Name'] = final_names
    df['Suggestion'] = [
        s.strip() if s and isinstance(s, str) and len(s.strip()) > 20
        else generate_fallback_suggestion(f, t)
        for s, t, f in zip(sugs, df['TopicID'].tolist(), df['Feedback'].tolist())
    ]
    df['Sentiment'] = sents
    df = df.sort_values('_orig_pos').reset_index(drop=True)
    return assign_ids(df[['Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion']])

# ══════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════
for k, v in [
    ('results_df', None), ('analyzed', False), ('session_id', None),
    ('short_uuid', None), ('user_name', ''), ('source_type', ''),
    ('nav_mode', 'CSV'),
    ('db_step', 1), ('db_connected', False), ('db_type', ''),
    ('db_conn_params', {}), ('db_raw_df', None),
    ('db_available_tables', []), ('db_selected_table', ''),
    ('db_src_label', ''), ('db_tmp_path', ''),
    ('db_detected_cols', {}), ('db_metadata', []),
]:
    if k not in st.session_state:
        st.session_state[k] = v

# ══════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════
NAV_ITEMS = [("CSV","📄","CSV Upload"),("DB","🗄️","DB Connect"),("URL","🔗","URL Scraper")]

with st.sidebar:
    st.markdown('<div style="font-family:Space Mono,monospace;font-size:0.65rem;letter-spacing:3px;color:#f9a825;text-transform:uppercase;margin-bottom:0.8rem;">⚙ Configuration</div>', unsafe_allow_html=True)
    uname = st.text_input("👤 Your Name", value=st.session_state.user_name, placeholder="e.g. Priya")
    if uname: st.session_state.user_name = uname
    if st.session_state.session_id:
        st.markdown(f'<div class="session-box">🔑 Active Session<br><strong style="font-size:0.65rem;">{st.session_state.session_id}</strong></div>', unsafe_allow_html=True)
    st.divider()
    st.markdown('<div style="font-family:Space Mono,monospace;font-size:0.65rem;letter-spacing:3px;color:#f9a825;text-transform:uppercase;margin-bottom:0.6rem;">📍 Data Source</div>', unsafe_allow_html=True)
    for key, icon, label in NAV_ITEMS:
        is_active = st.session_state.nav_mode == key
        css_class = "nav-btn-active" if is_active else "nav-btn-inactive"
        st.markdown(f'<div class="{css_class}">', unsafe_allow_html=True)
        if st.button(f"{icon}  {label}", key=f"nav_{key}", use_container_width=True):
            st.session_state.nav_mode = key; st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
    st.divider()
    if openai.api_key: st.success("✅ OpenAI key loaded")
    else: st.warning("⚠️ No OPENAI_API_KEY — rule-based mode")
    if SELENIUM_AVAILABLE: st.success("✅ Selenium ready")
    else: st.info("Selenium not installed")
    st.divider()
    st.markdown("**🗄️ DB Drivers**")
    for dbname, (pkg, avail) in DB_DRIVERS.items():
        short = dbname.split()[0]
        badge = (f'<span class="db-badge-avail">✓ {short}</span>'
                 if avail else f'<span class="db-badge-na">✗ {short}</span>')
        st.markdown(badge, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# HERO
# ══════════════════════════════════════════════════════════════
st.markdown('<div class="hero-title">⚡ ZEUS FEEDBACK ANALYZER</div>', unsafe_allow_html=True)
st.markdown('<div class="hero-sub">Customer Intelligence · Topic Modeling · Real Sentiment · Multi-DB · Smart Column Detection</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# SHARED RENDERERS
# ══════════════════════════════════════════════════════════════
def render_results_tab(df, sid, uname_val):
    total = len(df); tc_data = df['Topic'].value_counts()
    st.markdown(f'<div class="session-box">🔑 <strong>{sid}</strong> &nbsp;·&nbsp; 👤 {uname_val} &nbsp;·&nbsp; 📅 {datetime.now().strftime("%d %b %Y, %H:%M")} &nbsp;·&nbsp; 📊 {total} entries</div>', unsafe_allow_html=True)
    c1,c2,c3,c4,c5,c6 = st.columns(6)
    c1.metric("Total Entries", f"{total:,}"); c2.metric("Topics", df['Topic'].nunique())
    c3.metric("Sources", df['Source'].nunique())
    named = df['Reviewer_Name'].astype(str).str.strip().ne('').sum() if 'Reviewer_Name' in df.columns else 0
    c4.metric("Named Reviewers", named)
    pos = df['Sentiment'].eq('Positive').sum() if 'Sentiment' in df.columns else 0
    neg = df['Sentiment'].eq('Negative').sum() if 'Sentiment' in df.columns else 0
    c5.metric("😊 Positive", pos); c6.metric("😞 Negative", neg)
    if 'Sentiment' in df.columns:
        st.divider(); st.markdown("### 🎯 Sentiment Distribution")
        sc = df['Sentiment'].value_counts(); sent_cols = st.columns(len(sc))
        SENT_S = {'Positive':('#4caf50','😊'),'Negative':('#e91e63','😞'),'Neutral':('#9e9e9e','😐'),'Mixed':('#ff9800','🤔')}
        for i,(sent,cnt) in enumerate(sc.items()):
            col_h, icon = SENT_S.get(sent,('#888','·'))
            with sent_cols[i]:
                st.markdown(f'<div style="background:rgba(255,255,255,0.04);border:1px solid {col_h}40;border-radius:12px;padding:1rem;text-align:center;"><div style="font-size:1.6rem;">{icon}</div><div style="font-size:1.4rem;font-weight:800;color:{col_h};font-family:Space Mono,monospace;">{cnt}</div><div style="font-size:0.72rem;color:#aaa;margin:2px 0;">{sent}</div><div style="font-size:0.75rem;color:#555;">{round(cnt/total*100)}%</div></div>', unsafe_allow_html=True)
    st.divider(); st.markdown("### 🗂️ Topic Distribution")
    cols = st.columns(min(len(tc_data),5))
    for i,(topic,count) in enumerate(tc_data.items()):
        with cols[i%5]:
            st.markdown(f'<div style="background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:1rem;text-align:center;"><div style="font-size:1.6rem;font-weight:800;color:#f9a825;font-family:Space Mono,monospace;">{count}</div><div style="font-size:0.68rem;color:#aaa;margin:4px 0;">{topic}</div><div style="font-size:0.75rem;color:#666;">{round(count/total*100)}%</div></div>', unsafe_allow_html=True)
    st.divider(); st.markdown("### 🔍 Browse & Filter")
    cf1,cf2,cf3,cf4 = st.columns(4)
    with cf1: tf = st.multiselect("Topic", df['Topic'].unique().tolist(), default=df['Topic'].unique().tolist(), key=f"tf_{sid[:8]}")
    with cf2: sf = st.multiselect("Source", df['Source'].unique().tolist(), default=df['Source'].unique().tolist(), key=f"sf_{sid[:8]}")
    with cf3:
        sent_opts = ['All'] + sorted(df['Sentiment'].unique().tolist()) if 'Sentiment' in df.columns else ['All']
        sent_filter = st.selectbox("Sentiment", sent_opts, key=f"sent_{sid[:8]}")
    with cf4: search_q = st.text_input("🔎 Search", placeholder="keyword...", key=f"sq_{sid[:8]}")
    filtered = df[df['Topic'].isin(tf) & df['Source'].isin(sf)]
    if sent_filter != 'All' and 'Sentiment' in filtered.columns:
        filtered = filtered[filtered['Sentiment']==sent_filter]
    if search_q.strip():
        filtered = filtered[filtered['Feedback'].str.contains(search_q.strip(), case=False, na=False)]
    st.markdown(f'<div style="font-family:Space Mono,monospace;font-size:0.75rem;color:#888;margin-bottom:1rem;">Showing {len(filtered):,} of {total:,}</div>', unsafe_allow_html=True)
    view = st.radio("View", ["📋 Table","🃏 Cards"], horizontal=True, key=f"view_{sid[:8]}")
    if view == "📋 Table":
        dcols = [c for c in ['Feedback_ID','Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion'] if c in filtered.columns]
        st.dataframe(filtered[dcols], use_container_width=True, height=420,
            column_config={"Feedback_ID":st.column_config.TextColumn("ID",width="small"),
                "Feedback":st.column_config.TextColumn("Feedback",width="large"),
                "Suggestion":st.column_config.TextColumn("AI Suggestion",width="large"),
                "Sentiment":st.column_config.TextColumn("Sentiment",width="small"),
                "Topic":st.column_config.TextColumn("Topic",width="medium"),
                "Reviewer_Name":st.column_config.TextColumn("Reviewer",width="small"),
                "Feedback_Date":st.column_config.TextColumn("Date",width="small")})
    else:
        for _, row in filtered.head(30).iterrows():
            tid = [k for k,v in TOPIC_LABELS.items() if v==row['Topic']]
            bc = TOPIC_COLORS.get(tid[0] if tid else 0,"badge-0")
            fid = row.get('Feedback_ID',''); name = row.get('Reviewer_Name','')
            date = row.get('Feedback_Date',''); sent = row.get('Sentiment','')
            SENT_C = {'Positive':'#4caf50','Negative':'#e91e63','Neutral':'#9e9e9e','Mixed':'#ff9800'}
            sc_hex = SENT_C.get(sent,'#9e9e9e')
            SENT_I = {'Positive':'😊','Negative':'😞','Neutral':'😐','Mixed':'🤔'}
            sent_icon = SENT_I.get(sent,'·')
            meta = []
            if name and str(name).strip(): meta.append(f"👤 {name}")
            if date and str(date).strip(): meta.append(f"📅 {date}")
            mh = "  &nbsp;·&nbsp;  ".join(meta)
            fb_full = row['Feedback']; fb_display = fb_full[:1200]
            fb_overflow = (f'<span style="color:#f9a825;font-family:Space Mono,monospace;font-size:0.7rem;"> …[{len(fb_full)-1200} more chars]</span>' if len(fb_full) > 1200 else '')
            st.markdown(f"""<div class="card"><div style="display:flex;align-items:center;gap:8px;margin-bottom:0.8rem;flex-wrap:wrap;">
    <span class="feedback-id">{fid}</span><span class="source-tag">{row['Source']}</span>
    <span class="badge {bc}">{row['Topic']}</span>
    {f'<span style="background:rgba(0,0,0,0.3);border:1px solid {sc_hex};border-radius:12px;padding:1px 10px;font-family:Space Mono,monospace;font-size:0.7rem;color:{sc_hex};font-weight:700;">{sent_icon} {sent}</span>' if sent else ''}
    {f'<span style="font-family:Space Mono,monospace;font-size:0.68rem;color:#888;">{mh}</span>' if mh else ''}
</div><div style="color:#ddd;font-size:0.9rem;line-height:1.7;margin-bottom:0.8rem;">{fb_display}{fb_overflow}</div>
<div style="border-top:1px solid rgba(255,255,255,0.06);padding-top:0.8rem;">
    <span style="font-family:Space Mono,monospace;font-size:0.65rem;color:#f9a825;letter-spacing:2px;text-transform:uppercase;">💡 AI Suggestion</span>
    <div style="color:#b0b0b0;font-size:0.85rem;margin-top:4px;">{row.get('Suggestion','')}</div>
</div></div>""", unsafe_allow_html=True)
        if len(filtered) > 30: st.info(f"Showing 30 cards. Switch to Table for all {len(filtered)}.")
    st.divider(); st.markdown("### 📥 Download")
    d1,d2,d3 = st.columns(3)
    with d1:
        ecols = [c for c in ['Feedback_ID','Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion'] if c in filtered.columns]
        st.download_button("⬇️ CSV — Full Data", data=filtered[ecols].to_csv(index=False).encode('utf-8'),
            file_name=f"zeus_feedback_{sid}.csv", mime='text/csv', use_container_width=True, key=f"dl_csv_{sid[:8]}")
    with d2:
        txt = make_summary_txt(filtered, sid, uname_val)
        st.download_button("⬇️ TXT Summary Report", data=txt.encode('utf-8'),
            file_name=f"zeus_report_{sid}.txt", mime='text/plain', use_container_width=True, key=f"dl_txt_{sid[:8]}")
    with d3:
        try:
            xls = make_feedback_excel(filtered, sid, uname_val)
            if xls:
                st.download_button("⬇️ Excel Report (.xlsx)", data=xls,
                    file_name=f"zeus_feedback_{sid}.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True, key=f"dl_xls_{sid[:8]}")
            else: st.info("`pip install openpyxl` for Excel")
        except: st.info("`pip install openpyxl` for Excel")
    st.divider()
    if st.button("🔄 Clear & Start Over", key=f"clear_{sid[:8]}"):
        for k in ['results_df','analyzed','session_id','short_uuid','db_step','db_connected',
                  'db_raw_df','db_detected_cols','db_metadata']:
            st.session_state[k] = None if k in ('results_df','session_id','short_uuid') \
                else (False if k in ('analyzed','db_connected') \
                else (1 if k=='db_step' else ({} if k=='db_detected_cols' else ([] if k=='db_metadata' else ''))))
        st.rerun()

def render_history_tab():
    st.markdown('<div class="card-title">🗄️ Session History</div>', unsafe_allow_html=True)
    sessions_df = get_all_sessions()
    if sessions_df.empty: st.info("No sessions yet."); return
    st.markdown(f"**{len(sessions_df)} session(s)** in database")
    st.dataframe(sessions_df, use_container_width=True, column_config={
        "session_id":st.column_config.TextColumn("Session ID",width="large"),
        "user_name":st.column_config.TextColumn("Analyst"),
        "created_date":st.column_config.TextColumn("Date"),
        "created_time":st.column_config.TextColumn("Time"),
        "source_type":st.column_config.TextColumn("Source"),
        "total_entries":st.column_config.NumberColumn("Entries"),
        "notes":st.column_config.TextColumn("Notes",width="medium")})
    st.divider(); st.markdown("### 🔍 Load & Export Session")
    sel_sid = st.selectbox("Select Session", sessions_df['session_id'].tolist(),
        format_func=lambda x: f"{x}  ·  {sessions_df[sessions_df['session_id']==x]['created_date'].values[0]}  {sessions_df[sessions_df['session_id']==x]['created_time'].values[0]}")
    if sel_sid:
        entries = get_session_entries(sel_sid)
        if not entries.empty:
            meta = sessions_df[sessions_df['session_id']==sel_sid].iloc[0]
            st.markdown(f'<div class="session-box">🔑 <strong>{sel_sid}</strong><br>👤 {meta["user_name"]} &nbsp;·&nbsp; 📅 {meta["created_date"]} {meta["created_time"]} &nbsp;·&nbsp; 📊 {len(entries)} entries &nbsp;·&nbsp; 🌐 {meta["source_type"]}</div>', unsafe_allow_html=True)
            st.dataframe(entries[['feedback_id','source','reviewer_name','feedback_date','feedback','topic','sentiment','suggestion','analyzed_at']], use_container_width=True, height=320)
            hist_df = entries.rename(columns={'feedback_id':'Feedback_ID','source':'Source','reviewer_name':'Reviewer_Name','feedback_date':'Feedback_Date','feedback':'Feedback','topic':'Topic','suggestion':'Suggestion','sentiment':'Sentiment'})
            h1,h2,h3 = st.columns(3)
            with h1:
                ecols = [c for c in ['Feedback_ID','Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion'] if c in hist_df.columns]
                st.download_button("⬇️ CSV", data=hist_df[ecols].to_csv(index=False).encode('utf-8'), file_name=f"zeus_{sel_sid}.csv", mime='text/csv', use_container_width=True)
            with h2:
                txt = make_summary_txt(hist_df, sel_sid, meta['user_name'])
                st.download_button("⬇️ TXT Report", data=txt.encode('utf-8'), file_name=f"zeus_{sel_sid}_report.txt", mime='text/plain', use_container_width=True)
            with h3:
                try:
                    xls = make_feedback_excel(hist_df, sel_sid, meta['user_name'])
                    if xls:
                        st.download_button("⬇️ Excel Report", data=xls, file_name=f"zeus_{sel_sid}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', use_container_width=True)
                except: pass
            full_db = get_full_db_export()
            st.download_button("⬇️ Full Database CSV", data=full_db.to_csv(index=False).encode('utf-8'), file_name="zeus_full_database.csv", mime='text/csv', use_container_width=True)
            st.divider()
            if st.button(f"🗑️ Delete Session `{sel_sid}`"):
                delete_session(sel_sid); st.success("Session deleted."); st.rerun()
        else: st.warning("No entries found for this session.")
    st.divider()
    with st.expander("⚠️ Danger Zone — Clear All History"):
        st.warning("Permanently deletes ALL sessions and entries.")
        if st.button("🔥 Wipe All History"):
            conn = sqlite3.connect(DB_PATH)
            conn.execute("DELETE FROM feedback_entries"); conn.execute("DELETE FROM sessions")
            conn.commit(); conn.close()
            global _HASH_CACHE, _HASH_CACHE_LOADED
            _HASH_CACHE = set(); _HASH_CACHE_LOADED = False
            st.success("All history cleared."); st.rerun()

def render_step_bar(current_step, steps):
    html = '<div class="step-bar">'
    for i, label in enumerate(steps, 1):
        if i < current_step: css, prefix = "step-item step-done", "✓ "
        elif i == current_step: css, prefix = "step-item step-active", ""
        else: css, prefix = "step-item step-pending", ""
        html += f'<div class="{css}">{prefix}{label}</div>'
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# TAB LAYOUT
# ══════════════════════════════════════════════════════════════
mode = st.session_state.nav_mode
has_data = st.session_state.analyzed and st.session_state.results_df is not None
TAB_DEFS = {
    "CSV": ["📄 CSV Upload","📊 Results","📁 History"] if has_data else ["📄 CSV Upload","📁 History"],
    "DB":  ["🗄️ DB Connect","📊 Results","📁 History"] if has_data else ["🗄️ DB Connect","📁 History"],
    "URL": ["🔗 URL Scraper","📊 Results","📁 History"] if has_data else ["🔗 URL Scraper","📁 History"],
}
tabs = st.tabs(TAB_DEFS[mode])
tab_main    = tabs[0]
tab_results = tabs[1] if has_data else None
tab_history = tabs[2] if has_data else tabs[1]

# ══════════════════════════════════════════════════════════════
# CSV TAB
# ══════════════════════════════════════════════════════════════
if mode == "CSV":
    with tab_main:
        st.markdown('<div class="card"><div class="card-title">📤 Upload Feedback CSV</div>', unsafe_allow_html=True)
        uploaded = st.file_uploader("Drop your CSV file here", type=['csv'], label_visibility="visible")
        if uploaded:
            try:
                df_raw = pd.read_csv(uploaded); df_raw.columns = df_raw.columns.str.strip()
                total_rows, total_cols = len(df_raw), len(df_raw.columns)
                st.markdown(f'<div style="background:rgba(76,175,80,0.08);border:1px solid rgba(76,175,80,0.4);border-radius:10px;padding:1rem 1.4rem;margin-bottom:1rem;"><div style="font-family:Space Mono,monospace;font-size:0.72rem;color:#4caf50;letter-spacing:2px;text-transform:uppercase;margin-bottom:0.5rem;">✅ File loaded successfully</div><div style="display:flex;gap:2rem;font-family:Space Mono,monospace;font-size:0.78rem;color:#ddd;"><span>📊 <strong style="color:#f9a825;">{total_rows:,}</strong> rows</span><span>🗂 <strong style="color:#f9a825;">{total_cols}</strong> columns</span><span>📄 <strong style="color:#f9a825;">{uploaded.name}</strong></span></div></div>', unsafe_allow_html=True)
                chips_html = ''.join([f'<span class="col-chip">{col}</span>' for col in df_raw.columns])
                st.markdown(f'<div style="margin-bottom:1rem;">{chips_html}</div>', unsafe_allow_html=True)
                with st.expander("👁️ Preview (first 10 rows)", expanded=True):
                    st.dataframe(df_raw.head(10), use_container_width=True)

                # Smart detection for CSV too
                detected = smart_detect_columns(df_raw)
                auto_feedback_col = detected.get("feedback_col")
                auto_name_col     = detected.get("name_col")
                auto_date_col     = detected.get("date_col")
                layer_used        = detected.get("layer_used", "heuristic")

                st.markdown(
                    f'<div style="background:rgba(0,188,212,0.07);border:1px solid rgba(0,188,212,0.25);'
                    f'border-radius:10px;padding:1rem 1.4rem;margin-bottom:1rem;">'
                    f'<div style="font-family:Space Mono,monospace;font-size:0.7rem;color:#00bcd4;'
                    f'letter-spacing:2px;text-transform:uppercase;margin-bottom:0.6rem;">'
                    f'🤖 Auto-detected mapping · {layer_used}</div>'
                    f'<div style="font-family:Space Mono,monospace;font-size:0.75rem;color:#ddd;line-height:2;">'
                    f'<span style="color:#f9a825;">Feedback column →</span> <strong>{auto_feedback_col or "(not found)"}</strong><br>'
                    f'<span style="color:#f9a825;">Reviewer name   →</span> <strong>{auto_name_col or "(not found — AI extraction)"}</strong><br>'
                    f'<span style="color:#f9a825;">Date column     →</span> <strong>{auto_date_col or "(not found)"}</strong>'
                    f'</div></div>',
                    unsafe_allow_html=True
                )

                if not auto_feedback_col:
                    st.error("❌ Could not auto-detect a feedback/text column.")
                else:
                    notes = st.text_input("Session notes (optional)", placeholder="e.g. Q2 2025 product feedback")
                    if st.button("🚀 Analyze CSV Feedback", key="csv_btn"):
                        nm = st.session_state.user_name or "USER"
                        sid, uid8 = gen_sid(nm); st.session_state.session_id = sid; st.session_state.short_uuid = uid8
                        fbs  = df_raw[auto_feedback_col].dropna().astype(str).tolist()
                        revs = df_raw[auto_name_col].astype(str).tolist() if auto_name_col else ['']*len(fbs)
                        dts  = df_raw[auto_date_col].astype(str).tolist() if auto_date_col else ['']*len(fbs)
                        dts  = [extract_exact_date(d) or d for d in dts]
                        with st.spinner("Analyzing..."):
                            results = build_results(fbs, "CSV Upload", revs, dts, mode='suggestion')
                        if results.empty: st.error("❌ No processable feedback found.")
                        else:
                            st.session_state.results_df = results; st.session_state.analyzed = True
                            saved, skipped = save_entries(sid, results)
                            save_session(sid, nm, uid8, "CSV Upload", saved, notes)
                            msg = f"✅ **{len(results):,}** results · **{saved:,}** new saved"
                            if skipped > 0: msg += f" · **{skipped}** skipped (duplicates)"
                            st.success(msg + " → switch to **📊 Results** tab"); st.rerun()
            except Exception as e:
                st.error(f"Error reading file: {e}")
        else:
            st.markdown('<div style="border:2px dashed rgba(249,168,37,0.4);border-radius:12px;padding:3rem 2rem;text-align:center;background:rgba(249,168,37,0.03);margin-top:1rem;"><div style="font-size:3rem;margin-bottom:0.8rem;">📂</div><div style="font-family:Space Mono,monospace;font-size:0.85rem;color:#888;letter-spacing:1px;">DROP YOUR CSV FILE HERE</div><div style="font-family:Space Mono,monospace;font-size:0.68rem;color:#555;margin-top:0.4rem;">All columns auto-detected · No manual selection needed</div></div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# DB CONNECT TAB  — 2-step flow with smart detection + metadata
# ══════════════════════════════════════════════════════════════
elif mode == "DB":
    with tab_main:
        STEPS = ["1 · Configure & Connect", "2 · Select, Detect & Analyze"]
        render_step_bar(st.session_state.db_step, STEPS)

        # ──────────────────────── STEP 1 ─────────────────────────────────────
        if st.session_state.db_step == 1:
            st.markdown('<div class="card"><div class="card-title">🔌 Step 1 — Configure Database Connection</div>', unsafe_allow_html=True)
            db_type = st.selectbox("Database Type", ALL_DB_TYPES, key="db_type_select")
            st.session_state.db_type = db_type
            pkg, avail = DB_DRIVERS[db_type]
            if not avail and pkg != "built-in":
                st.warning(f"⚠️ Driver missing — `pip install {pkg}`")

            conn_params = {}

            if db_type == "SQLite (file)":
                f = st.file_uploader("Upload .db / .sqlite file", type=["db","sqlite","sqlite3"], key="sqlite_upload")
                conn_params = {"file": f}

            elif db_type == "MySQL / MariaDB":
                c1,c2 = st.columns([3,1])
                with c1: h  = st.text_input("Host","localhost",key="my_h")
                with c2: p  = st.text_input("Port","3306",key="my_p")
                c3,c4 = st.columns(2)
                with c3: u  = st.text_input("Username",key="my_u")
                with c4: pw = st.text_input("Password",type="password",key="my_pw")
                db = st.text_input("Database name",key="my_db")
                conn_params = dict(host=h,port=p,user=u,password=pw,database=db)

            elif db_type == "PostgreSQL":
                c1,c2 = st.columns([3,1])
                with c1: h  = st.text_input("Host","localhost",key="pg_h")
                with c2: p  = st.text_input("Port","5432",key="pg_p")
                c3,c4 = st.columns(2)
                with c3: u  = st.text_input("Username",key="pg_u")
                with c4: pw = st.text_input("Password",type="password",key="pg_pw")
                db = st.text_input("Database name",key="pg_db")
                conn_params = dict(host=h,port=p,user=u,password=pw,database=db)

            elif db_type == "SQL Server (MSSQL)":
                c1,c2 = st.columns([3,1])
                with c1: h  = st.text_input("Host / Server",key="ms_h")
                with c2: p  = st.text_input("Port","1433",key="ms_p")
                c3,c4 = st.columns(2)
                with c3: u  = st.text_input("Username",key="ms_u")
                with c4: pw = st.text_input("Password",type="password",key="ms_pw")
                db = st.text_input("Database name",key="ms_db")
                conn_params = dict(host=h,port=p,user=u,password=pw,database=db)

            elif db_type == "Snowflake":
                acct = st.text_input("Account (e.g. xy12345.us-east-1)",key="sf_acct")
                c1,c2 = st.columns(2)
                with c1: u  = st.text_input("Username",key="sf_u")
                with c2: pw = st.text_input("Password",type="password",key="sf_pw")
                c3,c4,c5 = st.columns(3)
                with c3: wh = st.text_input("Warehouse",key="sf_wh")
                with c4: db = st.text_input("Database",key="sf_db")
                with c5: sc = st.text_input("Schema","PUBLIC",key="sf_sc")
                conn_params = dict(account=acct,user=u,password=pw,warehouse=wh,database=db,schema=sc)

            elif db_type == "MongoDB":
                st.markdown('<div style="background:rgba(249,168,37,0.06);border:1px solid rgba(249,168,37,0.25);border-radius:8px;padding:0.7rem 1rem;margin-bottom:0.8rem;font-family:Space Mono,monospace;font-size:0.7rem;color:#f9a825;">⚡ Special characters auto-escaped — paste URI as-is.</div>', unsafe_allow_html=True)
                mongo_mode = st.radio("Input mode", ["📋 Full URI", "🔑 Host + Credentials"], horizontal=True, key="mg_mode")
                if mongo_mode == "📋 Full URI":
                    uri   = st.text_input("Connection URI", placeholder="mongodb+srv://user:pass@cluster0.xxxxx.mongodb.net/", key="mg_uri")
                    mg_db = st.text_input("Database name", key="mg_db")
                    conn_params = dict(uri=uri, database=mg_db)
                else:
                    from urllib.parse import quote_plus
                    mg_scheme = st.selectbox("Scheme", ["mongodb+srv","mongodb"], key="mg_scheme")
                    c1,c2 = st.columns(2)
                    with c1: mg_user = st.text_input("Username", key="mg_user_split")
                    with c2: mg_pass = st.text_input("Password", type="password", key="mg_pass_split")
                    mg_host = st.text_input("Host / Cluster", placeholder="cluster0.xxxxx.mongodb.net", key="mg_host_split")
                    mg_db   = st.text_input("Database name", key="mg_db_split")
                    enc_u   = quote_plus(mg_user) if mg_user else ""
                    enc_p   = quote_plus(mg_pass) if mg_pass else ""
                    cred    = f"{enc_u}:{enc_p}" if enc_p else enc_u
                    built_uri = f"{mg_scheme}://{cred}@{mg_host}/" if enc_u and mg_host else ""
                    conn_params = dict(uri=built_uri, database=mg_db)

            elif db_type == "Google BigQuery":
                proj      = st.text_input("GCP Project ID", key="bq_proj")
                cred_file = st.file_uploader("Service Account JSON (optional)", type=["json"], key="bq_cred")
                cred_json = cred_file.read().decode() if cred_file else None
                conn_params = dict(project=proj, credentials_json=cred_json)

            elif db_type == "Google Analytics 4":
                prop_id   = st.text_input("GA4 Property ID (numeric)", key="ga_prop")
                cred_file = st.file_uploader("Service Account JSON", type=["json"], key="ga_cred")
                cred_json = cred_file.read().decode() if cred_file else None
                c1,c2 = st.columns(2)
                with c1: ga_start = st.text_input("Start date","30daysAgo",key="ga_start")
                with c2: ga_end   = st.text_input("End date","today",key="ga_end")
                conn_params = dict(property_id=prop_id, credentials_json=cred_json,
                                   start_date=ga_start, end_date=ga_end)

            elif db_type == "Databricks SQL":
                db_host = st.text_input("Server Hostname", placeholder="adb-xxxx.azuredatabricks.net", key="dbr_host")
                db_http = st.text_input("HTTP Path", placeholder="/sql/1.0/warehouses/xxxx", key="dbr_http")
                db_tok  = st.text_input("Access Token", type="password", key="dbr_tok")
                c1,c2 = st.columns(2)
                with c1: db_cat = st.text_input("Catalog","main",key="dbr_cat")
                with c2: db_sc  = st.text_input("Schema","default",key="dbr_sc")
                conn_params = dict(host=db_host,http_path=db_http,token=db_tok,
                                   catalog=db_cat,schema=db_sc)

            st.session_state.db_conn_params = conn_params

            if st.button("🔌 Connect & Discover Tables →", key="db_connect_btn"):
                with st.spinner("Connecting…"):
                    try:
                        tables = []; src_label = db_type.split()[0]

                        if db_type == "SQLite (file)":
                            if not conn_params.get("file"): st.error("❌ Upload a SQLite file first."); st.stop()
                            tmp = f"/tmp/zeus_{uuid.uuid4().hex[:8]}.db"
                            with open(tmp,"wb") as fh: fh.write(conn_params["file"].read())
                            st.session_state.db_tmp_path = tmp
                            tables = _list_sqlite_tables(tmp); src_label = "SQLite"

                        elif db_type == "MySQL / MariaDB":
                            tables = _list_mysql_tables(conn_params['host'],conn_params['port'],conn_params['user'],conn_params['password'],conn_params['database']); src_label = "MySQL"

                        elif db_type == "PostgreSQL":
                            tables = _list_postgres_tables(conn_params['host'],conn_params['port'],conn_params['user'],conn_params['password'],conn_params['database']); src_label = "PostgreSQL"

                        elif db_type == "SQL Server (MSSQL)":
                            tables = _list_mssql_tables(conn_params['host'],conn_params['port'],conn_params['user'],conn_params['password'],conn_params['database']); src_label = "SQLServer"

                        elif db_type == "Snowflake":
                            tables = _list_snowflake_tables(conn_params['account'],conn_params['user'],conn_params['password'],conn_params['warehouse'],conn_params['database'],conn_params['schema']); src_label = "Snowflake"

                        elif db_type == "MongoDB":
                            tables = _list_mongodb_collections(conn_params['uri'],conn_params['database']); src_label = "MongoDB"

                        elif db_type == "Google BigQuery":
                            tables = _list_bigquery_datasets(conn_params['project'],conn_params.get('credentials_json')); src_label = "BigQuery"

                        elif db_type == "Google Analytics 4":
                            tables = ["(GA4 engagement report)"]; src_label = "GA4"

                        elif db_type == "Databricks SQL":
                            tables = _list_databricks_tables(conn_params['host'],conn_params['http_path'],conn_params['token'],conn_params.get('catalog','main'),conn_params.get('schema','default')); src_label = "Databricks"

                        st.session_state.db_available_tables = tables
                        st.session_state.db_src_label        = src_label
                        st.session_state.db_connected        = True
                        st.session_state.db_step             = 2
                        st.success(f"✅ Connected to **{src_label}** — found **{len(tables)}** object(s)")
                        st.rerun()

                    except ImportError as e: st.error(f"❌ Missing driver: {e}")
                    except Exception as e:   st.error(f"❌ Connection failed: {str(e)[:300]}")

            st.markdown('</div>', unsafe_allow_html=True)

        # ──────────────────────── STEP 2 ─────────────────────────────────────
        # Select table → Fetch ALL rows → Show metadata → Smart detect → Analyze
        # ─────────────────────────────────────────────────────────────────────
        elif st.session_state.db_step == 2:
            tables      = st.session_state.db_available_tables
            src_label   = st.session_state.db_src_label
            db_type     = st.session_state.db_type
            conn_params = st.session_state.db_conn_params
            tmp_path    = st.session_state.get("db_tmp_path","")

            st.markdown(
                f'<div class="session-box">✅ Connected to <strong>{src_label}</strong>'
                f' &nbsp;·&nbsp; <strong>{len(tables)}</strong> object(s) found</div>',
                unsafe_allow_html=True
            )

            st.markdown('<div class="card"><div class="card-title">📂 Step 2 — Select Table, Inspect Metadata & Analyze</div>', unsafe_allow_html=True)

            # ── Table selector ────────────────────────────────────────────────
            selected_table = st.selectbox(
                "Select table / collection", tables, key="db_table_select"
            )
            st.session_state.db_selected_table = selected_table

            # ── Fetch button ──────────────────────────────────────────────────
            if st.button("🔍 Fetch All Rows & Inspect", key="s2_fetch_all"):
                with st.spinner(f"Fetching all rows from **{selected_table}**…"):
                    try:
                        raw_df = fetch_all_rows(db_type, conn_params, selected_table, tmp_path)
                        if raw_df is None or raw_df.empty:
                            st.error("❌ Table is empty or returned no rows."); st.stop()

                        # Build metadata FIRST (uses raw_df for full stats)
                        with st.spinner("📊 Building column metadata…"):
                            meta = get_table_metadata(db_type, conn_params, selected_table,
                                                      tmp_path, raw_df=raw_df)

                        # Smart detection using pre-defined dictionary + data
                        with st.spinner("🤖 Detecting feedback column…"):
                            detected = smart_detect_columns(raw_df)

                        st.session_state.db_raw_df        = raw_df
                        st.session_state.db_detected_cols = detected
                        st.session_state.db_metadata      = meta
                        st.rerun()

                    except Exception as ex:
                        st.error(f"❌ {str(ex)[:400]}")

            # ── Show full panel once data loaded ──────────────────────────────
            raw_df   = st.session_state.get("db_raw_df")
            detected = st.session_state.get("db_detected_cols", {})
            meta     = st.session_state.get("db_metadata", [])

            if raw_df is not None and not raw_df.empty and detected:
                total_rows = len(raw_df); total_cols = len(raw_df.columns)

                # ── Stats banner ──────────────────────────────────────────────
                m1,m2,m3,m4 = st.columns(4)
                m1.metric("Total Rows",    f"{total_rows:,}")
                m2.metric("Total Columns", f"{total_cols}")
                m3.metric("Table",         selected_table)
                m4.metric("Source",        src_label)

                st.divider()

                # ── METADATA PANEL ────────────────────────────────────────────
                with st.expander("🗂️ Column Metadata Inspector", expanded=True):
                    render_metadata_panel(meta, detected_fb_col=detected.get("feedback_col",""))

                st.divider()

                # ── Detection result banner ───────────────────────────────────
                layer      = detected.get("layer_used", "Unknown")
                conf_score = detected.get("confidence_score", 0)
                reason     = detected.get("reason", "")
                layer_colors = {
                    "Layer 1": "#4caf50", "Layer 2": "#f9a825",
                    "Layer 3": "#00bcd4", "Layer 4": "#e91e63",
                }
                lc = next((v for k, v in layer_colors.items() if k in layer), "#888")

                st.markdown(
                    f'<div style="background:rgba(0,0,0,0.25);border:1px solid {lc}40;'
                    f'border-radius:10px;padding:0.9rem 1.2rem;margin-bottom:1rem;">'
                    f'<div style="display:flex;align-items:center;gap:1.5rem;flex-wrap:wrap;">'
                    f'<span style="font-family:Space Mono,monospace;font-size:0.72rem;'
                    f'color:{lc};font-weight:700;">🎯 {layer}</span>'
                    f'<span style="font-family:Space Mono,monospace;font-size:0.8rem;'
                    f'color:#f9a825;font-weight:800;">Confidence: {conf_score}/100</span>'
                    f'</div>'
                    f'<div style="font-size:0.8rem;color:#bbb;margin-top:4px;">{reason}</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )

                # ── Per-column confidence bars ────────────────────────────────
                scores = detected.get("column_scores", {})
                if scores:
                    st.markdown('<div style="font-family:Space Mono,monospace;font-size:0.7rem;color:#888;margin-bottom:0.5rem;letter-spacing:1px;text-transform:uppercase;">Feedback probability per column</div>', unsafe_allow_html=True)
                    sorted_scores = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:6]
                    bar_cols = st.columns(len(sorted_scores))
                    for i, (col, score) in enumerate(sorted_scores):
                        is_top = i == 0
                        css = "conf-card-top" if is_top else "conf-card"
                        clr = "#f9a825" if is_top else "#aaa"
                        with bar_cols[i]:
                            st.markdown(
                                f'<div class="{css}">'
                                f'<div style="font-family:Space Mono,monospace;font-size:0.62rem;color:{clr};margin-bottom:3px;">{col[:16]}</div>'
                                f'<div style="font-size:1.1rem;font-weight:800;color:#f9a825;font-family:Space Mono,monospace;">{score}%</div>'
                                f'<div style="background:rgba(249,168,37,0.12);border-radius:3px;height:4px;margin-top:4px;">'
                                f'<div style="background:#f9a825;width:{score}%;height:4px;border-radius:3px;"></div></div></div>',
                                unsafe_allow_html=True
                            )

                st.divider()

                # ── Column mapping confirmation ────────────────────────────────
                st.markdown("**📌 Confirm column mapping** *(auto-detected — override if needed)*")
                all_cols = ["(none)"] + list(raw_df.columns)
                col_list = list(raw_df.columns)

                def _si(lst, val):
                    try: return lst.index(val) if val in lst else 0
                    except: return 0

                cm1,cm2,cm3 = st.columns(3)
                with cm1:
                    fb_col = st.selectbox("📝 Feedback column *", col_list,
                        index=_si(col_list, detected.get("feedback_col") or ""), key="final_fb_col")
                with cm2:
                    nm_col = st.selectbox("👤 Reviewer name", all_cols,
                        index=_si(all_cols, detected.get("name_col") or "(none)"), key="final_nm_col")
                with cm3:
                    dt_col = st.selectbox("📅 Date column", all_cols,
                        index=_si(all_cols, detected.get("date_col") or "(none)"), key="final_dt_col")

                # ── Sample preview ────────────────────────────────────────────
                if fb_col:
                    samples = raw_df[fb_col].dropna().astype(str).head(3).tolist()
                    sample_html = "".join(
                        f'<div style="font-size:0.8rem;color:#ccc;padding:4px 0;border-bottom:1px solid rgba(255,255,255,0.05);">"{s[:250]}"</div>'
                        for s in samples
                    )
                    st.markdown(
                        f'<div style="background:rgba(249,168,37,0.04);border:1px solid rgba(249,168,37,0.2);'
                        f'border-radius:8px;padding:0.8rem 1.2rem;margin:0.8rem 0;">'
                        f'<div style="font-family:Space Mono,monospace;font-size:0.68rem;color:#f9a825;margin-bottom:6px;">'
                        f'📋 SAMPLE VALUES — {fb_col}</div>{sample_html}</div>',
                        unsafe_allow_html=True
                    )

                # ── Summary confirmation line ──────────────────────────────────
                st.markdown(
                    f'<div style="background:rgba(249,168,37,0.07);border:1px solid rgba(249,168,37,0.3);'
                    f'border-radius:8px;padding:0.8rem 1.2rem;margin:0.8rem 0;">'
                    f'<div style="font-family:Space Mono,monospace;font-size:0.72rem;color:#f9a825;">'
                    f'Will analyze: <strong>{fb_col}</strong> as feedback'
                    f'{f" · <strong>{nm_col}</strong> as name" if nm_col != "(none)" else " · AI will extract names"}'
                    f'{f" · <strong>{dt_col}</strong> as date" if dt_col != "(none)" else ""}'
                    f' · <strong>{total_rows:,}</strong> total rows'
                    f'</div></div>',
                    unsafe_allow_html=True
                )

                db_notes = st.text_input("Session notes (optional)",
                    placeholder="e.g. Production DB — Q2 reviews", key="db_notes_s2")

                b1, b2 = st.columns([1,4])
                with b1:
                    if st.button("← Back", key="s2_back"):
                        st.session_state.db_step          = 1
                        st.session_state.db_connected     = False
                        st.session_state.db_raw_df        = None
                        st.session_state.db_detected_cols = {}
                        st.session_state.db_metadata      = []
                        st.rerun()
                with b2:
                    if st.button("🚀 Run Feedback Analysis →", key="s2_analyze"):
                        nm = st.session_state.user_name or "USER"
                        sid, uid8 = gen_sid(nm)
                        st.session_state.session_id = sid
                        st.session_state.short_uuid = uid8

                        fbs  = raw_df[fb_col].dropna().astype(str).tolist()
                        revs = raw_df[nm_col].astype(str).tolist() if nm_col != "(none)" else ['']*len(fbs)
                        dts  = raw_df[dt_col].astype(str).tolist() if dt_col != "(none)" else ['']*len(fbs)
                        dts  = [extract_exact_date(d) or d for d in dts]

                        with st.spinner("Analyzing feedback…"):
                            results = build_results(fbs, src_label, revs, dts, mode='suggestion')

                        if results.empty:
                            st.error("❌ No processable feedback found.")
                        else:
                            st.session_state.results_df = results
                            st.session_state.analyzed   = True
                            saved, skipped = save_entries(sid, results)
                            save_session(sid, nm, uid8, src_label, saved, db_notes)
                            msg = f"✅ **{len(results):,}** results · **{saved:,}** saved"
                            if skipped > 0: msg += f" · **{skipped}** duplicates skipped"
                            st.success(msg + " → switch to **📊 Results** tab")
                            st.session_state.db_step          = 1
                            st.session_state.db_connected     = False
                            st.session_state.db_raw_df        = None
                            st.session_state.db_detected_cols = {}
                            st.session_state.db_metadata      = []
                            st.rerun()

            st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# URL SCRAPER TAB
# ══════════════════════════════════════════════════════════════
elif mode == "URL":
    _BAD_CLS = re.compile(r'\bnav\b|\bmenu\b|\bheader\b|\bfooter\b|\bsidebar\b|\bcookie\b|\bbanner\b|\bmodal\b|\bpopup\b', re.I)

    def _get_headers():
        ua = random.choice(["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120 Safari/537.36",
                            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/119 Safari/537.36"])
        return {"User-Agent":ua,"Accept":"text/html,application/xhtml+xml,*/*;q=0.8","Accept-Language":"en-US,en;q=0.9","Referer":"https://www.google.com/"}

    def _fetch_url(url, timeout=15):
        try:
            resp = requests.get(url,headers=_get_headers(),timeout=timeout,allow_redirects=True)
            if resp.status_code == 200: return resp
        except: pass
        return None

    def extract_blocks_generic(soup, min_len=80):
        seen, blocks, pos = set(), [], 0
        containers = soup.find_all(True, class_=re.compile(r'review|comment|feedback|answer|post|testimonial|customer.?review|rating.?item|user.?review|opinion',re.I), limit=500)
        def process(text,name='',date=''):
            nonlocal pos
            cl = clean_text(text)
            if not cl or is_junk(cl) or len(cl) < min_len or is_truncated(cl): return
            k = cl[:120].lower()
            if k in seen: return
            seen.add(k)
            if len(cl) > 3000:
                tr = cl[:3000]; lp = max(tr.rfind('.'),tr.rfind('!'),tr.rfind('?'))
                cl = cl[:lp+1] if lp > 600 else tr
            blocks.append({'text':cl,'name':name,'date':date,'pos':pos}); pos += 1
        for c in containers:
            cls_str = ' '.join(c.get('class') or [])
            if _BAD_CLS.search(cls_str): continue
            process(c.get_text(' ',strip=True))
        if not blocks:
            from bs4 import Tag
            for el in soup.find_all(['div','p','section','article','li','blockquote']):
                if not isinstance(el,Tag): continue
                cls_str = ' '.join(el.get('class') or [])
                if _BAD_CLS.search(cls_str): continue
                if len(el.find_all(['div','section','article'])) > 5: continue
                raw = el.get_text(' ',strip=True)
                if len(raw) < min_len or len(raw) > 8000: continue
                process(raw)
        blocks.sort(key=lambda x:x['pos']); return blocks

    with tab_main:
        st.markdown('<div class="card"><div class="card-title">🌐 Scrape Reviews from URL</div>', unsafe_allow_html=True)
        st.text_input("Paste URL", placeholder="enter url", label_visibility="collapsed")
        st.info("🚫 URL Scraper temporarily disabled by admin", icon="🚫")
        if st.button("🔍 Scrape & Analyze", key="url_btn", disabled=True): pass
        st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# RESULTS TAB
# ══════════════════════════════════════════════════════════════
if tab_results is not None:
    with tab_results:
        render_results_tab(st.session_state.results_df, st.session_state.session_id or "SESSION", st.session_state.user_name or "USER")

# ══════════════════════════════════════════════════════════════
# HISTORY TAB
# ══════════════════════════════════════════════════════════════
with tab_history:
    render_history_tab()
