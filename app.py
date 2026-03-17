import streamlit as st
import pandas as pd
import re, os, sqlite3, uuid, time, io, hashlib, json, random
from datetime import datetime
from bs4 import BeautifulSoup
import requests
import nltk
import openai
from dotenv import load_dotenv
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import NMF
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import WordNetLemmatizer

# ── Streamlit Cloud: NO Selenium / ChromeDriver ─────────────
# All scraping uses requests + BeautifulSoup only.
# JS-heavy sites handled via: mobile URLs, AMP pages,
# internal JSON APIs, and per-domain parsing strategies.

load_dotenv()
openai.api_key = os.getenv("OPENAI_API_KEY")

st.set_page_config(
    page_title="ZEUS FEEDBACK ANALYZER",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@400;600;800&display=swap');
html,body,[class*="css"]{font-family:'Syne',sans-serif;background-color:#0d0d0d;color:#f0ece2;}
#MainMenu,footer,header{visibility:hidden;}
.stApp{background:linear-gradient(135deg,#0d0d0d 0%,#111827 100%);}
.hero-title{font-family:'Syne',sans-serif;font-weight:800;font-size:3.2rem;background:linear-gradient(90deg,#f9a825,#ff6f00,#e91e63);-webkit-background-clip:text;-webkit-text-fill-color:transparent;letter-spacing:-1px;line-height:1.1;margin-bottom:0.2rem;}
.hero-sub{font-family:'Space Mono',monospace;font-size:0.85rem;color:#888;letter-spacing:2px;text-transform:uppercase;margin-bottom:2rem;}
.card{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:16px;padding:1.5rem 2rem;margin-bottom:1.5rem;backdrop-filter:blur(10px);}
.card-title{font-family:'Space Mono',monospace;font-size:0.75rem;letter-spacing:3px;text-transform:uppercase;color:#f9a825;margin-bottom:1rem;}
.badge-0{background:rgba(249,168,37,0.15);border:1px solid #f9a825;color:#f9a825;}
.badge-1{background:rgba(233,30,99,0.15);border:1px solid #e91e63;color:#e91e63;}
.badge-2{background:rgba(0,188,212,0.15);border:1px solid #00bcd4;color:#00bcd4;}
.badge-3{background:rgba(76,175,80,0.15);border:1px solid #4caf50;color:#4caf50;}
.badge-4{background:rgba(156,39,176,0.15);border:1px solid #9c27b0;color:#9c27b0;}
.badge{border-radius:20px;padding:2px 10px;font-size:0.72rem;font-family:'Space Mono',monospace;font-weight:700;}
.source-tag{display:inline-block;background:rgba(249,168,37,0.1);border:1px solid rgba(249,168,37,0.4);border-radius:4px;font-family:'Space Mono',monospace;font-size:0.65rem;padding:2px 8px;color:#f9a825;margin-right:6px;}
.session-box{background:rgba(0,188,212,0.07);border:1px solid rgba(0,188,212,0.25);border-radius:10px;padding:0.8rem 1.2rem;font-family:'Space Mono',monospace;font-size:0.72rem;color:#00bcd4;letter-spacing:0.5px;margin-bottom:1rem;}
.feedback-id{display:inline-block;background:rgba(0,188,212,0.15);border:1px solid rgba(0,188,212,0.5);border-radius:6px;padding:2px 10px;font-family:'Space Mono',monospace;font-size:0.72rem;color:#00bcd4;margin-right:8px;font-weight:700;}
.stTabs [data-baseweb="tab-list"]{gap:8px;border-bottom:1px solid rgba(255,255,255,0.1);}
.stTabs [data-baseweb="tab"]{border-radius:8px 8px 0 0;font-family:'Space Mono',monospace;font-size:0.78rem;letter-spacing:1px;text-transform:uppercase;color:#888;padding:10px 20px;border:none;background:transparent;}
.stTabs [aria-selected="true"]{background:rgba(249,168,37,0.1)!important;color:#f9a825!important;border-bottom:2px solid #f9a825!important;}
.stButton>button{background:linear-gradient(135deg,#f9a825,#ff6f00);color:#0d0d0d;font-family:'Space Mono',monospace;font-weight:700;font-size:0.82rem;letter-spacing:2px;text-transform:uppercase;border:none;border-radius:8px;padding:0.6rem 2rem;width:100%;transition:all 0.2s;}
.stButton>button:hover{transform:translateY(-2px);box-shadow:0 8px 24px rgba(249,168,37,0.4);}
.stTextInput input,.stTextArea textarea{background:rgba(255,255,255,0.05)!important;border:1px solid rgba(255,255,255,0.12)!important;border-radius:8px!important;color:#f0ece2!important;font-family:'Space Mono',monospace!important;font-size:0.85rem!important;}
[data-testid="stSidebar"]{background:rgba(0,0,0,0.4)!important;border-right:1px solid rgba(255,255,255,0.06);}
[data-testid="stMetric"]{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:1rem;}
[data-testid="stMetricValue"]{color:#f9a825!important;font-family:'Space Mono',monospace!important;}
hr{border-color:rgba(255,255,255,0.08);}
.pulse{animation:pulse 2s infinite;}
@keyframes pulse{0%,100%{opacity:1;}50%{opacity:0.5;}}
</style>
""", unsafe_allow_html=True)

@st.cache_resource
def download_nltk():
    for p in ['stopwords', 'punkt', 'wordnet', 'punkt_tab']:
        nltk.download(p, quiet=True)
download_nltk()

TOPIC_LABELS = {
    0: "⚡ Service Quality", 1: "📦 Product Issues",
    2: "💬 Communication",  3: "⏱️ Speed & Delays",
    4: "💡 Innovation & Features"
}
TOPIC_COLORS = {0:"badge-0",1:"badge-1",2:"badge-2",3:"badge-3",4:"badge-4"}

# ══════════════════════════════════════════════════════════════
# SENTIMENT ENGINE — rule-based + optional AI upgrade
# ══════════════════════════════════════════════════════════════

_POS = re.compile(
    r'\b(excellent|amazing|outstanding|wonderful|fantastic|great|good|love|loved|brilliant|superb|perfect|'
    r'delicious|awesome|fabulous|impressive|enjoy|enjoyed|pleasant|satisfied|satisfaction|happy|pleased|'
    r'best|highly recommend|recommend|worth|value for money|value for price|tasty|fresh|'
    r'friendly|helpful|attentive|professional|polite|courteous|efficient|prompt|quick|fast|'
    r'clean|hygienic|comfortable|cozy|ambience|nice place|will visit again|visit again|'
    r'must visit|must try|5 star|five star|10/10|loved it|exceeded expectations|top notch|'
    r'world class|phenomenal|spectacular|mouth.?watering|finger.?licking|well cooked|'
    r'perfectly cooked|good service|great service|good food|great food|amazing food|'
    r'yummy|scrumptious|lip.?smacking)\b', re.I)

_NEG = re.compile(
    r'\b(terrible|horrible|awful|worst|bad|poor|disappointing|disappointed|disgusting|pathetic|'
    r'rude|arrogant|unprofessional|impolite|hostile|aggressive|misbehav|attitude problem|'
    r'tasteless|bland|overcooked|undercooked|half.?cooked|raw|stale|rubbery|dry|cold food|'
    r'slow|delay|wait|waited|too long|overcrowded|understaffed|dirty|unclean|unhygienic|'
    r'overpriced|expensive|rip.?off|not worth|waste of money|never again|avoid|'
    r'do not recommend|not recommend|waste|pest|cockroach|insect|filth|stink|bad smell|'
    r'missing|wrong order|incorrect order|food poisoning|sick|fell ill)\b', re.I)

_MIXED_IND = re.compile(
    r'\bbut\b|\bhowever\b|\bthough\b|\balthough\b|\bexcept\b|\bdespite\b|\byet\b|'
    r'\bon the other hand\b|\bpartially\b|\bnot bad\b|\bok but\b', re.I)

def rule_based_sentiment(text: str) -> str:
    if not text or len(text.strip()) < 10: return 'Neutral'
    t = text.lower().strip()
    pos = len(_POS.findall(t))
    neg = len(_NEG.findall(t))
    if pos >= 1 and neg >= 1 and _MIXED_IND.search(t): return 'Mixed'
    if pos > neg * 1.5: return 'Positive'
    if neg > pos * 1.5: return 'Negative'
    if pos > 0 and neg > 0: return 'Mixed'
    if pos > 0: return 'Positive'
    if neg > 0: return 'Negative'
    mild_pos = re.search(r'\b(nice|okay|ok|fine|decent|average|alright|not bad|acceptable)\b', t)
    mild_neg = re.search(r'\b(not great|could be better|needs improvement|lacking|mediocre|underwhelming)\b', t)
    if mild_pos and not mild_neg: return 'Positive'
    if mild_neg and not mild_pos: return 'Negative'
    if mild_pos and mild_neg: return 'Mixed'
    return 'Neutral'

def get_sentiment_batch_ai(feedback_list: list) -> list:
    if not openai.api_key or not feedback_list:
        return [rule_based_sentiment(fb) for fb in feedback_list]
    try:
        prompt = ("Classify sentiment of each feedback. Reply ONLY with numbered lines:\n"
                  "1. Positive\n2. Negative\n...\nValid: Positive, Negative, Neutral, Mixed\n\nFeedbacks:\n")
        for i, fb in enumerate(feedback_list, 1):
            prompt += f"{i}. {str(fb)[:300]}\n"
        resp = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[{"role":"system","content":"Precise sentiment classifier. Output only numbered labels."},
                      {"role":"user","content":prompt}],
            max_tokens=len(feedback_list)*8+50, temperature=0.0)
        raw = resp["choices"][0]["message"]["content"].strip()
        results = {}
        for line in raw.split('\n'):
            m = re.match(r'^(\d+)\.\s*(Positive|Negative|Neutral|Mixed)', line.strip(), re.I)
            if m: results[int(m.group(1))-1] = m.group(2).capitalize()
        return [results.get(i, rule_based_sentiment(fb)) for i, fb in enumerate(feedback_list)]
    except Exception:
        return [rule_based_sentiment(fb) for fb in feedback_list]

def analyze_sentiments_all(feedback_list: list) -> list:
    rule_results = [rule_based_sentiment(fb) for fb in feedback_list]
    if not openai.api_key: return rule_results
    uncertain_idx = [i for i, s in enumerate(rule_results) if s == 'Neutral']
    if not uncertain_idx: return rule_results
    uncertain_texts = [feedback_list[i] for i in uncertain_idx]
    ai_results = []
    for i in range(0, len(uncertain_texts), 20):
        ai_results.extend(get_sentiment_batch_ai(uncertain_texts[i:i+20]))
    for orig_idx, ai_sent in zip(uncertain_idx, ai_results):
        rule_results[orig_idx] = ai_sent
    return rule_results

# ══════════════════════════════════════════════════════════════
# FALLBACK SUGGESTIONS
# ══════════════════════════════════════════════════════════════

_KW_ACTIONS = [
    (r'made.*wait|waited.*long|took.*too long|long.*queue|slow.*service|delay.*order|waited.*hour|no.*refill',
     "Audit the service timeline at peak hours, assign a dedicated expediter role during rush periods, and set a 10-minute maximum wait target with manager alerts when breached."),
    (r'rude|behaviour|attitude|unprofessional|hostile|impolite|arrogant|disrespectful|misbehav',
     "Launch a monthly frontline staff empathy workshop, introduce peer-review scorecards, and tie guest-satisfaction ratings directly to service team performance incentives."),
    (r'tasteless|no taste|bland|overcooked|undercooked|half.?cooked|raw.*meat|cold.*food|stale|rubbery|insipid|dry.*chicken',
     "Establish a kitchen quality-control checkpoint: assign a dedicated quality officer per shift to inspect every dish before it leaves the kitchen, with a discard protocol for sub-standard items."),
    (r'overpriced|not worth|rip.?off|too expensive|poor.*value|price.*hike|price.*increase|not.*worth.*money',
     "Introduce a transparent value communication strategy — highlight premium ingredients on menus, publish cost-per-portion data for high-value items, and launch a loyalty tier with exclusive repeat-visitor pricing."),
    (r'dirty|unhygienic|unclean|pest|cockroach|insect|filth|stink|smell.*bad|bad.*smell|not.*clean',
     "Schedule bi-weekly hygiene audits with a third-party inspector, post the latest inspection score at the entrance, and create an anonymous staff reporting channel for cleanliness violations."),
    (r'less variety|limited.*option|same.*menu|menu.*not.*change|fewer.*item|not enough.*veg|lack.*choice|no.*option',
     "Conduct a quarterly menu refresh, guarantee a minimum of 6 rotating seasonal items, and introduce a dedicated vegetarian section with at least 8 distinct options."),
    (r'too.*noisy|very.*loud|overcrowd|uncomfortable.*seat|dirty.*floor|broken|torn.*seat|ambience.*bad|bad.*ambience',
     "Commission an acoustic and ambience assessment, install sound-dampening panels in high-density zones, and establish a monthly décor inspection checklist for all outlets."),
    (r'no.*table|could not.*seat|reservation.*not.*honour|lost.*booking|waited.*despite.*booking|table.*not.*ready',
     "Implement a real-time table management system with automated SMS confirmation 30 minutes before arrival, and introduce a 15-minute hold policy to eliminate booking-loss failures."),
    (r'dessert.*finish|no.*dessert|sweet.*not.*available|ice.?cream.*over|cake.*finish|dessert.*empty|dessert.*ran out',
     "Assign a dedicated dessert replenishment attendant during peak hours with a 5-minute restocking SLA, and add 3 rotating regional dessert specials monthly."),
    (r'starter.*finish|no.*starter|grill.*empty|waiting.*for.*starter|prawn.*not.*available|chicken.*finish|item.*not.*available',
     "Install a live grill-station display showing estimated replenishment time per item and empower grill staff to proactively alert tables when a popular item is 5 minutes away."),
    (r'discount.*not.*applied|coupon.*not.*work|offer.*reject|promo.*not.*accept',
     "Audit third-party discount redemption workflows monthly, ensure POS systems auto-validate all active promotions, and train billing staff on same-day coupon escalation procedures."),
    (r'birthday.*ruin|anniversary.*disappoint|event.*mess|party.*not.*organis|celebration.*bad|occasion.*ruin',
     "Create a dedicated celebrations concierge role — confirming dietary needs, arranging décor, and briefing floor staff 30 minutes before the party arrives."),
    (r'app.*crash|app.*not.*work|website.*down|online.*order.*fail|delivery.*wrong|order.*missing|payment.*fail',
     "Conduct a quarterly UX audit of the ordering platform, prioritise the top 3 friction points in user drop-off analytics, and commit to a 2-week fix cycle for critical bugs."),
    (r'no.*parking|parking.*issue|far.*away|location.*bad|difficult.*reach|hard.*find|difficult.*locate',
     "Update Google Maps and Zomato location pins with precise coordinates and real-time parking info; introduce a weekend valet service to eliminate the parking friction."),
    (r'no.*staff|understaffed|no one came|staff.*shortage|ignored.*table|unattended.*table',
     "Implement a table-to-staff ratio cap (max 1:4 during peak hours), deploy a digital floor-management tool to flag unattended tables after 3 minutes, and maintain a trained part-time staff pool for weekends."),
    (r'love|amazing|excellent|outstanding|wonderful|fantastic|highly recommend|best.*place|worth.*visit|thoroughly enjoyed',
     "Capitalise on this positive sentiment by launching a referral programme — offer a complimentary dessert for every first-visit guest brought by a verified loyalist, tracked through the app."),
    (r'loyalt|loyalty.*card|membership|reward|point.*system|regulars|return.*customer|repeat.*visit',
     "Redesign the loyalty programme with a tiered structure (Bronze → Silver → Gold) visible in the app, where Gold members receive a dedicated host, priority seating, and a monthly personalised offer."),
]

_TOPIC_ACTIONS = {
    0:["Introduce a real-time floor-monitoring dashboard for managers, flagging tables with no staff contact beyond 4 minutes, and tie resolution speed to shift-level performance reviews.",
       "Deploy a post-meal digital feedback kiosk at exit points and route sub-4-star ratings directly to the duty manager's phone within 60 seconds for immediate recovery.",
       "Establish a mystery dining programme — quarterly audits by trained evaluators — to surface service gaps that are invisible to in-house management.",
       "Create a 'Service Champion' recognition system where floor staff earn points for positive guest mentions, redeemable monthly.",
       "Map the guest journey from entry to exit and identify the three highest-friction touchpoints; redesign those in a 30-day sprint with frontline staff input."],
    1:["Implement a live product quality log where kitchen staff flag any batch that deviates from standard, triggering an automatic re-prep before it reaches the buffet.",
       "Introduce a guest-facing 'freshness timer' display at each buffet station showing when each dish was last replenished.",
       "Set up a weekly tasting panel — including non-chef staff — to score each dish against a standardised flavour benchmark, with mandatory rework for anything below 7/10.",
       "Conduct a root-cause analysis on the top 3 most-complained-about dishes this quarter and pilot improved recipes at one outlet before rolling out chain-wide.",
       "Create a feedback card at each table specifically rating food quality; share bottom-ranked items with the executive chef weekly for corrective action."],
    2:["Train all floor staff on a standard 3-step proactive communication script — greet within 90 seconds, present menu with one recommendation, check back after first course.",
       "Introduce a pre-visit automated WhatsApp message confirming reservations, sharing the day's special, and inviting dietary preferences.",
       "Set up a post-visit email sequence: a thank-you within 2 hours, a feedback survey within 24 hours, and a personalised offer within 7 days.",
       "Create a laminated 'Guest FAQ' on each table covering top 10 common questions (billing, allergens, refills).",
       "Implement a structured complaint-handling protocol: acknowledge within 1 minute, offer a concrete resolution within 3 minutes, follow up before guest leaves."],
    3:["Introduce staggered entry slots (every 15 minutes) for large group reservations to distribute kitchen load and reduce peak-hour bottlenecks.",
       "Install a kitchen display system (KDS) linking front-of-house seating data to prep workloads in real time.",
       "Set a maximum 8-minute starter replenishment SLA during peak hours with one dedicated runner per grill station.",
       "Analyse historical reservation data to identify the busiest 2-hour window weekly and schedule 20% additional staff specifically for that window.",
       "Pilot a 'Fast Lane' table category for guests with 60-minute dining windows — pre-set with starters already grilled."],
    4:["Launch a 'Guest Chef' quarterly event where top-rated customer recipe suggestions are cooked live at the outlet.",
       "Introduce a digital loyalty passport — every visit earns stamps redeemable for exclusive menu access or table upgrades.",
       "Create a seasonal limited-edition menu released every 90 days, promoted exclusively to loyalty members 48 hours early.",
       "Set up a live innovation lab at one flagship outlet where new dishes are beta-tested with willing diners who receive a discount in exchange for structured feedback.",
       "Partner with a local culinary school for a rotating 'Student Signature Dish' feature — fresh ideas and a consistent stream of menu innovation at low cost."],
}
_topic_ctr = {k: 0 for k in _TOPIC_ACTIONS}

def generate_fallback_suggestion(feedback_text: str, topic_id: int) -> str:
    t = (feedback_text or '').lower()
    for pattern, suggestion in _KW_ACTIONS:
        if re.search(pattern, t): return suggestion
    bucket = int(hashlib.md5(t[:300].encode()).hexdigest()[:8], 16)
    templates = _TOPIC_ACTIONS.get(topic_id, _TOPIC_ACTIONS[0])
    idx = bucket % len(templates)
    prev = _topic_ctr.get(topic_id, 0)
    if idx == prev and len(templates) > 1: idx = (idx+1) % len(templates)
    _topic_ctr[topic_id] = idx
    return templates[idx]

# ══════════════════════════════════════════════════════════════
# DATABASE  (SQLite — persists in /tmp on Streamlit Cloud per session)
# For production: swap sqlite3 for st.connection("postgresql") or TinyDB
# ══════════════════════════════════════════════════════════════

DB_PATH = "zeus_feedback.db"

SOURCE_PREFIX_MAP = {
    'quora':'Q','reddit':'RE','trustpilot':'TR','yelp':'YL','g2':'G2','capterra':'CP',
    'amazon':'AM','amazon reviews':'AM','tripadvisor':'TA','glassdoor':'GL','indeed':'ID',
    'producthunt':'PH','zomato':'ZO','swiggy':'SG','booking':'BK','booking.com':'BK',
    'csv upload':'CS','csv':'CS',
}
MONTH_MAP = {
    'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,'jul':7,'aug':8,'sep':9,
    'oct':10,'nov':11,'dec':12,'january':1,'february':2,'march':3,'april':4,
    'june':6,'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
}

def get_prefix(src):
    sl = src.lower().strip()
    for k, p in SOURCE_PREFIX_MAP.items():
        if k in sl: return p
    c = re.sub(r'[^a-zA-Z]', '', src).upper()
    return c[:2] if len(c) >= 2 else 'FB'

def assign_ids(df):
    ctr = {}
    try:
        conn = sqlite3.connect(DB_PATH)
        ex = pd.read_sql_query("SELECT feedback_id FROM feedback_entries", conn); conn.close()
        for fid in ex['feedback_id'].dropna():
            m = re.match(r'^([A-Z]+)(\d+)$', str(fid).strip())
            if m:
                p, n = m.group(1), int(m.group(2))
                if ctr.get(p, 0) <= n: ctr[p] = n+1
    except: pass
    ids, local = [], {}
    for _, row in df.iterrows():
        p = get_prefix(row['Source'])
        if p not in local: local[p] = ctr.get(p, 1)
        n = local[p]
        ids.append(f"{p}{str(n).zfill(2) if n < 100 else n}")
        local[p] = n+1
    df = df.copy(); df.insert(0, 'Feedback_ID', ids)
    return df

def extract_exact_date(text):
    if not text or not isinstance(text, str): return ''
    m = re.search(r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})', text)
    if m:
        try:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            if 1990 <= y <= 2100 and 1 <= mo <= 12 and 1 <= d <= 31:
                return datetime(y, mo, d).strftime("%b %d, %Y")
        except: pass
    m = re.search(r'(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\.?\s+(\d{1,2}),?\s+(\d{4})', text, re.I)
    if m:
        try:
            mo = MONTH_MAP[m.group(1).lower()[:3]]; d, y = int(m.group(2)), int(m.group(3))
            return datetime(y, mo, d).strftime("%b %d, %Y")
        except: pass
    m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', text)
    if m:
        try:
            a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            mo, d = (a, b) if a <= 12 else (b, a)
            if 1 <= mo <= 12 and 1 <= d <= 31: return datetime(y, mo, d).strftime("%b %d, %Y")
        except: pass
    return ''

def _normalize_for_hash(text: str) -> str:
    if not text or not isinstance(text, str): return ''
    t = re.sub(r'http\S+|www\S+', '', text.lower())
    t = re.sub(r'[^\w\s]', ' ', t)
    return re.sub(r'\s+', ' ', t).strip()[:500]

def make_content_hash(fb_text: str) -> str:
    return hashlib.sha256(_normalize_for_hash(fb_text).encode('utf-8')).hexdigest()

def init_db():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS sessions(
        session_id TEXT PRIMARY KEY, user_name TEXT, short_uuid TEXT,
        created_date TEXT, created_time TEXT, created_at TEXT,
        source_type TEXT, total_entries INTEGER, notes TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS feedback_entries(
        id INTEGER PRIMARY KEY AUTOINCREMENT, session_id TEXT, entry_uuid TEXT UNIQUE,
        feedback_id TEXT, source TEXT, reviewer_name TEXT, feedback_date TEXT,
        feedback TEXT, topic TEXT, sentiment TEXT, suggestion TEXT, analyzed_at TEXT,
        content_hash TEXT,
        FOREIGN KEY(session_id) REFERENCES sessions(session_id))""")
    for col in ['feedback_id','reviewer_name','feedback_date','sentiment','content_hash']:
        try: c.execute(f"ALTER TABLE feedback_entries ADD COLUMN {col} TEXT")
        except: pass
    try: c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_content_hash ON feedback_entries(content_hash)")
    except: pass
    try:
        missing = conn.execute("SELECT id, feedback FROM feedback_entries WHERE content_hash IS NULL OR content_hash = ''").fetchall()
        for row_id, fb_text in missing:
            c.execute("UPDATE feedback_entries SET content_hash=? WHERE id=?", (make_content_hash(fb_text or ''), row_id))
        if missing: conn.commit()
    except: pass
    conn.commit(); conn.close()

_HASH_CACHE: set = set()
_HASH_CACHE_LOADED: bool = False

def _ensure_cache_loaded():
    global _HASH_CACHE, _HASH_CACHE_LOADED
    if _HASH_CACHE_LOADED: return
    try:
        conn = sqlite3.connect(DB_PATH)
        rows = conn.execute("SELECT content_hash FROM feedback_entries WHERE content_hash IS NOT NULL AND content_hash != ''").fetchall()
        conn.close()
        _HASH_CACHE.update(r[0] for r in rows)
    except: pass
    _HASH_CACHE_LOADED = True

def is_dup(fb_text: str) -> bool:
    h = make_content_hash(fb_text)
    _ensure_cache_loaded()
    if h in _HASH_CACHE: return True
    try:
        conn = sqlite3.connect(DB_PATH)
        row = conn.execute("SELECT 1 FROM feedback_entries WHERE content_hash=? LIMIT 1", (h,)).fetchone()
        conn.close()
        if row: _HASH_CACHE.add(h); return True
    except: pass
    return False

def save_session(sid, uname, uid8, src, total, notes=""):
    now = datetime.now(); conn = sqlite3.connect(DB_PATH)
    conn.execute("INSERT OR REPLACE INTO sessions(session_id,user_name,short_uuid,created_date,created_time,created_at,source_type,total_entries,notes)VALUES(?,?,?,?,?,?,?,?,?)",
        (sid, uname, uid8, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), now.strftime("%Y-%m-%d %H:%M:%S"), src, total, notes))
    conn.commit(); conn.close()

def save_entries(session_id, df):
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    at = datetime.now().strftime("%Y-%m-%d %H:%M:%S"); saved = 0; skipped = 0
    _ensure_cache_loaded()
    for _, row in df.iterrows():
        fb = row.get('Feedback', ''); h = make_content_hash(fb)
        if h in _HASH_CACHE: skipped += 1; continue
        try:
            c.execute("INSERT INTO feedback_entries(session_id,entry_uuid,feedback_id,source,reviewer_name,feedback_date,feedback,topic,sentiment,suggestion,analyzed_at,content_hash)VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (session_id, str(uuid.uuid4()), row.get('Feedback_ID',''), row.get('Source',''), row.get('Reviewer_Name',''),
                 row.get('Feedback_Date',''), fb, row.get('Topic',''), row.get('Sentiment','Neutral'),
                 row.get('Suggestion',''), at, h))
            _HASH_CACHE.add(h); saved += 1
        except sqlite3.IntegrityError: _HASH_CACHE.add(h); skipped += 1
    conn.commit(); conn.close()
    return saved, skipped

def get_all_sessions():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT session_id,user_name,created_date,created_time,source_type,total_entries,notes FROM sessions ORDER BY created_at DESC", conn)
    conn.close(); return df

def get_session_entries(sid):
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT feedback_id,source,reviewer_name,feedback_date,feedback,topic,sentiment,suggestion,analyzed_at FROM feedback_entries WHERE session_id=? ORDER BY id", conn, params=(sid,))
    conn.close(); return df

def get_full_db_export():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("""SELECT f.feedback_id,f.source,f.reviewer_name,f.feedback_date,
        f.feedback,f.topic,f.sentiment,f.suggestion,f.analyzed_at,s.user_name
        FROM sessions s JOIN feedback_entries f ON s.session_id=f.session_id
        ORDER BY s.created_at DESC,f.id""", conn)
    conn.close(); return df

def delete_session(sid):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM feedback_entries WHERE session_id=?", (sid,))
    conn.execute("DELETE FROM sessions WHERE session_id=?", (sid,))
    conn.commit(); conn.close()

init_db()

def gen_sid(uname):
    now = datetime.now(); safe = re.sub(r'[^a-zA-Z0-9]', '', uname.upper())[:12] or "USER"
    uid8 = str(uuid.uuid4()).replace('-', '')[:8].upper()
    return f"{safe}-{now.strftime('%Y%m%d')}-{now.strftime('%H%M%S')}-{uid8}", uid8

# ══════════════════════════════════════════════════════════════
# NLP
# ══════════════════════════════════════════════════════════════

def preprocess(text):
    if not text or not isinstance(text, str): return ''
    text = text.lower(); text = re.sub(r'http\S+|www\S+', '', text)
    text = re.sub(r'[^a-z0-9\s]', '', text)
    try: words = word_tokenize(text)
    except: words = text.split()
    sw = set(stopwords.words('english')); words = [w for w in words if w not in sw and len(w) > 2]
    try: lem = WordNetLemmatizer(); words = [lem.lemmatize(w) for w in words]
    except: pass
    return ' '.join(words) if len(words) >= 2 else ''

def topic_model(texts, n=5):
    if len(texts) < 2: return [0]*len(texts)
    n = min(n, len(texts))
    try:
        vec = TfidfVectorizer(stop_words='english', max_features=500, min_df=1)
        X = vec.fit_transform(texts)
        W = NMF(n_components=n, random_state=42, max_iter=300).fit_transform(X)
        return W.argmax(axis=1).tolist()
    except: return [0]*len(texts)

# ══════════════════════════════════════════════════════════════
# AI PROMPTS
# ══════════════════════════════════════════════════════════════

CSV_SYSTEM = """You are a Chief Customer Experience (CCX) Officer with 25+ years at Fortune 500 companies.
Identify the specific operational root cause for each feedback and prescribe one concrete business fix.
Rules: Name the department/workflow/tool. Max 2 sentences. Each suggestion unique. Executive tone."""

URL_SYSTEM = """You are a CCX Officer analyzing real-time scraped reviews.
For each review: 1) Extract reviewer name (or Anonymous) 2) Write ONE precise operational recommendation.
Rules: Reference what this reviewer experienced. Name the team/process. Max 2 sentences. Vary action verbs."""

def get_suggestions_openai(feedback_list):
    if not openai.api_key: return [""] * len(feedback_list)
    try:
        prompt = "For each feedback, prescribe ONE specific operational action. Max 2 sentences. Be unique per item.\n\nFeedbacks:\n"
        for i, fb in enumerate(feedback_list, 1):
            prompt += f"{i}. {str(fb)[:400]}\n"
        resp = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[{"role":"system","content":CSV_SYSTEM},{"role":"user","content":prompt}],
            max_tokens=3000, temperature=0.85)
        raw = resp["choices"][0]["message"]["content"].strip()
        segments = re.split(r'(?m)^(\d+)[\.):]\s+', raw)
        suggestions_map = {}
        i = 1
        while i < len(segments) - 1:
            try:
                num = int(segments[i]); text = re.sub(r'\s*\n\s*', ' ', segments[i+1].strip()).strip()
                if len(text) > 10: suggestions_map[num] = text
            except: pass
            i += 2
        return [suggestions_map.get(idx, "") for idx in range(1, len(feedback_list)+1)]
    except Exception as e:
        st.warning(f"⚠️ AI suggestion error: {e}")
        return [""] * len(feedback_list)

def get_ai_names_and_suggestions(feedback_list, existing_names=None):
    if existing_names is None: existing_names = [''] * len(feedback_list)
    try:
        prompt = "For each review, extract reviewer name and write ONE operational action.\nFormat: <n>. [NAME: Full Name] Recommendation.\n\nReviews:\n"
        for i, (fb, nm) in enumerate(zip(feedback_list, existing_names), 1):
            hint = f" [Reviewer: {nm}]" if nm and nm.strip() else ""
            prompt += f"{i}. {fb}{hint}\n"
        resp = openai.ChatCompletion.create(
            model="gpt-4o-mini",
            messages=[{"role":"system","content":URL_SYSTEM},{"role":"user","content":prompt}],
            max_tokens=2500, temperature=0.85)
        text = resp["choices"][0]["message"]["content"]
        parsed = {}
        for line in text.strip().split('\n'):
            line = line.strip()
            if not line: continue
            num_m = re.match(r'^(\d+)[\.\)]\s*', line)
            if not num_m: continue
            idx = int(num_m.group(1)) - 1; line = line[num_m.end():]
            name = ""
            nm_m = re.search(r'\[NAME:\s*([^\]]+)\]', line, re.I)
            if nm_m:
                name = nm_m.group(1).strip()
                if name.lower() in ("anonymous","unknown","n/a"): name = ""
                line = line.replace(nm_m.group(0), "").strip()
            if idx not in parsed: parsed[idx] = (name, line.strip())
        return [parsed.get(i, ("", "")) for i in range(len(feedback_list))]
    except: return [("", "")] * len(feedback_list)

def run_ai_analysis(fb_df, mode='full'):
    ai_names, sugs, sents = [], [], []
    existing = [fb_df['Reviewer_Name'].iloc[i] if 'Reviewer_Name' in fb_df.columns and i < len(fb_df) else '' for i in range(len(fb_df))]
    bs = 10; prog = st.progress(0); stat = st.empty(); tb = max(1, (len(fb_df)+bs-1)//bs)
    if openai.api_key:
        for i in range(0, len(fb_df), bs):
            batch = fb_df['Feedback'].iloc[i:i+bs].tolist(); bn = i//bs+1
            stat.markdown(f'<p class="pulse" style="color:#f9a825;font-family:Space Mono,monospace;font-size:0.8rem;">🤖 Processing batch {bn}/{tb}...</p>', unsafe_allow_html=True)
            if mode == 'full':
                for nm, sg in get_ai_names_and_suggestions(batch, existing[i:i+bs]):
                    ai_names.append(nm); sugs.append(sg)
            else:
                for sg in get_suggestions_openai(batch):
                    ai_names.append(''); sugs.append(sg)
            prog.progress(min(0.7, (i+bs)/len(fb_df))); time.sleep(0.3)
    else:
        ai_names = [''] * len(fb_df); sugs = [''] * len(fb_df)
    stat.markdown('<p class="pulse" style="color:#00bcd4;font-family:Space Mono,monospace;font-size:0.8rem;">🎯 Analyzing sentiment...</p>', unsafe_allow_html=True)
    sents = analyze_sentiments_all(fb_df['Feedback'].tolist())
    prog.progress(1.0); stat.empty(); prog.empty()
    return ai_names, sugs, sents

def build_results(fbs, src, names=None, dates=None, mode='full'):
    if names is None: names = [''] * len(fbs)
    if dates is None: dates = [''] * len(fbs)
    valid = [(fb, nm, dt) for fb, nm, dt in zip(fbs, names, dates)
             if fb and str(fb).strip() and not is_junk(str(fb))]
    if not valid: return pd.DataFrame()
    fbs2, names2, dates2 = map(list, zip(*valid))
    df = pd.DataFrame({'Feedback': fbs2, 'Source': src, 'Reviewer_Name': names2, 'Feedback_Date': dates2})
    df['_c'] = df['Feedback'].apply(preprocess)
    df = df[df['_c'].str.strip() != ''].reset_index(drop=True)
    if df.empty: return df
    df['TopicID'] = topic_model(df['_c'].tolist()); df['Topic'] = df['TopicID'].map(TOPIC_LABELS)
    ai_names, sugs, sents = run_ai_analysis(df, mode=mode)
    final_names = []
    for scraped, ai_nm in zip(df['Reviewer_Name'].tolist(), ai_names):
        if scraped and str(scraped).strip() and scraped not in ('nan','None',''): final_names.append(str(scraped).strip())
        elif ai_nm and str(ai_nm).strip(): final_names.append(str(ai_nm).strip())
        else: final_names.append('')
    df['Reviewer_Name'] = final_names
    df['Suggestion'] = [s.strip() if s and isinstance(s,str) and len(s.strip())>20 else generate_fallback_suggestion(f, t)
                        for s, t, f in zip(sugs, df['TopicID'], df['Feedback'])]
    df['Sentiment'] = sents
    return assign_ids(df[['Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion']])

# ══════════════════════════════════════════════════════════════
# SCRAPING ENGINE — Streamlit Cloud compatible (no Selenium)
#
# Strategy per domain:
#   Quora       → /sitemap answer pages + JSON API mining + AMP pages
#   Reddit      → old.reddit.com (static HTML, no JS required)
#   Trustpilot  → paginated static HTML
#   Amazon      → mobile product review pages
#   TripAdvisor → paginated static HTML
#   Zomato      → mobile web + JSON API
#   G2/Capterra → static paginated
#   Glassdoor   → static HTML attempt + clear error if blocked
#   Generic     → BS4 multi-page with rotating UA
# ══════════════════════════════════════════════════════════════

_USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (Linux; Android 13; Pixel 7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Mobile Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/121.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
]

def _get_headers(mobile=False):
    ua = random.choice(_USER_AGENTS)
    if mobile:
        ua = "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
    return {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Referer": "https://www.google.com/",
        "Cache-Control": "no-cache",
        "DNT": "1",
    }

def _fetch(url, timeout=15, mobile=False, retries=2):
    """Robust fetch with retries and rotating UA."""
    session = requests.Session()
    for attempt in range(retries + 1):
        try:
            resp = session.get(url, headers=_get_headers(mobile), timeout=timeout,
                               allow_redirects=True)
            if resp.status_code == 200: return resp
            if resp.status_code == 403: return None  # blocked, don't retry
            if resp.status_code == 429:
                time.sleep(2 ** attempt); continue
        except requests.exceptions.Timeout:
            if attempt == retries: return None
            time.sleep(1)
        except Exception: return None
    return None

# ── Text quality helpers ───────────────────────────────────

_JUNK_RE = re.compile(
    r'^(upvote|share|comment|follow|log\s?in|sign\s?in|sign\s?up|related|sponsored'
    r'|privacy|terms|contact|your\s+ad\s+choices|\u00a9|copyright|page\s+not\s+found'
    r'|404|security\s+service|please\s+enable\s+javascript|cloudflare|access\s+denied'
    r'|you\s+have\s+been\s+blocked|cookies?\s+policy|all\s+rights\s+reserved'
    r'|follow\s+us|subscribe|newsletter|download\s+the\s+app|write\s+a\s+review'
    r'|add\s+a\s+review|loading\.\.\.|please\s+wait|redirecting)', re.I)

_JUNK_PATTERNS = [
    re.compile(r'privacy\s*[\u00b7·]\s*terms', re.I),
    re.compile(r'©\s*\w+.*\d{4}', re.I),
    re.compile(r'security\s+service\s+to\s+protect', re.I),
    re.compile(r'enable\s+javascript', re.I),
    re.compile(r'you\s+have\s+been\s+blocked', re.I),
    re.compile(r'^[\s\u00b7·\-\|,]+$'),
]

def is_junk(text: str) -> bool:
    if not text: return True
    t = text.strip().lower()
    if len(t) < 50: return True
    if _JUNK_RE.search(t): return True
    for pat in _JUNK_PATTERNS:
        if pat.search(text): return True
    alpha = sum(1 for c in t if c.isalpha())
    return len(t) > 0 and alpha / len(t) < 0.4

_TRUNC_RE = re.compile(
    r'(\.{2,}|…|\u2026)\s*$|\(more\)\s*$|\[more\]\s*$|'
    r'(?:read more|see more|show more|view more|continue reading|show full review)\s*$', re.I)

def is_truncated(text: str) -> bool:
    if not text: return False
    s = text.strip()
    if _TRUNC_RE.search(s): return True
    if len(s) > 100 and not re.search(r'[.!?"\'।]$', s): return True
    return False

def clean_text(text: str) -> str:
    if not text or not isinstance(text, str): return ''
    text = re.sub(
        r'\s*[\u2026\.]{2,}\s*$|\s*\(more[\.\u2026]?\)\s*$|\s*\[more\]\s*$|'
        r'\s*(?:read|see|show|view)\s+more\s*$|\s*continue\s+reading\s*$|\s*show\s+full\s+review\s*$',
        '', text.strip(), flags=re.I).strip()
    text = re.sub(r'[\.\u2026]+$', '', text).strip()
    return re.sub(r'\s+', ' ', text).strip()

_BAD_CLS = re.compile(r'nav|menu|header|footer|sidebar|cookie|banner|modal|popup', re.I)

_NAME_STOP = {'nation','india','food','what','how','why','management','engineer','doctor',
              'anonymous','user','guest','customer','reviewer','author','member','profile',
              'review','rating','verified','purchase','buyer','seller','owner','manager'}

def _is_valid_name(name: str) -> bool:
    if not name or not (2 < len(name) < 60): return False
    if re.search(r'http|@|\.(com|net|org)|\d{3,}|[<>{}\[\]]', name, re.I): return False
    if not any(c.isalpha() for c in name): return False
    if name.isupper() and len(name) > 4: return False
    if {w.lower() for w in name.split()} & _NAME_STOP: return False
    return True

def _extract_name(el) -> str:
    from bs4 import Tag
    for attr in ['itemprop']:
        for val in ['reviewerName','reviewer','author','name']:
            found = el.find(True, {attr: val})
            if found and isinstance(found, Tag):
                n = found.get_text(strip=True)
                if _is_valid_name(n): return n
    for a in el.find_all('a', href=True):
        href = a.get('href', '')
        if re.search(r'/user/|/profile/|/u/|/member/', href, re.I):
            n = a.get_text(strip=True)
            if _is_valid_name(n): return n
    for cls_pat in [re.compile(r'reviewer|author|username|display.?name|person.?name', re.I)]:
        found = el.find(True, class_=cls_pat)
        if found and isinstance(found, Tag):
            n = found.get_text(strip=True)
            if _is_valid_name(n): return n
    return ''

def _extract_date(el) -> str:
    t = el.find('time')
    if t:
        d = extract_exact_date(t.get('datetime','') or t.get('data-date','') or t.get_text(strip=True))
        if d: return d
    for e in el.find_all(class_=re.compile(r'date|time|posted|published|timestamp', re.I)):
        d = extract_exact_date(e.get('datetime','') or e.get('title','') or e.get_text(strip=True))
        if d: return d
    return ''

def extract_blocks_generic(soup, min_len=80) -> list:
    """Generic BS4 block extractor — works on any site."""
    seen, blocks = set(), []
    containers = soup.find_all(True,
        class_=re.compile(r'review|comment|feedback|answer|post|testimonial|customer[-_]?review|rating[-_]?item', re.I),
        limit=500)
    def process(text, name='', date=''):
        cl = clean_text(text)
        if is_junk(cl) or len(cl) < min_len: return
        if is_truncated(cl): return  # skip preview snippets
        k = cl[:120].lower()
        if k in seen: return
        seen.add(k)
        if len(cl) > 3000:
            tr = cl[:3000]; lp = max(tr.rfind('.'), tr.rfind('!'), tr.rfind('?'))
            cl = cl[:lp+1] if lp > 600 else tr
        blocks.append({'text': cl, 'name': name, 'date': date})
    for c in containers:
        if _BAD_CLS.search(' '.join(c.get('class') or [])): continue
        raw = c.get_text(' ', strip=True)
        process(raw, _extract_name(c), _extract_date(c))
    if not blocks:
        from bs4 import Tag
        for el in soup.find_all(['div','p','section','article','li','blockquote']):
            if not isinstance(el, Tag): continue
            if _BAD_CLS.search(' '.join(el.get('class') or [])): continue
            if len(el.find_all(['div','section','article'])) > 3: continue
            raw = el.get_text(' ', strip=True)
            if len(raw) < min_len: continue
            bad = any(_BAD_CLS.search(' '.join(p.get('class') or '')) or getattr(p,'name','') in ('header','footer','nav','aside')
                      for p in el.parents if hasattr(p,'get'))
            if bad: continue
            process(raw, date=_extract_date(el))
    return blocks

def json_mine(html_src: str) -> list:
    """Extract long text strings from embedded JSON in page source."""
    results, seen = [], set()
    for m in re.finditer(r'"(?:text|content|body|answer|description|reviewText|fullText)"\s*:\s*"([^"]{150,})"', html_src):
        cand = m.group(1).replace('\\n',' ').replace('\\t',' ').replace('\\"','"')
        cand = re.sub(r'\\u[0-9a-fA-F]{4}', ' ', cand)
        cl = clean_text(cand)
        if is_junk(cl) or is_truncated(cl): continue
        k = cl[:120].lower()
        if k in seen: continue
        seen.add(k); results.append({'text': cl, 'name': '', 'date': ''})
    return results

# ── Domain-specific scrapers ──────────────────────────────

def scrape_quora(url: str):
    """
    Quora scraper — no Selenium required.

    Quora blocks headless browsers on web pages but their AMP pages
    (amp.quora.com) and sitemap/answer pages return full static HTML
    including complete answer text without any collapsing.

    Strategy (tried in order):
    1. AMP URL — amp.quora.com gives full expanded answers as static HTML
    2. JSON-LD mining from the standard page (some answers appear in structured data)
    3. Mobile UA request to www.quora.com (sometimes bypasses collapse)
    4. Direct answer URL if a single /answer/ URL is provided
    5. Search redirect with clear guidance if all fail
    """
    fbs, names, dates = [], [], []; seen = set()

    # Normalise: if /topic/ URL, convert to search
    base_urls = [url]
    if '/topic/' in url:
        slug_m = re.search(r'/topic/([^/?#]+)', url)
        if slug_m:
            slug = slug_m.group(1)
            q = re.sub(r'[-_%]',' ', slug).strip()
            qe = requests.utils.quote(q)
            base_urls = [
                f"https://www.quora.com/search?q={qe}&type=answer",
                url,
            ]

    def _add(text, name='', date=''):
        cl = clean_text(text)
        if is_junk(cl) or len(cl) < 150: return
        if is_truncated(cl): return
        k = cl[:120].lower()
        if k in seen: return
        seen.add(k); fbs.append(cl); names.append(name); dates.append(date)

    for src_url in base_urls:
        # ── Attempt 1: AMP version ──────────────────────────
        # amp.quora.com serves full static HTML with complete answer text
        amp_url = src_url.replace('www.quora.com', 'amp.quora.com').replace('quora.com', 'amp.quora.com')
        if not amp_url.startswith('https://amp.'): amp_url = src_url  # skip if already non-www

        for try_url in [amp_url, src_url]:
            resp = _fetch(try_url, timeout=15, mobile=True)
            if not resp: continue
            soup = BeautifulSoup(resp.text, 'html.parser')

            # Quora AMP: answers in <div class="Answer"> or similar
            # Also try JSON-LD structured data first (most reliable)
            for script in soup.find_all('script', type='application/ld+json'):
                try:
                    data = json.loads(script.string or '')
                    items = data if isinstance(data, list) else [data]
                    for item in items:
                        # Review/QAPage schema
                        for key in ('acceptedAnswer','suggestedAnswer','answer'):
                            ans = item.get(key, [])
                            if isinstance(ans, dict): ans = [ans]
                            for a in ans:
                                txt = a.get('text','') or a.get('description','')
                                if txt and len(txt) > 150:
                                    _add(txt, a.get('author',{}).get('name','') if isinstance(a.get('author'),dict) else '')
                except: pass

            # Generic block extraction — works on AMP pages
            blocks = extract_blocks_generic(soup, min_len=150)
            for b in blocks: _add(b['text'], b['name'], b['date'])

            # JSON mining from page source
            for b in json_mine(resp.text):
                if len(b['text']) >= 150: _add(b['text'])

            if len(fbs) >= 5: break

        # ── Attempt 2: Individual answer pages (if answer URL provided) ──
        if '/answer/' in src_url and len(fbs) < 3:
            resp = _fetch(src_url, timeout=15, mobile=False)
            if resp:
                soup = BeautifulSoup(resp.text, 'html.parser')
                # Full answer text often in a <div> with data attributes
                for el in soup.find_all('div', attrs={'data-testid': True}):
                    raw = el.get_text(' ', strip=True)
                    if len(raw) > 200: _add(raw)
                for b in json_mine(resp.text):
                    if len(b['text']) >= 150: _add(b['text'])

        if len(fbs) >= 5: break

    if not fbs:
        return [], [], [], (
            "⚠️ Quora requires login to view answers on this page.\n\n"
            "**Best options:**\n"
            "1. Copy the answers manually → paste into a CSV → use CSV Upload tab\n"
            "2. Use a direct answer URL: `quora.com/What-is-X/answer/Person-Name`\n"
            "3. Try: `quora.com/search?q=your+topic&type=answer`"
        )
    return fbs[:200], names[:200], dates[:200], None


def scrape_reddit(url: str):
    """Reddit: use old.reddit.com which returns full static HTML."""
    url = url.replace('www.reddit.com','old.reddit.com').replace('reddit.com','old.reddit.com')
    if 'old.reddit.com' not in url: url = url.replace('reddit.com','old.reddit.com')
    fbs, names, dates = [], [], []; seen = set()
    for page in range(1, 4):
        resp = _fetch(url, timeout=12)
        if not resp: break
        soup = BeautifulSoup(resp.text, 'html.parser')
        # old.reddit comments: div.usertext-body > div.md
        for el in soup.select('div.usertext-body div.md'):
            raw = el.get_text(' ', strip=True)
            cl = clean_text(raw)
            if is_junk(cl) or len(cl) < 80: continue
            k = cl[:120].lower()
            if k in seen: continue
            seen.add(k)
            parent = el.find_parent('div', class_=re.compile(r'comment'))
            name = ''
            if parent:
                a = parent.find('a', class_=re.compile(r'author'))
                if a: name = a.get_text(strip=True)
            fbs.append(cl); names.append(name); dates.append(_extract_date(el))
        # Also get post body
        for el in soup.select('div.expando div.md'):
            raw = el.get_text(' ', strip=True)
            cl = clean_text(raw)
            if not is_junk(cl) and len(cl) > 80:
                k = cl[:120].lower()
                if k not in seen: seen.add(k); fbs.append(cl); names.append(''); dates.append('')
        # Next page
        nxt = soup.find('a', rel='next')
        if nxt and nxt.get('href'):
            url = nxt['href'] if nxt['href'].startswith('http') else 'https://old.reddit.com' + nxt['href']
        else: break
        time.sleep(0.5)
    return fbs[:200], names[:200], dates[:200], None if fbs else "⚠️ No content found"


def scrape_trustpilot(url: str):
    """Trustpilot: fully static paginated HTML."""
    base = re.sub(r'[?&]page=\d+', '', url).rstrip('/')
    fbs, names, dates = [], [], []; seen = set()
    for page in range(1, 8):
        page_url = f"{base}?page={page}" if page > 1 else base
        resp = _fetch(page_url, timeout=12)
        if not resp: break
        soup = BeautifulSoup(resp.text, 'html.parser')
        reviews = soup.select('article[data-service-review-business-unit-display-name], div[class*="reviewCard"], section[class*="review"]')
        if not reviews:
            reviews = soup.find_all('article')
        if not reviews: break
        new_found = False
        for r in reviews:
            # Full review text
            body = r.find('p', attrs={'data-service-review-text-typography': True}) or \
                   r.find('p', class_=re.compile(r'typography_body|review-content|reviewText', re.I))
            raw = body.get_text(' ', strip=True) if body else r.get_text(' ', strip=True)
            cl = clean_text(raw)
            if is_junk(cl) or len(cl) < 60: continue
            k = cl[:120].lower()
            if k in seen: continue
            seen.add(k); new_found = True
            name = _extract_name(r); date = _extract_date(r)
            fbs.append(cl); names.append(name); dates.append(date)
        if not new_found: break
        time.sleep(0.4)
    return fbs[:200], names[:200], dates[:200], None if fbs else "⚠️ No reviews found on this Trustpilot page"


def scrape_amazon(url: str):
    """Amazon reviews: use mobile URL + JSON-LD + structured review selectors."""
    # Convert to mobile URL for less JS
    mob_url = url.replace('www.amazon.', 'www.amazon.').replace('/dp/', '/dp/')
    # Amazon review pages
    if '/product-reviews/' not in url and '/dp/' in url:
        asin_m = re.search(r'/dp/([A-Z0-9]{10})', url)
        if asin_m:
            asin = asin_m.group(1)
            domain = re.sub(r'https?://(www\.)?','', url).split('/')[0]
            mob_url = f"https://{domain}/product-reviews/{asin}?reviewerType=all_reviews&pageNumber=1"
    fbs, names, dates = [], [], []; seen = set()
    for page in range(1, 6):
        page_url = re.sub(r'pageNumber=\d+', f'pageNumber={page}', mob_url)
        if 'pageNumber' not in page_url: page_url += f"&pageNumber={page}"
        resp = _fetch(page_url, timeout=12, mobile=True)
        if not resp: break
        soup = BeautifulSoup(resp.text, 'html.parser')
        reviews = soup.select('div[data-hook="review"]')
        if not reviews: break
        new_found = False
        for r in reviews:
            body = r.select_one('span[data-hook="review-body"]')
            if not body: continue
            raw = body.get_text(' ', strip=True)
            cl = clean_text(raw)
            if is_junk(cl) or len(cl) < 60: continue
            k = cl[:120].lower()
            if k in seen: continue
            seen.add(k); new_found = True
            name_el = r.select_one('span.a-profile-name')
            name = name_el.get_text(strip=True) if name_el else ''
            fbs.append(cl); names.append(name); dates.append(_extract_date(r))
        if not new_found: break
        time.sleep(0.4)
    return fbs[:200], names[:200], dates[:200], None if fbs else "⚠️ Amazon review page blocked or no reviews found"


def scrape_tripadvisor(url: str):
    """TripAdvisor: static paginated reviews."""
    base = re.sub(r'-or\d+\.html', '.html', url)
    fbs, names, dates = [], [], []; seen = set()
    for page in range(0, 6):
        offset = page * 10
        if page == 0: page_url = base
        else:
            page_url = re.sub(r'\.html$', f'-or{offset}.html', base)
            if '-or' not in page_url: page_url = base.replace('.html', f'-or{offset}.html')
        resp = _fetch(page_url, timeout=15, mobile=False)
        if not resp: break
        soup = BeautifulSoup(resp.text, 'html.parser')
        # JSON-LD has full review text on TripAdvisor
        for script in soup.find_all('script', type='application/ld+json'):
            try:
                data = json.loads(script.string or '')
                items = data if isinstance(data, list) else [data]
                for item in items:
                    for rev in item.get('review', []) or item.get('reviews', []):
                        body = rev.get('reviewBody','') or rev.get('description','')
                        if body and len(body) > 60:
                            cl = clean_text(body); k = cl[:120].lower()
                            if k not in seen:
                                seen.add(k)
                                author = rev.get('author',{})
                                name = author.get('name','') if isinstance(author,dict) else str(author)
                                fbs.append(cl); names.append(name); dates.append(rev.get('datePublished',''))
            except: pass
        # HTML fallback
        blocks = extract_blocks_generic(soup, min_len=80)
        for b in blocks:
            k = b['text'][:120].lower()
            if k not in seen: seen.add(k); fbs.append(b['text']); names.append(b['name']); dates.append(b['date'])
        if not blocks and not any(s.find('script',type='application/ld+json') for s in [soup]): break
        time.sleep(0.5)
    return fbs[:200], names[:200], dates[:200], None if fbs else "⚠️ TripAdvisor page returned no reviews"


def scrape_zomato(url: str):
    """Zomato: try mobile web + JSON API endpoint."""
    fbs, names, dates = [], [], []; seen = set()
    # Zomato exposes reviews via internal API for restaurant pages
    # Pattern: zomato.com/city/restaurant-name-id/reviews
    res_id_m = re.search(r'\.com/[^/]+/[^/]+-(\d+)(?:/|$)', url)
    if res_id_m:
        res_id = res_id_m.group(1)
        api_url = f"https://www.zomato.com/webroutes/getPage?page_url=/restaurant/{res_id}/reviews&location=&isMobile=1"
        resp = _fetch(api_url, timeout=12, mobile=True)
        if resp:
            try:
                data = resp.json()
                reviews = data.get('page', {}).get('sections', [{}])[0].get('reviewsList', {}).get('reviewsList', [])
                if not reviews:
                    for b in json_mine(resp.text):
                        k = b['text'][:120].lower()
                        if k not in seen: seen.add(k); fbs.append(b['text']); names.append(''); dates.append('')
                for rev in reviews:
                    body = rev.get('reviewText','') or rev.get('text','')
                    if body and len(body) > 60:
                        cl = clean_text(body); k = cl[:120].lower()
                        if k not in seen: seen.add(k); fbs.append(cl); names.append(rev.get('reviewerName','')); dates.append('')
            except: pass
    # Fallback: mobile page
    if not fbs:
        mob = url.replace('www.zomato.com','www.zomato.com') + ('?reviews&page=1' if '?' not in url else '&page=1')
        for page in range(1, 4):
            resp = _fetch(mob, timeout=12, mobile=True)
            if not resp: break
            soup = BeautifulSoup(resp.text, 'html.parser')
            blocks = extract_blocks_generic(soup, 80)
            for b in blocks:
                k = b['text'][:120].lower()
                if k not in seen: seen.add(k); fbs.append(b['text']); names.append(b['name']); dates.append(b['date'])
            for b in json_mine(resp.text):
                k = b['text'][:120].lower()
                if k not in seen: seen.add(k); fbs.append(b['text']); names.append(''); dates.append('')
            if not blocks: break
            time.sleep(0.4)
    return fbs[:200], names[:200], dates[:200], None if fbs else "⚠️ Zomato page returned no reviews (may require login)"


def scrape_glassdoor(url: str):
    """Glassdoor: attempt static fetch with correct headers."""
    fbs, names, dates = [], [], []; seen = set()
    headers = _get_headers()
    headers['Referer'] = 'https://www.glassdoor.com/'
    headers['Cookie'] = 'GSESSIONID=undefined'
    for page in range(1, 4):
        page_url = re.sub(r'_P\d+\.htm', f'_P{page}.htm', url) if '_P' in url else url
        resp = _fetch(page_url, timeout=15)
        if not resp: break
        soup = BeautifulSoup(resp.text, 'html.parser')
        for b in json_mine(resp.text):
            k = b['text'][:120].lower()
            if k not in seen: seen.add(k); fbs.append(b['text']); names.append(''); dates.append('')
        blocks = extract_blocks_generic(soup, 80)
        for b in blocks:
            k = b['text'][:120].lower()
            if k not in seen: seen.add(k); fbs.append(b['text']); names.append(b['name']); dates.append(b['date'])
        if not blocks: break
        time.sleep(0.5)
    err = None if fbs else "⚠️ Glassdoor blocked the request (login required). Copy reviews manually → CSV Upload."
    return fbs[:200], names[:200], dates[:200], err


def scrape_generic(url: str, max_pages=5):
    """Generic multi-page BS4 scraper for any site."""
    domain = re.sub(r'https?://(www\.)?','', url).split('/')[0].lower()
    base = re.sub(r'[?&]page=\d+|[?&]start=\d+', '', url).rstrip('/')
    current = url; fbs, names, dates = [], [], []; seen = set(); err = None
    for page in range(1, max_pages+1):
        resp = _fetch(current, timeout=12)
        if not resp:
            err = "🚫 Site blocked or unreachable"; break
        soup = BeautifulSoup(resp.text, 'html.parser')
        blocks = extract_blocks_generic(soup, 80)
        new = [b for b in blocks if b['text'][:120].lower() not in seen]
        if not new: break
        for b in new:
            seen.add(b['text'][:120].lower())
            fbs.append(b['text']); names.append(b['name']); dates.append(b['date'])
        # Pagination
        nxt = soup.find('a', rel='next') or soup.find('a', class_=re.compile(r'\bnext\b', re.I))
        if nxt and nxt.get('href'):
            h = nxt['href']
            current = h if h.startswith('http') else f"https://{domain}{h}"
        else:
            sep = '&' if '?' in base else '?'
            current = f"{base}{sep}page={page+1}"
        time.sleep(0.4)
    if not fbs and not err:
        # Try JSON mining as last resort
        resp = _fetch(url, timeout=12)
        if resp:
            for b in json_mine(resp.text):
                k = b['text'][:120].lower()
                if k not in seen: seen.add(k); fbs.append(b['text']); names.append(''); dates.append('')
        if not fbs: err = "⚠️ No feedback text found. Site may require login or use heavy JavaScript."
    return fbs[:200], names[:200], dates[:200], err


def scrape_url(url: str):
    """
    Master scraper — routes to the right strategy based on domain.
    All methods use requests + BeautifulSoup only (Streamlit Cloud compatible).
    """
    domain = re.sub(r'https?://(www\.)?','', url).split('/')[0].lower()
    source_map = {
        'quora':'Quora', 'reddit':'Reddit', 'trustpilot':'Trustpilot',
        'yelp':'Yelp', 'g2':'G2', 'capterra':'Capterra',
        'amazon':'Amazon Reviews', 'glassdoor':'Glassdoor', 'indeed':'Indeed',
        'tripadvisor':'TripAdvisor', 'booking':'Booking.com',
        'producthunt':'ProductHunt', 'zomato':'Zomato', 'swiggy':'Swiggy'
    }
    source = next((v for k, v in source_map.items() if k in domain), domain.split('.')[0].capitalize())
    if   'quora'       in domain: fbs,nms,dts,err = scrape_quora(url)
    elif 'reddit'      in domain: fbs,nms,dts,err = scrape_reddit(url)
    elif 'trustpilot'  in domain: fbs,nms,dts,err = scrape_trustpilot(url)
    elif 'amazon'      in domain: fbs,nms,dts,err = scrape_amazon(url)
    elif 'tripadvisor' in domain: fbs,nms,dts,err = scrape_tripadvisor(url)
    elif 'zomato'      in domain: fbs,nms,dts,err = scrape_zomato(url)
    elif 'glassdoor'   in domain: fbs,nms,dts,err = scrape_glassdoor(url)
    elif 'swiggy'      in domain: fbs,nms,dts,err = scrape_generic(url)
    elif 'yelp'        in domain: fbs,nms,dts,err = scrape_generic(url, max_pages=4)
    elif 'g2'          in domain: fbs,nms,dts,err = scrape_generic(url, max_pages=4)
    elif 'capterra'    in domain: fbs,nms,dts,err = scrape_generic(url, max_pages=4)
    else:                         fbs,nms,dts,err = scrape_generic(url)
    return fbs, nms, dts, source, err

# ══════════════════════════════════════════════════════════════
# EXPORTS
# ══════════════════════════════════════════════════════════════

def make_summary_txt(df, session_id, user_name):
    tc = df['Topic'].value_counts(); total = len(df)
    lines = ["╔═══════════════════════════════════════════════╗",
             "║      ZEUS FEEDBACK ANALYZER  —  REPORT        ║",
             "╚═══════════════════════════════════════════════╝",
             f"  Session ID   : {session_id}", f"  Analyst      : {user_name}",
             f"  Generated    : {datetime.now().strftime('%d %B %Y, %H:%M:%S')}",
             f"  Total Entries: {total}",
             "─────────────────────────────────────────────────","","TOPIC DISTRIBUTION","──────────────────"]
    for topic, count in tc.items():
        bar = "█" * int(count/total*30)
        lines.append(f"  {topic:<35} {count:>4} ({round(count/total*100):>2}%)  {bar}")
    if 'Sentiment' in df.columns:
        sc = df['Sentiment'].value_counts()
        lines += ["","SENTIMENT BREAKDOWN","───────────────────"]
        for sent, cnt in sc.items():
            lines.append(f"  {sent:<20} {cnt:>4} ({round(cnt/total*100):>2}%)")
    lines += ["","TOP AI SUGGESTION PER TOPIC","────────────────────────────"]
    for tid, label in TOPIC_LABELS.items():
        sub = df[df['Topic']==label]
        if not sub.empty: lines += [f"\n  {label}", f"  → {sub.iloc[0]['Suggestion']}"]
    lines += ["","","ALL ENTRIES","───────────"]
    for i, row in df.iterrows():
        fid = row.get('Feedback_ID', f"#{i+1}"); name = row.get('Reviewer_Name','')
        date = row.get('Feedback_Date',''); sent = row.get('Sentiment','')
        lines += [f"\n  [{fid}] [{row['Source']}]  {row['Topic']}  [{sent}]",
                  f"  Reviewer   : {name if name else 'Anonymous'}",
                  f"  Date       : {date if date else 'Unknown'}",
                  f"  Feedback   : {row['Feedback']}",
                  f"  Suggestion : {row['Suggestion']}", "  "+"·"*65]
    return '\n'.join(lines)

def make_feedback_excel(df, session_id, user_name):
    from openpyxl import Workbook as WB
    from openpyxl.styles import PatternFill as PF, Font as FT, Alignment as AL, Border as BD, Side as SD
    from openpyxl.utils import get_column_letter as gcl
    from openpyxl.chart import BarChart as BC, Reference as RF
    DARK='1A1A2E'; NAV='0F3460'; GOLD='F9A825'; WHITE='FFFFFF'; LGT='F5F5F5'; GRY='E0E0E0'; DGR='666666'
    TC={'⚡ Service Quality':('F9A825','1A1A2E'),'📦 Product Issues':('E53935','FFFFFF'),
        '💬 Communication':('0288D1','FFFFFF'),'⏱️ Speed & Delays':('F57C00','FFFFFF'),
        '💡 Innovation & Features':('7B1FA2','FFFFFF')}
    SC_EXCEL={'Positive':('2E7D32','FFFFFF'),'Negative':('C62828','FFFFFF'),'Neutral':('455A64','FFFFFF'),'Mixed':('E65100','FFFFFF')}
    def ft(sz=10,b=False,co='111111'): return FT(name='Arial',size=sz,bold=b,color=co)
    def fl(c): return PF('solid',fgColor=c)
    def bd():
        s = SD(style='thin',color=GRY); return BD(left=s,right=s,top=s,bottom=s)
    def al(h='left',v='center',w=False): return AL(horizontal=h,vertical=v,wrap_text=w)
    def gc(cols): return next((c for c in cols if c in df.columns),'')
    idc=gc(['Feedback_ID','feedback_id']); rvc=gc(['Reviewer_Name','reviewer_name'])
    dtc=gc(['Feedback_Date','feedback_date']); fbc=gc(['Feedback','feedback'])
    tpc=gc(['Topic','topic']); sgc=gc(['Suggestion','suggestion']); stc=gc(['Sentiment','sentiment'])
    src=gc(['Source','source'])
    df2 = df.copy()
    if rvc:
        named = df2[df2[rvc].str.strip().ne('')].drop_duplicates(subset=[rvc],keep='first')
        unnamed = df2[df2[rvc].str.strip().eq('')]
        df2 = pd.concat([named,unnamed]).sort_index().reset_index(drop=True)
    src_label = ', '.join(df2[src].unique()[:3]) if src else 'N/A'
    wb = WB(); ws1 = wb.active; ws1.title = 'Feedback Data'; ws1.sheet_view.showGridLines = False
    ws1.merge_cells(start_row=1,start_column=1,end_row=1,end_column=8)
    c = ws1.cell(row=1,column=1,value=f'  ⚡ ZEUS — {src_label}  |  {len(df2)} Entries  |  {datetime.now().strftime("%d %b %Y")}')
    c.font=ft(13,True,GOLD); c.fill=fl(DARK); c.alignment=al('left','center'); ws1.row_dimensions[1].height=32
    ws1.merge_cells(start_row=2,start_column=1,end_row=2,end_column=8)
    c = ws1.cell(row=2,column=1,value=f'  Analyst: {user_name}  ·  Session: {session_id}  ·  Unique entries only')
    c.font=ft(9,False,WHITE); c.fill=fl(NAV); c.alignment=al('left','center'); ws1.row_dimensions[2].height=16
    ws1.row_dimensions[3].height = 4
    HDRS = ['S.No','ID','Name of Feedbacker','Date','Feedback','Type','Sentiment','AI Suggestion']
    for i,w in enumerate([6,9,20,13,55,22,14,55],1): ws1.column_dimensions[gcl(i)].width = w
    for ci,lbl in enumerate(HDRS,1):
        c = ws1.cell(row=4,column=ci,value=lbl)
        c.font=ft(10,True,WHITE); c.fill=fl(DARK); c.border=bd(); c.alignment=al('center','center')
    ws1.row_dimensions[4].height = 22
    for ri,(_,row) in enumerate(df2.iterrows(),start=5):
        alt=(ri%2==0); rbg=LGT if alt else WHITE
        tp=str(row[tpc] if tpc else ''); sent=str(row[stc] if stc else 'Neutral')
        tbg,tfg=TC.get(tp,('888888','FFFFFF')); sbg,sfg=SC_EXCEL.get(sent,('455A64','FFFFFF'))
        sug=str(row[sgc] if sgc else '')
        vals=[ri-4,row[idc] if idc else '',row[rvc] if rvc else '',row[dtc] if dtc else '',row[fbc] if fbc else '',tp,sent,sug]
        styles=[(rbg,DGR,9,False,'center'),(rbg,'0D47A1',9,True,'center'),(rbg,'1A237E',9,True,'center'),
                (rbg,DGR,9,False,'center'),(rbg,'212121',9,False,'left'),(tbg,tfg,8,True,'center'),
                (sbg,sfg,8,True,'center'),('FFFDE7','4E342E',9,False,'left')]
        for ci,(v,(bg,fg,sz,b,h)) in enumerate(zip(vals,styles),1):
            c = ws1.cell(row=ri,column=ci,value=str(v) if v is not None else '')
            c.fill=fl(bg); c.font=ft(sz,b,fg); c.border=bd(); c.alignment=al(h,'top' if h=='left' else 'center',w=(h=='left'))
        fb_len = len(str(row[fbc])) if fbc and row[fbc] else 0
        ws1.row_dimensions[ri].height = max(45, min(409, 45 + (fb_len//80)*14))
    ws1.freeze_panes = 'A5'; ws1.auto_filter.ref = f'A4:{gcl(8)}4'
    # Sentiment sheet
    ws3 = wb.create_sheet('Sentiment Analysis'); ws3.sheet_view.showGridLines = False
    ws3.merge_cells('A1:D1')
    c = ws3.cell(row=1,column=1,value='  🎯 Sentiment Analysis')
    c.font=ft(12,True,GOLD); c.fill=fl(DARK); c.alignment=al('left','center'); ws3.row_dimensions[1].height=28
    for ci,lbl in enumerate(['Sentiment','Count','%','Avg Length'],1):
        c=ws3.cell(row=2,column=ci,value=lbl); c.font=ft(10,True,WHITE); c.fill=fl(DARK); c.border=bd(); c.alignment=al('center','center')
    for i,w in enumerate([16,10,8,12],1): ws3.column_dimensions[gcl(i)].width = w
    if stc and fbc:
        sg2 = df2.groupby(stc).agg(count=(fbc,'count'),avg_len=(fbc,lambda x:int(x.str.len().mean()))).reset_index()
        total_s = sg2['count'].sum()
        for ri,(_,row) in enumerate(sg2.iterrows(),start=3):
            s=str(row[stc]); bg,fg=SC_EXCEL.get(s,('455A64','FFFFFF'))
            for ci,v in enumerate([s,int(row['count']),f"{round(row['count']/total_s*100)}%",row['avg_len']],1):
                c=ws3.cell(row=ri,column=ci,value=v)
                c.fill=fl(bg if ci==1 else ('F5F5F5' if ri%2==0 else 'FFFFFF'))
                c.font=ft(9,ci==1,fg if ci==1 else '212121'); c.border=bd(); c.alignment=al('center','center')
            ws3.row_dimensions[ri].height=22
    # Summary sheet
    ws2 = wb.create_sheet('Summary'); ws2.sheet_view.showGridLines = False
    ws2.merge_cells('A1:E1')
    c=ws2.cell(row=1,column=1,value='  📊 Topic & Sentiment Summary')
    c.font=ft(12,True,GOLD); c.fill=fl(DARK); c.alignment=al('left','center'); ws2.row_dimensions[1].height=28
    for i,w in enumerate([5,28,10,12,60],1): ws2.column_dimensions[gcl(i)].width = w
    for ci,lbl in enumerate(['#','Topic','Count','Sentiment','Top Action'],1):
        c=ws2.cell(row=2,column=ci,value=lbl); c.font=ft(10,True,WHITE); c.fill=fl(DARK); c.border=bd(); c.alignment=al('center','center')
    tc = df2[tpc].value_counts().reset_index() if tpc else pd.DataFrame(columns=['Topic','Count'])
    tc.columns = ['Topic','Count']
    for ri,(_,row) in enumerate(tc.iterrows(),start=3):
        tp=row['Topic']; tbg,tfg=TC.get(tp,('888888','FFFFFF')); alt=(ri%2==0); rbg=LGT if alt else WHITE
        top_sent = df2[df2[tpc]==tp][stc].mode()[0] if (stc and tpc and not df2[df2[tpc]==tp].empty) else ''
        sug = df2[df2[tpc]==tp][sgc].iloc[0] if (tpc and sgc and len(df2[df2[tpc]==tp])>0) else ''
        for ci,(v,bg,fg) in enumerate(zip([ri-2,tp,int(row['Count']),top_sent,sug],[rbg,tbg,rbg,rbg,'FFFDE7'],[DGR,tfg,'1A1A2E',DGR,'4E342E']),1):
            c=ws2.cell(row=ri,column=ci,value=v)
            c.fill=fl(bg); c.font=ft(9,False,fg); c.border=bd(); c.alignment=al('center','center',w=(ci==5))
        ws2.row_dimensions[ri].height=50
    if len(tc) > 1:
        bar=BC(); bar.type='col'; bar.title='Feedback by Topic'; bar.style=10; bar.width=22; bar.height=12
        bar.add_data(RF(ws2,min_col=3,min_row=2,max_row=2+len(tc)),titles_from_data=True)
        bar.set_categories(RF(ws2,min_col=2,min_row=3,max_row=2+len(tc)))
        ws2.add_chart(bar,f'A{len(tc)+5}')
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def make_db_excel():
    from openpyxl import Workbook as WB
    from openpyxl.styles import PatternFill as PF, Font as FT, Alignment as AL, Border as BD, Side as SD
    from openpyxl.utils import get_column_letter as gcl
    df = get_full_db_export()
    if df.empty: return None
    DARK='1A1A2E'; NAV='0F3460'; GOLD='F9A825'; WHITE='FFFFFF'; LGT='F5F5F5'; GRY='E0E0E0'; DGR='666666'
    TC={'⚡ Service Quality':('F9A825','1A1A2E'),'📦 Product Issues':('E53935','FFFFFF'),
        '💬 Communication':('0288D1','FFFFFF'),'⏱️ Speed & Delays':('F57C00','FFFFFF'),
        '💡 Innovation & Features':('7B1FA2','FFFFFF')}
    SC_EXCEL={'Positive':('2E7D32','FFFFFF'),'Negative':('C62828','FFFFFF'),'Neutral':('455A64','FFFFFF'),'Mixed':('E65100','FFFFFF')}
    def ft(sz=10,b=False,co='111111'): return FT(name='Arial',size=sz,bold=b,color=co)
    def fl(c): return PF('solid',fgColor=c)
    def bd():
        s=SD(style='thin',color=GRY); return BD(left=s,right=s,top=s,bottom=s)
    def al(h='left',v='center',w=False): return AL(horizontal=h,vertical=v,wrap_text=w)
    wb=WB(); ws=wb.active; ws.title='All Feedback'; ws.sheet_view.showGridLines=False
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=8)
    c=ws.cell(row=1,column=1,value=f'  🗄️ ZEUS Full Database  |  {len(df)} Entries  |  {datetime.now().strftime("%d %b %Y")}')
    c.font=ft(13,True,GOLD); c.fill=fl(DARK); c.alignment=al('left','center'); ws.row_dimensions[1].height=32
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=8)
    c=ws.cell(row=2,column=1,value='  All sessions combined  ·  Unique IDs  ·  No repeated reviewers')
    c.font=ft(9,False,WHITE); c.fill=fl(NAV); c.alignment=al('left','center'); ws.row_dimensions[2].height=16
    ws.row_dimensions[3].height=4
    HDRS=['S.No','ID','Name of Feedbacker','Date','Feedback','Type','Sentiment','AI Suggestion']
    for i,w in enumerate([6,9,20,13,55,22,14,55],1): ws.column_dimensions[gcl(i)].width=w
    for ci,lbl in enumerate(HDRS,1):
        c=ws.cell(row=4,column=ci,value=lbl); c.font=ft(10,True,WHITE); c.fill=fl(DARK); c.border=bd(); c.alignment=al('center','center')
    ws.row_dimensions[4].height=22
    for ri,(_,row) in enumerate(df.iterrows(),start=5):
        alt=(ri%2==0); rbg=LGT if alt else WHITE
        tp=str(row.get('topic','')); tbg,tfg=TC.get(tp,('888888','FFFFFF'))
        sent=str(row.get('sentiment','Neutral')); sbg,sfg=SC_EXCEL.get(sent,('455A64','FFFFFF'))
        vals=[ri-4,row.get('feedback_id',''),row.get('reviewer_name',''),row.get('feedback_date',''),
              row.get('feedback',''),tp,sent,str(row.get('suggestion',''))]
        styles=[(rbg,DGR,9,False,'center'),(rbg,'0D47A1',9,True,'center'),(rbg,'1A237E',9,True,'center'),
                (rbg,DGR,9,False,'center'),(rbg,'212121',9,False,'left'),(tbg,tfg,8,True,'center'),
                (sbg,sfg,8,True,'center'),('FFFDE7','4E342E',9,False,'left')]
        for ci,(v,(bg,fg,sz,b,h)) in enumerate(zip(vals,styles),1):
            c=ws.cell(row=ri,column=ci,value=str(v) if v is not None else '')
            c.fill=fl(bg); c.font=ft(sz,b,fg); c.border=bd(); c.alignment=al(h,'top' if h=='left' else 'center',w=(h=='left'))
        fb_len = len(str(row.get('feedback','')))
        ws.row_dimensions[ri].height = max(45, min(409, 45+(fb_len//80)*14))
    ws.freeze_panes='A5'; ws.auto_filter.ref=f'A4:{gcl(8)}4'
    ws2=wb.create_sheet('By Source'); ws2.sheet_view.showGridLines=False
    ws2.merge_cells('A1:E1')
    c=ws2.cell(row=1,column=1,value='  📊 Entries by Source')
    c.font=ft(12,True,GOLD); c.fill=fl(DARK); c.alignment=al('left','center'); ws2.row_dimensions[1].height=28
    for ci,lbl in enumerate(['#','Source','Entries','Named','Analyst'],1):
        c=ws2.cell(row=2,column=ci,value=lbl); c.font=ft(10,True,WHITE); c.fill=fl(DARK); c.border=bd(); c.alignment=al('center','center')
    for i,w in enumerate([5,22,10,10,18],1): ws2.column_dimensions[gcl(i)].width=w
    grp=df.groupby('source').agg(entries=('feedback_id','count'),named=('reviewer_name',lambda x:(x.str.strip()!='').sum()),analyst=('user_name','first')).reset_index()
    for ri,(_,row) in enumerate(grp.iterrows(),start=3):
        alt=(ri%2==0); rbg=LGT if alt else WHITE
        for ci,v in enumerate([ri-2,row['source'],int(row['entries']),int(row['named']),row['analyst']],1):
            c=ws2.cell(row=ri,column=ci,value=v); c.fill=fl(rbg); c.font=ft(9,False,DGR); c.border=bd(); c.alignment=al('center','center')
        ws2.row_dimensions[ri].height=20
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

# ══════════════════════════════════════════════════════════════
# SESSION STATE
# ══════════════════════════════════════════════════════════════

for k,v in [('results_df',None),('analyzed',False),('session_id',None),('short_uuid',None),('user_name',''),('source_type','')]:
    if k not in st.session_state: st.session_state[k]=v

# ══════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════

with st.sidebar:
    st.markdown('<div style="font-family:Space Mono,monospace;font-size:0.65rem;letter-spacing:3px;color:#f9a825;text-transform:uppercase;margin-bottom:0.5rem;">⚙ Configuration</div>', unsafe_allow_html=True)
    uname = st.text_input("👤 Your Name", value=st.session_state.user_name, placeholder="e.g. Priya")
    if uname: st.session_state.user_name = uname
    if st.session_state.session_id:
        st.markdown(f'<div class="session-box">🔑 Active Session<br><strong style="font-size:0.65rem;">{st.session_state.session_id}</strong></div>', unsafe_allow_html=True)
    st.divider()
    st.success("✅ OpenAI key loaded") if openai.api_key else st.warning("⚠️ No OPENAI_API_KEY — rule-based sentiment + keyword suggestions active")
    st.divider()
    st.markdown("**🎯 Sentiment Engine**")
    st.markdown('<div style="font-family:Space Mono,monospace;font-size:0.68rem;color:#888;line-height:1.8;">Rule-based → always runs.<br>AI upgrade → uncertain cases only.<br>Positive | Negative | Neutral | Mixed</div>', unsafe_allow_html=True)
    st.divider()
    st.markdown("**🔍 Cloud Scraper**")
    st.markdown('<div style="font-family:Space Mono,monospace;font-size:0.68rem;color:#888;line-height:1.8;">requests + BeautifulSoup only.<br>No Selenium / ChromeDriver.<br>Domain-aware strategies.<br>AMP · JSON API · Mobile UA.<br>Full text — no truncation.</div>', unsafe_allow_html=True)
    st.divider()
    st.markdown("**📌 Supported Sources**")
    for s in ["CSV Upload","Quora (AMP + JSON-LD)","Reddit (old.reddit.com)","Trustpilot","Amazon Reviews","TripAdvisor","Zomato","Glassdoor (best effort)","G2 / Capterra","Any public URL"]:
        st.markdown(f'<div style="font-family:Space Mono,monospace;font-size:0.7rem;color:#aaa;margin:2px 0;">· {s}</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# HERO
# ══════════════════════════════════════════════════════════════

st.markdown('<div class="hero-title">⚡ ZEUS FEEDBACK ANALYZER</div>', unsafe_allow_html=True)
st.markdown('<div class="hero-sub">Customer Intelligence · Topic Modeling · Real Sentiment · Streamlit Cloud Ready</div>', unsafe_allow_html=True)

tab1, tab2, tab3, tab4 = st.tabs(["📄 CSV Upload","🔗 URL Scraper","📊 Results","🗄️ History"])

# ══════════════════════════════════════════════════════════════
# TAB 1 — CSV UPLOAD
# ══════════════════════════════════════════════════════════════

with tab1:
    st.markdown('<div class="card"><div class="card-title">📤 Upload Feedback CSV</div>', unsafe_allow_html=True)
    uploaded = st.file_uploader("Drop CSV here", type=['csv'], label_visibility="collapsed")
    if uploaded:
        try:
            df_raw = pd.read_csv(uploaded); df_raw.columns = df_raw.columns.str.strip()
            st.success(f"✅ **{len(df_raw):,}** rows · Columns: `{'`, `'.join(df_raw.columns)}`")
            text_cols = [c for c in df_raw.columns if df_raw[c].dtype==object]
            default = next((c for c in text_cols if any(k in c.lower() for k in ['feedback','comment','review','text','response'])), text_cols[0] if text_cols else None)
            if not text_cols: st.error("❌ No text columns found.")
            else:
                col_sel = st.selectbox("Feedback column", text_cols, index=text_cols.index(default) if default else 0)
                ac = ['(None)'] + list(df_raw.columns)
                cn, cd = st.columns(2)
                with cn: name_col = st.selectbox("Reviewer Name column (optional)", ac, index=0)
                with cd: date_col = st.selectbox("Date column (optional)", ac, index=0)
                st.dataframe(df_raw[[col_sel]].head(5), use_container_width=True)
                notes = st.text_input("Session notes (optional)", placeholder="e.g. Q2 2025 product feedback")
                if st.button("🚀 Analyze CSV Feedback", key="csv_btn"):
                    nm = st.session_state.user_name or "USER"
                    sid, uid8 = gen_sid(nm); st.session_state.session_id = sid; st.session_state.short_uuid = uid8
                    fbs = df_raw[col_sel].dropna().astype(str).tolist()
                    revs = df_raw[name_col].astype(str).tolist() if name_col!='(None)' else ['']*len(fbs)
                    dts = df_raw[date_col].astype(str).tolist() if date_col!='(None)' else ['']*len(fbs)
                    dts = [extract_exact_date(d) or d for d in dts]
                    with st.spinner("Analyzing..."):
                        results = build_results(fbs, "CSV Upload", revs, dts, mode='suggestion')
                    if results.empty: st.error("❌ No processable feedback found.")
                    else:
                        st.session_state.results_df = results; st.session_state.analyzed = True
                        saved, skipped = save_entries(sid, results)
                        save_session(sid, nm, uid8, "CSV Upload", saved, notes)
                        if 'Sentiment' in results.columns:
                            sc = results['Sentiment'].value_counts()
                            st.info("🎯 Sentiment: " + " · ".join([f"**{s}**: {c}" for s,c in sc.items()]))
                        msg = f"✅ **{len(results):,}** results · **{saved:,}** new saved"
                        if skipped > 0: msg += f" · **{skipped}** skipped (duplicates)"
                        st.success(msg + " → **📊 Results**")
        except Exception as e: st.error(f"Error: {e}")
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# TAB 2 — URL SCRAPER
# ══════════════════════════════════════════════════════════════

with tab2:
    st.markdown('<div class="card"><div class="card-title">🌐 Scrape from URL</div>', unsafe_allow_html=True)
    c1, c2 = st.columns([3,1])
    with c2: multi_mode = st.checkbox("Multi-URL mode")
    url_input = st.text_input("Paste URL", placeholder="https://www.trustpilot.com/review/...", label_visibility="collapsed")
    multi_urls_text = ""
    if multi_mode: multi_urls_text = st.text_area("One URL per line", placeholder="https://...\nhttps://...", height=100, label_visibility="collapsed")
    notes_url = st.text_input("Session notes", placeholder="e.g. Competitor analysis", key="nu")

    # Cloud scraping info box
    st.markdown("""<div style="background:rgba(249,168,37,0.06);border:1px solid rgba(249,168,37,0.25);border-radius:10px;padding:0.8rem 1.2rem;margin-bottom:1rem;">
<div style="font-family:'Space Mono',monospace;font-size:0.68rem;color:#f9a825;letter-spacing:1px;margin-bottom:0.4rem;">ℹ️ CLOUD SCRAPER INFO</div>
<div style="font-family:'Space Mono',monospace;font-size:0.65rem;color:#888;line-height:1.8;">
Quora → AMP pages + JSON-LD (full answer text, no truncation)<br>
Reddit → old.reddit.com static HTML<br>
Trustpilot / TripAdvisor / Amazon → paginated static pages<br>
Zomato → mobile web + internal JSON API<br>
Sites requiring login → use CSV Upload instead
</div></div>""", unsafe_allow_html=True)

    if st.button("🔍 Scrape & Analyze", key="url_btn"):
        urls = []
        if multi_mode and multi_urls_text.strip(): urls = [u.strip() for u in multi_urls_text.strip().split('\n') if u.strip().startswith('http')]
        elif url_input.strip(): urls = [url_input.strip()]
        if not urls: st.warning("⚠️ Enter at least one valid URL.")
        else:
            nm = st.session_state.user_name or "USER"
            sid, uid8 = gen_sid(nm); st.session_state.session_id = sid; st.session_state.short_uuid = uid8
            all_fb, all_n, all_d, all_s = [], [], [], []
            for url in urls:
                lbl = f"`{url[:70]}...`" if len(url)>70 else f"`{url}`"
                with st.status(f"🔄 Scraping {lbl}"):
                    fbs, nms, dts, src, err = scrape_url(url)
                    if err:
                        if '\n' in str(err):
                            st.markdown(f'<div style="background:rgba(233,30,99,0.1);border:1px solid rgba(233,30,99,0.4);border-radius:8px;padding:0.8rem 1rem;font-size:0.82rem;color:#ff8a80;">{err.replace(chr(10),"<br>")}</div>', unsafe_allow_html=True)
                        else: st.error(err)
                    if fbs:
                        st.success(f"✅ **{len(fbs)}** entries from **{src}** · {sum(1 for n in nms if n)} named · {sum(1 for d in dts if d)} dated")
                        all_fb.extend(fbs); all_n.extend(nms); all_d.extend(dts); all_s.extend([src]*len(fbs))
            if all_fb:
                clean_fb, clean_n, clean_d, clean_s = [], [], [], []
                for fb, n, d, s in zip(all_fb, all_n, all_d, all_s):
                    if not is_junk(fb): clean_fb.append(fb); clean_n.append(n); clean_d.append(d); clean_s.append(s)
                junk_rm = len(all_fb) - len(clean_fb)
                if not clean_fb:
                    st.error("⚠️ All scraped content appears to be website UI / error pages. Try CSV Upload.")
                else:
                    info = f"📊 Scraped **{len(all_fb)}** · kept **{len(clean_fb)}** real entries"
                    if junk_rm > 0: info += f" · removed **{junk_rm}** junk"
                    st.info(info + " · Running analysis...")
                    fd = pd.DataFrame({'Feedback':clean_fb,'_s':clean_s,'_n':clean_n,'_d':clean_d})
                    fd['_c'] = fd['Feedback'].apply(preprocess)
                    fd = fd[fd['_c'].str.strip()!=''].reset_index(drop=True)
                    if fd.empty: st.error("❌ No processable feedback after cleaning.")
                    else:
                        fd['TopicID'] = topic_model(fd['_c'].tolist()); fd['Topic'] = fd['TopicID'].map(TOPIC_LABELS)
                        fd = fd.rename(columns={'_n':'Reviewer_Name'})
                        ai_names, sugs, sents = run_ai_analysis(fd, mode='full')
                        final_names = []
                        for scraped, ai_nm in zip(fd['Reviewer_Name'].tolist(), ai_names):
                            if scraped and str(scraped).strip() and scraped not in ('nan','None',''): final_names.append(str(scraped).strip())
                            elif ai_nm and str(ai_nm).strip(): final_names.append(str(ai_nm).strip())
                            else: final_names.append('')
                        fd['Reviewer_Name'] = final_names
                        fd['Suggestion'] = [s.strip() if s and isinstance(s,str) and len(s.strip())>20 else generate_fallback_suggestion(f,t)
                                            for s,t,f in zip(sugs, fd['TopicID'], fd['Feedback'])]
                        fd['Sentiment'] = sents
                        results = fd.rename(columns={'_s':'Source','_d':'Feedback_Date'})[['Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion']]
                        results = assign_ids(results)
                        st.session_state.results_df = results; st.session_state.analyzed = True
                        src_label = ", ".join(dict.fromkeys(clean_s))
                        saved, skipped = save_entries(sid, results); save_session(sid, nm, uid8, src_label, saved, notes_url)
                        if 'Sentiment' in results.columns:
                            sc = results['Sentiment'].value_counts()
                            st.info("🎯 Sentiment: " + " · ".join([f"**{s}**: {c}" for s,c in sc.items()]))
                        msg = f"✅ **{len(results):,}** results · **{saved:,}** new saved"
                        if skipped > 0: msg += f" · **{skipped}** already in DB"
                        st.success(msg + " → **📊 Results**")
            else:
                st.error("No feedback extracted. Try CSV Upload or a different URL.")
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════
# TAB 3 — RESULTS
# ══════════════════════════════════════════════════════════════

with tab3:
    if not st.session_state.analyzed or st.session_state.results_df is None:
        st.markdown("""<div style="text-align:center;padding:4rem;"><div style="font-size:4rem;">📭</div>
<div style="font-family:'Space Mono',monospace;font-size:0.9rem;color:#555;letter-spacing:2px;text-transform:uppercase;margin-top:1rem;">No results yet</div>
<div style="color:#444;font-size:0.85rem;margin-top:0.5rem;">Upload a CSV or scrape a URL first</div></div>""", unsafe_allow_html=True)
    else:
        df = st.session_state.results_df; total = len(df)
        sid = st.session_state.session_id or "—"; uname = st.session_state.user_name or "—"
        tc = df['Topic'].value_counts()
        st.markdown(f'<div class="session-box">🔑 <strong>{sid}</strong> &nbsp;·&nbsp; 👤 {uname} &nbsp;·&nbsp; 📅 {datetime.now().strftime("%d %b %Y, %H:%M")} &nbsp;·&nbsp; 📊 {total} entries</div>', unsafe_allow_html=True)
        # Metrics
        c1,c2,c3,c4,c5,c6 = st.columns(6)
        c1.metric("Total Entries", f"{total:,}"); c2.metric("Topics", df['Topic'].nunique()); c3.metric("Sources", df['Source'].nunique())
        named = df['Reviewer_Name'].astype(str).str.strip().ne('').sum() if 'Reviewer_Name' in df.columns else 0
        c4.metric("Named Reviewers", named)
        pos = df['Sentiment'].eq('Positive').sum() if 'Sentiment' in df.columns else 0
        neg = df['Sentiment'].eq('Negative').sum() if 'Sentiment' in df.columns else 0
        c5.metric("😊 Positive", pos); c6.metric("😞 Negative", neg)
        # Sentiment distribution
        if 'Sentiment' in df.columns:
            st.divider(); st.markdown("### 🎯 Sentiment Distribution")
            sc = df['Sentiment'].value_counts(); sent_cols = st.columns(len(sc))
            SENT_S = {'Positive':('#4caf50','😊'),'Negative':('#e91e63','😞'),'Neutral':('#9e9e9e','😐'),'Mixed':('#ff9800','🤔')}
            for i,(sent,cnt) in enumerate(sc.items()):
                col_h, icon = SENT_S.get(sent,('#888','·'))
                with sent_cols[i]:
                    st.markdown(f"""<div style="background:rgba(255,255,255,0.04);border:1px solid {col_h}40;border-radius:12px;padding:1rem;text-align:center;">
<div style="font-size:1.6rem;">{icon}</div>
<div style="font-size:1.4rem;font-weight:800;color:{col_h};font-family:'Space Mono',monospace;">{cnt}</div>
<div style="font-size:0.72rem;color:#aaa;margin:2px 0;">{sent}</div>
<div style="font-size:0.75rem;color:#555;">{round(cnt/total*100)}%</div></div>""", unsafe_allow_html=True)
        # Topic distribution
        st.divider(); st.markdown("### 🗂️ Topic Distribution")
        cols = st.columns(min(len(tc),5))
        for i,(topic,count) in enumerate(tc.items()):
            with cols[i%5]:
                st.markdown(f"""<div style="background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:1rem;text-align:center;">
<div style="font-size:1.6rem;font-weight:800;color:#f9a825;font-family:'Space Mono',monospace;">{count}</div>
<div style="font-size:0.68rem;color:#aaa;margin:4px 0;">{topic}</div>
<div style="font-size:0.75rem;color:#666;">{round(count/total*100)}%</div></div>""", unsafe_allow_html=True)
        # Filter
        st.divider(); st.markdown("### 🔍 Browse & Filter")
        cf1,cf2,cf3,cf4 = st.columns(4)
        with cf1: tf = st.multiselect("Topic", df['Topic'].unique().tolist(), default=df['Topic'].unique().tolist())
        with cf2: sf = st.multiselect("Source", df['Source'].unique().tolist(), default=df['Source'].unique().tolist())
        with cf3:
            sent_opts = ['All'] + sorted(df['Sentiment'].unique().tolist()) if 'Sentiment' in df.columns else ['All']
            sent_filter = st.selectbox("Sentiment", sent_opts)
        with cf4: search_q = st.text_input("🔎 Search feedback", placeholder="keyword...")
        filtered = df[df['Topic'].isin(tf) & df['Source'].isin(sf)]
        if sent_filter != 'All' and 'Sentiment' in filtered.columns: filtered = filtered[filtered['Sentiment']==sent_filter]
        if search_q.strip(): filtered = filtered[filtered['Feedback'].str.contains(search_q.strip(), case=False, na=False)]
        st.markdown(f'<div style="font-family:Space Mono,monospace;font-size:0.75rem;color:#888;margin-bottom:1rem;">Showing {len(filtered):,} of {total:,}</div>', unsafe_allow_html=True)
        view = st.radio("View", ["📋 Table","🃏 Cards"], horizontal=True)
        if view == "📋 Table":
            dcols = [c for c in ['Feedback_ID','Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion'] if c in filtered.columns]
            st.dataframe(filtered[dcols], use_container_width=True, height=420,
                column_config={
                    "Feedback_ID": st.column_config.TextColumn("ID",width="small"),
                    "Feedback": st.column_config.TextColumn("Feedback",width="large"),
                    "Suggestion": st.column_config.TextColumn("AI Suggestion",width="large"),
                    "Sentiment": st.column_config.TextColumn("Sentiment",width="small"),
                    "Topic": st.column_config.TextColumn("Topic",width="medium"),
                    "Reviewer_Name": st.column_config.TextColumn("Reviewer",width="small"),
                    "Feedback_Date": st.column_config.TextColumn("Date",width="small"),
                })
        else:
            for _, row in filtered.head(30).iterrows():
                tid = [k for k,v in TOPIC_LABELS.items() if v==row['Topic']]
                bc = TOPIC_COLORS.get(tid[0] if tid else 0,"badge-0")
                fid = row.get('Feedback_ID',''); name = row.get('Reviewer_Name',''); date = row.get('Feedback_Date','')
                sent = row.get('Sentiment','')
                SENT_C = {'Positive':'#4caf50','Negative':'#e91e63','Neutral':'#9e9e9e','Mixed':'#ff9800'}
                sc_hex = SENT_C.get(sent,'#9e9e9e')
                SENT_I = {'Positive':'😊','Negative':'😞','Neutral':'😐','Mixed':'🤔'}
                sent_icon = SENT_I.get(sent,'·')
                meta = []
                if name and str(name).strip(): meta.append(f"👤 {name}")
                if date and str(date).strip(): meta.append(f"📅 {date}")
                mh = "  &nbsp;·&nbsp;  ".join(meta)
                fb_full = row['Feedback']
                fb_display = fb_full[:1200]
                fb_overflow = f'<span style="color:#f9a825;font-family:Space Mono,monospace;font-size:0.7rem;"> …[{len(fb_full)-1200} more chars — see CSV/Excel export]</span>' if len(fb_full) > 1200 else ''
                st.markdown(f"""<div class="card">
<div style="display:flex;align-items:center;gap:8px;margin-bottom:0.8rem;flex-wrap:wrap;">
    <span class="feedback-id">{fid}</span>
    <span class="source-tag">{row['Source']}</span>
    <span class="badge {bc}">{row['Topic']}</span>
    {f'<span style="background:rgba(0,0,0,0.3);border:1px solid {sc_hex};border-radius:12px;padding:1px 10px;font-family:Space Mono,monospace;font-size:0.7rem;color:{sc_hex};font-weight:700;">{sent_icon} {sent}</span>' if sent else ''}
    {f'<span style="font-family:Space Mono,monospace;font-size:0.68rem;color:#888;">{mh}</span>' if mh else ''}
</div>
<div style="color:#ddd;font-size:0.9rem;line-height:1.7;margin-bottom:0.8rem;">{fb_display}{fb_overflow}</div>
<div style="border-top:1px solid rgba(255,255,255,0.06);padding-top:0.8rem;">
    <span style="font-family:'Space Mono',monospace;font-size:0.65rem;color:#f9a825;letter-spacing:2px;text-transform:uppercase;">💡 AI Suggestion</span>
    <div style="color:#b0b0b0;font-size:0.85rem;margin-top:4px;">{row.get('Suggestion','')}</div>
</div></div>""", unsafe_allow_html=True)
            if len(filtered) > 30: st.info(f"Showing 30 cards. Switch to Table for all {len(filtered)}.")
        # Downloads
        st.divider(); st.markdown("### 📥 Download")
        d1,d2,d3 = st.columns(3)
        with d1:
            ecols = [c for c in ['Feedback_ID','Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion'] if c in filtered.columns]
            st.download_button("⬇️ CSV — Full Data", data=filtered[ecols].to_csv(index=False).encode('utf-8'),
                file_name=f"zeus_feedback_{sid}.csv", mime='text/csv', use_container_width=True)
        with d2:
            txt = make_summary_txt(filtered, sid, uname)
            st.download_button("⬇️ TXT Summary Report", data=txt.encode('utf-8'),
                file_name=f"zeus_report_{sid}.txt", mime='text/plain', use_container_width=True)
        with d3:
            try:
                xls = make_feedback_excel(filtered, sid, uname)
                st.download_button("⬇️ Excel — Full Report (.xlsx)", data=xls,
                    file_name=f"zeus_feedback_{sid}.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True)
            except Exception as e: st.info(f"`pip install openpyxl` for Excel · {e}")
        st.divider()
        if st.button("🔄 Clear & Start Over"):
            st.session_state.results_df = None; st.session_state.analyzed = False
            st.session_state.session_id = None; st.session_state.short_uuid = None; st.rerun()

# ══════════════════════════════════════════════════════════════
# TAB 4 — HISTORY
# ══════════════════════════════════════════════════════════════

with tab4:
    st.markdown('<div class="card-title" style="font-family:Space Mono,monospace;font-size:0.75rem;letter-spacing:3px;text-transform:uppercase;color:#f9a825;">🗄️ Session History</div>', unsafe_allow_html=True)
    st.info("ℹ️ On Streamlit Cloud, session history persists within the same app instance. For permanent storage across deployments, connect a PostgreSQL database via `st.connection`.", icon="☁️")
    sessions_df = get_all_sessions()
    if sessions_df.empty: st.info("No sessions yet. Run an analysis to start building history.")
    else:
        st.markdown(f"**{len(sessions_df)} session(s)** in database")
        st.dataframe(sessions_df, use_container_width=True, column_config={
            "session_id": st.column_config.TextColumn("Session ID",width="large"),
            "user_name": st.column_config.TextColumn("Analyst"),
            "created_date": st.column_config.TextColumn("Date"),
            "created_time": st.column_config.TextColumn("Time"),
            "source_type": st.column_config.TextColumn("Source"),
            "total_entries": st.column_config.NumberColumn("Entries"),
            "notes": st.column_config.TextColumn("Notes",width="medium"),
        })
        st.divider(); st.markdown("### 🔍 Load & Export Session")
        sel_sid = st.selectbox("Select Session", sessions_df['session_id'].tolist(),
            format_func=lambda x: f"{x}  ·  {sessions_df[sessions_df['session_id']==x]['created_date'].values[0]}  {sessions_df[sessions_df['session_id']==x]['created_time'].values[0]}")
        if sel_sid:
            entries = get_session_entries(sel_sid)
            if not entries.empty:
                meta = sessions_df[sessions_df['session_id']==sel_sid].iloc[0]
                st.markdown(f'<div class="session-box">🔑 <strong>{sel_sid}</strong><br>👤 {meta["user_name"]} &nbsp;·&nbsp; 📅 {meta["created_date"]} {meta["created_time"]} &nbsp;·&nbsp; 📊 {len(entries)} entries &nbsp;·&nbsp; 🌐 {meta["source_type"]}</div>', unsafe_allow_html=True)
                if 'sentiment' in entries.columns:
                    sc_hist = entries['sentiment'].value_counts()
                    st.info("🎯 Sentiment: " + " · ".join([f"**{s}**: {c}" for s,c in sc_hist.items()]))
                st.dataframe(entries[['feedback_id','source','reviewer_name','feedback_date','feedback','topic','sentiment','suggestion','analyzed_at']], use_container_width=True, height=320)
                hist_df = entries.rename(columns={'feedback_id':'Feedback_ID','source':'Source','reviewer_name':'Reviewer_Name','feedback_date':'Feedback_Date','feedback':'Feedback','topic':'Topic','suggestion':'Suggestion','sentiment':'Sentiment'})
                h1,h2,h3 = st.columns(3)
                with h1:
                    ecols = [c for c in ['Feedback_ID','Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion'] if c in hist_df.columns]
                    st.download_button("⬇️ CSV", data=hist_df[ecols].to_csv(index=False).encode('utf-8'),
                        file_name=f"zeus_{sel_sid}.csv", mime='text/csv', use_container_width=True)
                with h2:
                    txt = make_summary_txt(hist_df, sel_sid, meta['user_name'])
                    st.download_button("⬇️ TXT Report", data=txt.encode('utf-8'),
                        file_name=f"zeus_{sel_sid}_report.txt", mime='text/plain', use_container_width=True)
                with h3:
                    try:
                        xls = make_feedback_excel(hist_df, sel_sid, meta['user_name'])
                        st.download_button("⬇️ Excel — Full Report", data=xls,
                            file_name=f"zeus_{sel_sid}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            use_container_width=True)
                    except: st.info("`pip install openpyxl`")
                st.markdown("<br>", unsafe_allow_html=True)
                try:
                    db_xls = make_db_excel()
                    if db_xls: st.download_button("⬇️ Excel — Full Database (.xlsx)", data=db_xls,
                        file_name="zeus_full_database.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        use_container_width=True)
                except: pass
                full_db = get_full_db_export()
                st.download_button("⬇️ CSV — Full Database", data=full_db.to_csv(index=False).encode('utf-8'),
                    file_name="zeus_full_database.csv", mime='text/csv', use_container_width=True)
                st.divider()
                if st.button(f"🗑️ Delete Session `{sel_sid}`"):
                    delete_session(sel_sid); st.success("Session deleted."); st.rerun()
            else: st.warning("No entries found for this session.")
        st.divider()
        with st.expander("⚠️ Danger Zone — Clear All History"):
            st.warning("Permanently deletes ALL sessions and entries from the database.")
            if st.button("🔥 Wipe All History"):
                conn = sqlite3.connect(DB_PATH)
                conn.execute("DELETE FROM feedback_entries"); conn.execute("DELETE FROM sessions")
                conn.commit(); conn.close()
                st.success("All history cleared."); st.rerun()
