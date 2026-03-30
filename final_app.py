import streamlit as st
import pandas as pd
import tempfile
import re, os, sqlite3, uuid, time, io, hashlib, json
from datetime import datetime
# import openai
from dotenv import load_dotenv
import nltk
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import NMF
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import WordNetLemmatizer
from openai import OpenAI
from dotenv import load_dotenv

try:
    import mysql.connector
    MYSQL_AVAILABLE = True
except ImportError:
    MYSQL_AVAILABLE = False

try:
    import psycopg2
    POSTGRES_AVAILABLE = True
except ImportError:
    POSTGRES_AVAILABLE = False

try:
    import snowflake.connector
    SNOWFLAKE_AVAILABLE = True
except ImportError:
    SNOWFLAKE_AVAILABLE = False

try:
    import pymongo
    MONGO_AVAILABLE = True
except ImportError:
    MONGO_AVAILABLE = False

try:
    import pymssql
    MSSQL_AVAILABLE = True
except ImportError:
    MSSQL_AVAILABLE = False

try:
    from google.cloud import bigquery as bq_lib
    BIGQUERY_AVAILABLE = True
except ImportError:
    BIGQUERY_AVAILABLE = False

try:
    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.analytics.data_v1beta.types import RunReportRequest, DateRange, Metric, Dimension
    GA_AVAILABLE = True
except ImportError:
    GA_AVAILABLE = False

try:
    from databricks import sql as dbsql
    DATABRICKS_AVAILABLE = True
except ImportError:
    DATABRICKS_AVAILABLE = False


env_path = r"F:\Zeus AI\zeusfeedback_analizer\.env"
load_dotenv(env_path)

api_key = "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"
client = OpenAI(api_key=api_key)


# st.write("ENV PATH:", env_path)
# st.write("API KEY LOADED:", api_key)


_POS = re.compile(
    r'\b(excellent|amazing|outstanding|wonderful|fantastic|great|good|love|loved|brilliant|superb|perfect|'
    r'delicious|awesome|fabulous|impressive|enjoy|enjoyed|pleasant|satisfied|satisfaction|happy|pleased|'
    r'best|highly recommend|recommend|worth|value for money|value for price|tasty|fresh|'
    r'friendly|helpful|attentive|professional|polite|courteous|efficient|prompt|quick|fast|'
    r'clean|hygienic|comfortable|cozy|ambience|nice place|will visit again|visit again|'
    r'must visit|must try|5 star|five star|10/10|loved it|exceeded expectations|top notch|'
    r'world class|phenomenal|spectacular|mouth.?watering|finger.?licking|well cooked|'
    r'perfectly cooked|good service|great service|good food|great food|amazing food|'
    r'yummy|scrumptious|lip.?smacking)\b', re.I)

_NEG = re.compile(
    r'\b(terrible|horrible|awful|worst|bad|poor|disappointing|disappointed|disgusting|pathetic|'
    r'rude|arrogant|unprofessional|impolite|hostile|aggressive|misbehav|attitude problem|'
    r'tasteless|bland|overcooked|undercooked|half.?cooked|raw|stale|rubbery|dry|cold food|'
    r'slow|delay|wait|waited|too long|overcrowded|understaffed|dirty|unclean|unhygienic|'
    r'overpriced|expensive|rip.?off|not worth|waste of money|never again|avoid|'
    r'do not recommend|not recommend|waste|pest|cockroach|insect|filth|stink|bad smell|'
    r'missing|wrong order|incorrect order|food poisoning|sick|fell ill)\b', re.I)

_MIXED_IND = re.compile(
    r'\bbut\b|\bhowever\b|\bthough\b|\balthough\b|\bexcept\b|\bdespite\b|\byet\b|'
    r'\bon the other hand\b|\bpartially\b|\bnot bad\b|\bok but\b', re.I)

DB_DRIVERS = {
    "MySQL / MariaDB":     ("mysql-connector-python",        MYSQL_AVAILABLE),
    "PostgreSQL":          ("psycopg2-binary",               POSTGRES_AVAILABLE),
    "MongoDB":             ("pymongo",                       MONGO_AVAILABLE),
    "SQLite (file)":       ("built-in",                      True),
    "SQL Server (MSSQL)":  ("pymssql",                       MSSQL_AVAILABLE),
    "Snowflake":           ("snowflake-connector-python",    SNOWFLAKE_AVAILABLE),
    "Google BigQuery":     ("google-cloud-bigquery",         BIGQUERY_AVAILABLE),
    "Google Analytics 4":  ("google-analytics-data",         GA_AVAILABLE),
    "Databricks SQL":      ("databricks-sql-connector",      DATABRICKS_AVAILABLE),
}

ALL_DB_TYPES = list(DB_DRIVERS.keys())

# from zeusfeedback_analizer.app4 import CSV_SYSTEM_PROMPT, CSV_USER_PROMPT_TEMPLATE
# -------------------------------
# SESSION STATE INIT (IMPORTANT)
# -------------------------------
st.set_page_config(page_title="ZEUS FEEDBACK ANALYZER", page_icon="⚡", layout="wide", initial_sidebar_state="expanded")
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@400;600;800&display=swap');
html,body,[class*="css"]{font-family:'Syne',sans-serif;background-color:#0d0d0d;color:#f0ece2;}
#MainMenu,footer,header{visibility:hidden;}
.stApp{background:linear-gradient(135deg,#0d0d0d 0%,#111827 100%);}
.hero-title{font-family:'Syne',sans-serif;font-weight:800;font-size:3.2rem;background:linear-gradient(90deg,#f9a825,#ff6f00,#e91e63);-webkit-background-clip:text;-webkit-text-fill-color:transparent;letter-spacing:-1px;line-height:1.1;margin-bottom:0.2rem;}
.hero-sub{font-family:'Space Mono',monospace;font-size:0.85rem;color:#888;letter-spacing:2px;text-transform:uppercase;margin-bottom:2rem;}
.card{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:16px;padding:1.5rem 2rem;margin-bottom:1.5rem;backdrop-filter:blur(10px);}
.card-title{font-family:'Space Mono',monospace;font-size:0.75rem;letter-spacing:3px;text-transform:uppercase;color:#f9a825;margin-bottom:1rem;}
.badge-0{background:rgba(249,168,37,0.15);border:1px solid #f9a825;color:#f9a825;}
.badge-1{background:rgba(233,30,99,0.15);border:1px solid #e91e63;color:#e91e63;}
.badge-2{background:rgba(0,188,212,0.15);border:1px solid #00bcd4;color:#00bcd4;}
.badge-3{background:rgba(76,175,80,0.15);border:1px solid #4caf50;color:#4caf50;}
.badge-4{background:rgba(156,39,176,0.15);border:1px solid #9c27b0;color:#9c27b0;}
.badge{border-radius:20px;padding:2px 10px;font-size:0.72rem;font-family:'Space Mono',monospace;font-weight:700;}
.source-tag{display:inline-block;background:rgba(249,168,37,0.1);border:1px solid rgba(249,168,37,0.4);border-radius:4px;font-family:'Space Mono',monospace;font-size:0.65rem;padding:2px 8px;color:#f9a825;margin-right:6px;}
.session-box{background:rgba(0,188,212,0.07);border:1px solid rgba(0,188,212,0.25);border-radius:10px;padding:0.8rem 1.2rem;font-family:'Space Mono',monospace;font-size:0.72rem;color:#00bcd4;letter-spacing:0.5px;margin-bottom:1rem;}
.feedback-id{display:inline-block;background:rgba(0,188,212,0.15);border:1px solid rgba(0,188,212,0.5);border-radius:6px;padding:2px 10px;font-family:'Space Mono',monospace;font-size:0.72rem;color:#00bcd4;margin-right:8px;font-weight:700;}
.stTabs [data-baseweb="tab-list"]{gap:8px;border-bottom:1px solid rgba(255,255,255,0.1);}
.stTabs [data-baseweb="tab"]{border-radius:8px 8px 0 0;font-family:'Space Mono',monospace;font-size:0.78rem;letter-spacing:1px;text-transform:uppercase;color:#888;padding:10px 20px;border:none;background:transparent;}
.stTabs [aria-selected="true"]{background:rgba(249,168,37,0.1)!important;color:#f9a825!important;border-bottom:2px solid #f9a825!important;}
.stButton>button{background:linear-gradient(135deg,#f9a825,#ff6f00);color:#0d0d0d;font-family:'Space Mono',monospace;font-weight:700;font-size:0.82rem;letter-spacing:2px;text-transform:uppercase;border:none;border-radius:8px;padding:0.6rem 2rem;width:100%;transition:all 0.2s;}
.stButton>button:hover{transform:translateY(-2px);box-shadow:0 8px 24px rgba(249,168,37,0.4);}
.stTextInput input,.stTextArea textarea{background:rgba(255,255,255,0.05)!important;border:1px solid rgba(255,255,255,0.12)!important;border-radius:8px!important;color:#f0ece2!important;font-family:'Space Mono',monospace!important;font-size:0.85rem!important;}
[data-testid="stSidebar"]{background:rgba(0,0,0,0.4)!important;border-right:1px solid rgba(255,255,255,0.06);}
[data-testid="stMetric"]{background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:1rem;}
[data-testid="stMetricValue"]{color:#f9a825!important;font-family:'Space Mono',monospace!important;}
hr{border-color:rgba(255,255,255,0.08);}
.pulse{animation:pulse 2s infinite;}
@keyframes pulse{0%,100%{opacity:1;}50%{opacity:0.5;}}
</style>
""", unsafe_allow_html=True)

@st.cache_resource
def download_nltk():
    for p in ['stopwords','punkt','wordnet','punkt_tab']:
        nltk.download(p, quiet=True)
download_nltk()

if "user_name" not in st.session_state:
    st.session_state.user_name = ""

if "data" not in st.session_state:
    st.session_state.data = None

if "history" not in st.session_state:
    st.session_state.history = []

DB_PATH = "zeus_feedback.db"
_HASH_CACHE: set = set()
_HASH_CACHE_LOADED: bool = False

SOURCE_PREFIX_MAP = {
    'quora': 'Q', 'reddit': 'RE', 'trustpilot': 'TR', 'yelp': 'YL',
    'g2': 'G2', 'capterra': 'CP', 'amazon': 'AM', 'amazon reviews': 'AM',
    'tripadvisor': 'TA', 'glassdoor': 'GL', 'indeed': 'ID',
    'producthunt': 'PH', 'zomato': 'ZO', 'swiggy': 'SG',
    'booking': 'BK', 'booking.com': 'BK', 'csv upload': 'CS', 'csv': 'CS',
}
CSV_SYSTEM_PROMPT = """
You are a Chief Customer Experience (CCX) Officer with 25+ years leading CX strategy for Fortune 500 companies.

You are analyzing structured feedback data uploaded from a CSV dataset.
Your job is to identify the specific operational root cause for each piece of feedback and prescribe one concrete business fix.

Rules for CSV analysis:
• Focus on the PATTERN hidden in the feedback — what process, team, or system is failing?
• Name the specific department, workflow, tool, or metric that needs to change
• Each suggestion must be unique and tailored to that specific feedback item
• Maximum 2 sentences per suggestion
• Avoid repeating the same action verb across consecutive suggestions
• Write as a strategic advisor, not a customer service agent

Tone: Executive. Tactical. Data-driven. No filler.
"""

CSV_USER_PROMPT_TEMPLATE = """
You are analyzing CSV-uploaded customer feedback for strategic CX improvement.

For each feedback entry below, prescribe ONE specific operational action.

Each action must:
- Name the team, system, or process responsible
- Specify the exact change or intervention needed
- Be unique and different from other suggestions
- Maximum 2 sentences

Customer Feedback (CSV dataset):
"""


MONTH_MAP = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
    'january': 1, 'february': 2, 'march': 3, 'april': 4, 'june': 6,
    'july': 7, 'august': 8, 'september': 9, 'october': 10,
    'november': 11, 'december': 12,
}
_MORE=re.compile(
    r'[\.\u2026]*\s*(?:\(more[\.\u2026]?\)|\[more\]|\.{2,}more|\u2026more|'
    r'read\s+more|see\s+more|show\s+more|view\s+more|continue\s+reading|show\s+full\s+review|expand)',
    re.I)
_BAD_CLS=re.compile(r'nav|menu|header|footer|sidebar|cookie|banner',re.I)

_JUNK_RE=re.compile(
    r'^(upvote|share|comment|follow|log\s?in|sign\s?in|sign\s?up|related|sponsored'
    r'|privacy|terms|contact|languages|your\s+ad\s+choices|press|\u00a9|copyright'
    r'|page\s+not\s+found|404|we\s+searched\s+everywhere|couldn\'t\s+find\s+the\s+page'
    r'|this\s+website\s+uses\s+a\s+security\s+service|security\s+service\s+to\s+protect'
    r'|please\s+enable\s+javascript|javascript\s+is\s+disabled|cloudflare'
    r'|access\s+denied|you\s+have\s+been\s+blocked|ddos\s+protection'
    r'|cookies?\s+policy|cookie\s+consent|accept\s+cookies'
    r'|all\s+rights\s+reserved|inc\.\s*\d{4}|llc\.\s*\d{4}'
    r'|follow\s+us|subscribe\s+to|newsletter|download\s+the\s+app'
    r'|write\s+a\s+review|add\s+a\s+review|submit\s+feedback'
    r'|loading\.\.\.|please\s+wait|redirecting)',
    re.I
)

_JUNK_PATTERNS = [
    re.compile(r'privacy\s*[\u00b7\·]\s*terms\s*[\u00b7\·]', re.I),
    re.compile(r'©\s*\w+.*\d{4}', re.I),
    re.compile(r'security\s+service\s+to\s+protect', re.I),
    re.compile(r'page\s+not\s+found', re.I),
    re.compile(r'enable\s+javascript', re.I),
    re.compile(r'you\s+have\s+been\s+blocked', re.I),
    re.compile(r'^[\s\u00b7\·\-\|,]+$'),
    re.compile(r'(privacy|terms|contact|languages|press)\s*[\u00b7\·].*[\u00b7\·]', re.I),
]
URL_USER_PROMPT_TEMPLATE = """
Analyze each scraped web review below. For every item:
- Extract the reviewer name (or Anonymous)
- Identify sentiment: Positive | Negative | Neutral | Mixed
- Write ONE specific operational action the business must take based on this exact review

Format EXACTLY:
<number>. [NAME: Full Name] [Sentiment] Your recommendation here.

Scraped Reviews:
"""
URL_SYSTEM_PROMPT = """
You are a world-class Chief Customer Experience (CCX) Officer analyzing real-time scraped reviews from live platforms like Trustpilot, Yelp, Google, Reddit, Amazon, TripAdvisor, and similar.

These are authentic customer voices captured directly from the web — treat each review as a live signal.

Your expertise:
• Extracting reviewer identity and sentiment from unstructured web text
• Mapping real complaints to operational failures in specific business units
• Turning a single review into an actionable CX directive

For each review:
1. Extract reviewer name if present anywhere in the text (or write: Anonymous)
2. Classify sentiment: Positive | Negative | Neutral | Mixed
3. Write ONE precise operational recommendation based on the specific context of this review

Rules:
• Each suggestion must directly reference what this specific reviewer experienced
• Mention the team, process, or channel that failed or excelled
• Maximum 2 sentences
• Vary your action verbs and recommended interventions — no two suggestions should sound identical
• Write like a CX executive advising a board, not a customer support bot

Tone: Direct. Evidence-based. Boardroom-ready.
"""

_FALLBACK_KEYWORD_ACTIONS = [
    # Complaint-specific patterns — require negative context, not just topic words
    (r'made.*wait|waited.*long|took.*too long|long.*queue|slow.*service|delay.*order|waited.*hour|no.*refill|refill.*wait',
     "Audit the service timeline at peak hours, assign a dedicated expediter role during rush periods, and set a 10-minute maximum wait target with manager alerts when breached."),
    (r'rude|behaviour|attitude|unprofessional|hostile|impolite|arrogant|disrespectful|misbehav',
     "Launch a monthly frontline staff empathy workshop, introduce peer-review scorecards, and tie guest-satisfaction ratings directly to service team performance incentives."),
    (r'tasteless|no taste|bland|overcooked|undercooked|half.?cooked|raw.*meat|cold.*food|stale|rubbery|insipid|dry.*chicken|dry.*meat',
     "Establish a kitchen quality-control checkpoint: assign a dedicated quality officer per shift to inspect every dish before it leaves the kitchen, with a discard protocol for sub-standard items."),
    (r'overpriced|not worth|rip.?off|too expensive|poor.*value|price.*hike|price.*increase|price.*went up|not.*worth.*money',
     "Introduce a transparent value communication strategy — highlight premium ingredients on menus, publish cost-per-portion data for high-value items, and launch a loyalty tier with exclusive repeat-visitor pricing."),
    (r'dirty|unhygienic|unclean|pest|cockroach|insect|filth|stink|smell.*bad|bad.*smell|not.*clean',
     "Schedule bi-weekly hygiene audits with a third-party inspector, post the latest inspection score at the entrance, and create an anonymous staff reporting channel for cleanliness violations."),
    (r'less variety|limited.*option|same.*menu|menu.*not.*change|fewer.*item|not enough.*veg|lack.*choice|no.*option',
     "Conduct a quarterly menu refresh, guarantee a minimum of 6 rotating seasonal items, and introduce a dedicated vegetarian section with at least 8 distinct options."),
    (r'too.*noisy|very.*loud|overcrowd|uncomfortable.*seat|dirty.*floor|broken|torn.*seat|ambience.*bad|bad.*ambience',
     "Commission an acoustic and ambience assessment, install sound-dampening panels in high-density zones, and establish a monthly décor inspection checklist for all outlets."),
    (r'no.*table|could not.*seat|reservation.*not.*honour|lost.*booking|waited.*despite.*booking|table.*not.*ready',
     "Implement a real-time table management system with automated SMS confirmation 30 minutes before arrival, and introduce a 15-minute hold policy to eliminate booking-loss failures."),
    (r'dessert.*finish|no.*dessert|sweet.*not.*available|ice.?cream.*over|cake.*finish|dessert.*empty|dessert.*ran out',
     "Assign a dedicated dessert replenishment attendant during peak hours with a 5-minute restocking SLA, and add 3 rotating regional dessert specials monthly."),
    (r'starter.*finish|no.*starter|grill.*empty|waiting.*for.*starter|prawn.*not.*available|mutton.*finish|chicken.*finish|item.*not.*available',
     "Install a live grill-station display showing estimated replenishment time per item and empower grill staff to proactively alert tables when a popular item is 5 minutes away."),
    (r'discount.*not.*applied|coupon.*not.*work|offer.*reject|promo.*not.*accept|zomato.*discount.*fail|swiggy.*offer.*fail',
     "Audit third-party discount redemption workflows monthly, ensure POS systems auto-validate all active promotions, and train billing staff on same-day coupon escalation procedures."),
    (r'birthday.*ruin|anniversary.*disappoint|event.*mess|party.*not.*organis|celebration.*bad|occasion.*ruin|staff.*forgot.*occasion',
     "Create a dedicated celebrations concierge role — confirming dietary needs, arranging décor, and briefing floor staff 30 minutes before the party arrives."),
    (r'app.*crash|app.*not.*work|website.*down|online.*order.*fail|delivery.*wrong|order.*missing|payment.*fail',
     "Conduct a quarterly UX audit of the ordering platform, prioritise the top 3 friction points in user drop-off analytics, and commit to a 2-week fix cycle for critical bugs."),
    (r'hard.*find|difficult.*locate|no.*parking|parking.*issue|far.*away|location.*bad|difficult.*reach',
     "Update Google Maps and Zomato location pins with precise coordinates and real-time parking info; introduce a weekend valet service to eliminate the parking friction reported by repeat visitors."),
    (r'no.*staff|understaffed|no one came|staff.*shortage|ignored.*table|unattended.*table|waited.*no.*one.*came',
     "Implement a table-to-staff ratio cap (max 1:4 during peak hours), deploy a digital floor-management tool to flag unattended tables after 3 minutes, and maintain a trained part-time staff pool for weekends."),
    (r'love|amazing|excellent|outstanding|wonderful|fantastic|highly recommend|best.*place|worth.*visit|thoroughly enjoyed',
     "Capitalise on this positive sentiment by launching a referral programme — offer a complimentary dessert for every first-visit guest brought by a verified loyalist, tracked through the app."),
    (r'better than|worse than|compared to|absolute barbeque|competition|competitor|vs barbeque|barbeque vs',
     "Commission a quarterly competitive benchmark study covering food quality, price, ambience, and service speed against the top 3 competitors in each city, and share findings with ops and product teams."),
    (r'ipo|stock.*price|share.*price|invest|nse|bse|rakesh jhunjhunwala|market cap|equity|revenue.*crore|financials',
     "Strengthen investor relations by publishing a quarterly CX metrics report alongside financial results — showing NPS trend, complaint resolution rate, and repeat-visit ratio to build long-term stakeholder confidence."),
    (r'smoke|smok|wood.*fire|charcoal|pit.*bbq|authentic.*bbq|real.*barbecue|hardwood|mesquite|hickory',
     "Introduce a Smoke Story menu callout explaining the wood type, temperature, and smoking duration for each grilled item — building perceived authenticity and differentiating from competitors who use gas grills."),
    (r'tip|strategy|trick|hack|maximize|recover.*money|get.*worth|eat.*more|buffet.*tip|save.*money.*buffet',
     "Develop a Diner Guide card placed on every table with recommended order sequencing and pairing tips — turning the buffet experience into a curated journey and reducing the filling-up-on-filler pattern."),
    (r'loyalt|loyalty.*card|membership|reward|point.*system|regulars|return.*customer|repeat.*visit',
     "Redesign the loyalty programme with a tiered structure (Bronze → Silver → Gold) visible in the app, where Gold members receive a dedicated host, priority seating, and a monthly personalised offer based on past orders."),
]
_FALLBACK_TOPIC_ACTIONS = {
    0: [  # Service Quality — 5 rotating templates
        "Introduce a real-time floor-monitoring dashboard for managers, flagging tables with no staff contact beyond 4 minutes, and tie resolution speed to shift-level performance reviews.",
        "Deploy a post-meal digital feedback kiosk at exit points and route sub-4-star ratings directly to the duty manager's phone within 60 seconds for immediate recovery.",
        "Establish a mystery dining programme — quarterly audits by trained evaluators — to surface service gaps that are invisible to in-house management.",
        "Create a 'Service Champion' recognition system where floor staff earn points for positive guest mentions, redeemable monthly, reinforcing a culture of proactive hospitality.",
        "Map the guest journey from entry to exit and identify the three highest-friction touchpoints; redesign those micro-moments in a 30-day sprint with frontline staff input.",
    ],
    1: [  # Product Issues — 5 rotating templates
        "Implement a live product quality log where kitchen staff flag any batch that deviates from standard, triggering an automatic re-prep before the item reaches the buffet.",
        "Introduce a guest-facing 'freshness timer' display at each buffet station showing when each dish was last replenished, building trust and driving kitchen accountability.",
        "Set up a weekly tasting panel — including non-chef staff — to score each dish against a standardised flavour benchmark, with mandatory rework for anything scoring below 7/10.",
        "Conduct a root-cause analysis on the top 3 most-complained-about dishes this quarter and pilot improved recipes at one outlet before rolling out chain-wide.",
        "Create a short feedback card at each table specifically rating food quality, aggregate scores weekly, and share the bottom-ranked items with the executive chef for corrective action.",
    ],
    2: [  # Communication — 5 rotating templates
        "Train all floor staff on a standard 3-step proactive communication script — greet within 90 seconds, present the menu with one recommendation, check back after the first course.",
        "Introduce a pre-visit automated WhatsApp message confirming reservations, sharing the day's special, and inviting dietary preferences — turning communication into a personalisation tool.",
        "Set up a post-visit email sequence: a thank-you within 2 hours, a feedback survey within 24 hours, and a personalised offer within 7 days to drive return visits.",
        "Create a single-page laminated 'Guest FAQ' on each table covering top 10 common questions (billing, allergens, refills), reducing staff interruptions and improving guest confidence.",
        "Implement a structured complaint-handling protocol: acknowledge within 1 minute, offer a concrete resolution within 3 minutes, and follow up with the guest before they leave.",
    ],
    3: [  # Speed & Delays — 5 rotating templates
        "Introduce staggered entry slots (every 15 minutes) for large group reservations to distribute kitchen load, reducing peak-hour bottlenecks by an estimated 30%.",
        "Install a kitchen display system (KDS) linking front-of-house seating data to prep workloads in real time, enabling the kitchen to front-load popular items before rush periods.",
        "Set a maximum 8-minute starter replenishment SLA during peak hours and assign one dedicated runner per grill station whose sole role is keeping buffet counters filled.",
        "Analyse historical reservation data to identify the busiest 2-hour window weekly and schedule 20% additional staff specifically for that window, pre-approved as standing overtime.",
        "Pilot a 'Fast Lane' table category for guests with 60-minute dining windows — pre-set with starters already grilled, ensuring the first course is ready upon seating.",
    ],
    4: [  # Innovation & Features — 5 rotating templates
        "Launch a 'Guest Chef' quarterly event where top-rated customer recipe suggestions are cooked live at the outlet, generating organic social content and deepening community loyalty.",
        "Introduce a digital loyalty passport — every visit earns stamps redeemable for exclusive menu access or table upgrades, with tier levels (Bronze → Gold) visible in the app.",
        "Create a seasonal limited-edition menu released every 90 days, promoted exclusively to loyalty members 48 hours early, driving repeat visits and social buzz.",
        "Set up a live innovation lab at one flagship outlet where new dishes are beta-tested with willing diners who receive a discount in exchange for structured feedback.",
        "Partner with a local culinary school for a rotating 'Student Signature Dish' feature — fresh ideas, brand goodwill, and a consistent stream of menu innovation at low cost.",
    ],
}

TOPIC_LABELS = {
    0: "⚡ Service Quality",
    1: "📦 Product Issues",
    2: "💬 Communication",
    3: "⏱️ Speed & Delays",
    4: "💡 Innovation & Features",
}
TOPIC_COLORS = {
    0: "badge-0",
    1: "badge-1",
    2: "badge-2",
    3: "badge-3",
    4: "badge-4",
}


FEEDBACK_COL_EXACT = [
    "feedback","review","comment","text","body","description","message","opinion",
    "response","remark","note","testimonial","content","reply","remarks","review_text",
    "comment_text","feedback_text","user_comment","customer_feedback","customer_review",
    "review_body","comment_body","post_text","tweet","post","answer","complaint",
    "suggestion","verbatim","open_ended","narrative","user_feedback","product_review",
    "service_review","user_review","review_description","feedback_description",
    "customer_comment","client_feedback","patient_feedback","employee_feedback",
    "survey_response","survey_text","survey_answer","nps_comment","nps_feedback",
    "csat_comment","open_text","free_text","freetext","qualitative","written_feedback",
    "written_review","text_response","long_text","details","additional_comments",
    "additional_feedback","comments","reviews","feedbacks","responses",
]
FEEDBACK_COL_PARTIAL = [
    "review","feedback","comment","opinion","remark","suggest","complain",
    "response","testimonial","verbatim","narrative","text","body","description",
    "message","reply","content","note","answer","complaint","survey",
]
NAME_COL_EXACT = [
    "name","author","reviewer","username","user","customer","user_name","author_name",
    "reviewer_name","customer_name","full_name","display_name","posted_by","written_by",
    "reviewer_id","customer_id","client_name","patient_name","employee_name",
    "respondent","submitter","first_name","screen_name","handle",
]
DATE_COL_EXACT = [
    "date","created_at","posted_at","timestamp","review_date","feedback_date",
    "created_date","submission_date","updated_at","time","when","datetime",
    "post_date","review_time","submitted_at","recorded_at","entry_date",
    "response_date","date_created","date_posted","date_submitted","date_reviewed",
]

_STRICT_DETECT_SYSTEM = (
    "You are a strict and highly accurate data schema analyst. "
    "Your job is to analyze a dataset schema and identify which column contains CUSTOMER FEEDBACK. "
    "Return ONLY valid JSON — no markdown, no explanation."
)
_FEEDBACK_DEFINITION = """\
A feedback column contains human-written natural language opinions, complaints, reviews, suggestions.
NOT feedback: IDs, ratings, status fields, emails, URLs, categories.
OUTPUT FORMAT (return ONLY this JSON):
{"feedback_col":"<col or null>","name_col":"<col or null>","date_col":"<col or null>",
 "confidence_score":<0-100>,"reason":"<one sentence>","column_scores":{"<col>":<0-100>}}"""

_fallback_topic_counters = {k: 0 for k in _FALLBACK_TOPIC_ACTIONS}
def init_db():
    conn = sqlite3.connect(DB_PATH); c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS sessions(
        session_id TEXT PRIMARY KEY, user_name TEXT, short_uuid TEXT,
        created_date TEXT, created_time TEXT, created_at TEXT,
        source_type TEXT, total_entries INTEGER, notes TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS feedback_entries(
        id INTEGER PRIMARY KEY AUTOINCREMENT, session_id TEXT, entry_uuid TEXT UNIQUE,
        feedback_id TEXT, source TEXT, reviewer_name TEXT, feedback_date TEXT,
        feedback TEXT, topic TEXT, sentiment TEXT, suggestion TEXT, analyzed_at TEXT,
        content_hash TEXT,
        FOREIGN KEY(session_id) REFERENCES sessions(session_id))""")
    for col in ['feedback_id','reviewer_name','feedback_date','sentiment','content_hash']:
        try: c.execute(f"ALTER TABLE feedback_entries ADD COLUMN {col} TEXT")
        except: pass
    try: c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_content_hash ON feedback_entries(content_hash)")
    except: pass
    try:
        missing = conn.execute("SELECT id, feedback FROM feedback_entries WHERE content_hash IS NULL OR content_hash = ''").fetchall()
        for row_id, fb_text in missing:
            c.execute("UPDATE feedback_entries SET content_hash=? WHERE id=?", (make_content_hash(fb_text or ''), row_id))
        if missing: conn.commit()
    except: pass
    conn.commit(); conn.close()

_HASH_CACHE: set = set()
_HASH_CACHE_LOADED: bool = False

def get_session_entries(sid):
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("SELECT feedback_id,source,reviewer_name,feedback_date,feedback,topic,sentiment,suggestion,analyzed_at FROM feedback_entries WHERE session_id=? ORDER BY id ASC", conn, params=(sid,))
    conn.close(); return df

def get_full_db_export():
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("""SELECT f.feedback_id,f.source,f.reviewer_name,f.feedback_date,
        f.feedback,f.topic,f.sentiment,f.suggestion,f.analyzed_at,s.user_name
        FROM sessions s JOIN feedback_entries f ON s.session_id=f.session_id
        ORDER BY s.created_at DESC, f.id ASC""", conn)
    conn.close(); return df

def delete_session(sid):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM feedback_entries WHERE session_id=?", (sid,))
    conn.execute("DELETE FROM sessions WHERE session_id=?", (sid,))
    conn.commit(); conn.close()

init_db()

def render_step_bar(current_step, steps):
    html = '<div class="step-bar">'
    
    for i, label in enumerate(steps, 1):
        if i < current_step:
            css, prefix = "step-item step-done", "✓ "
        elif i == current_step:
            css, prefix = "step-item step-active", ""
        else:
            css, prefix = "step-item step-pending", ""

        html += f'<div class="{css}">{prefix}{i}. {label}</div>'

    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)

def _fetch_sqlite(db_path, query):
    conn = sqlite3.connect(db_path)
    try: return pd.read_sql_query(query, conn)
    finally: conn.close()

def _list_sqlite_tables(db_path):
    conn = sqlite3.connect(db_path)
    tbls = pd.read_sql_query("SELECT name FROM sqlite_master WHERE type='table'", conn)["name"].tolist()
    conn.close(); return tbls

def _fetch_mysql(host, port, user, password, database, query):
    if not MYSQL_AVAILABLE: raise ImportError("pip install mysql-connector-python")
    conn = mysql.connector.connect(host=host, port=int(port), user=user, password=password, database=database)
    try: return pd.read_sql_query(query, conn)
    finally: conn.close()

def _list_mysql_tables(host, port, user, password, database):
    if not MYSQL_AVAILABLE: return []
    conn = mysql.connector.connect(host=host, port=int(port), user=user, password=password, database=database)
    cur = conn.cursor(); cur.execute("SHOW TABLES")
    tbls = [r[0] for r in cur.fetchall()]; conn.close(); return tbls

def _fetch_postgres(host, port, user, password, database, query):
    if not POSTGRES_AVAILABLE: raise ImportError("pip install psycopg2-binary")
    conn = psycopg2.connect(host=host, port=int(port), user=user, password=password, dbname=database)
    try: return pd.read_sql_query(query, conn)
    finally: conn.close()

def _list_postgres_tables(host, port, user, password, database):
    if not POSTGRES_AVAILABLE: return []
    conn = psycopg2.connect(host=host, port=int(port), user=user, password=password, dbname=database)
    df = pd.read_sql_query("SELECT table_name FROM information_schema.tables WHERE table_schema='public'", conn)
    conn.close(); return df['table_name'].tolist()

def _fetch_snowflake(account, user, password, warehouse, database, schema, query):
    if not SNOWFLAKE_AVAILABLE: raise ImportError("pip install snowflake-connector-python")
    conn = snowflake.connector.connect(account=account, user=user, password=password, warehouse=warehouse, database=database, schema=schema)
    try: return pd.read_sql_query(query, conn)
    finally: conn.close()

def _list_snowflake_tables(account, user, password, warehouse, database, schema):
    if not SNOWFLAKE_AVAILABLE: return []
    df = _fetch_snowflake(account, user, password, warehouse, database, schema, f"SHOW TABLES IN SCHEMA {database}.{schema}")
    return df["name"].tolist() if "name" in df.columns else []

def _sanitize_mongo_uri(uri):
    from urllib.parse import quote_plus
    if not uri or not uri.strip(): return uri
    uri = uri.strip()
    if MONGO_AVAILABLE:
        try:
            from pymongo import uri_parser
            uri_parser.parse_uri(uri); return uri
        except: pass
    if uri.startswith("mongodb+srv://"): scheme = "mongodb+srv"; rest = uri[len("mongodb+srv://"):]
    elif uri.startswith("mongodb://"): scheme = "mongodb"; rest = uri[len("mongodb://"):]
    else: return uri
    at_pos = next((i for i, ch in enumerate(rest) if ch == "@"), None)
    if at_pos is None: return uri
    creds_part = rest[:at_pos]; host_part = rest[at_pos+1:]
    if ":" in creds_part: username, password = creds_part.split(":", 1)
    else: username, password = creds_part, ""
    if "%" not in username: username = quote_plus(username)
    if "%" not in password: password = quote_plus(password)
    return f"{scheme}://{username}:{password}@{host_part}"

def _fetch_mongodb(uri, database, collection, fields, limit=0):
    if not MONGO_AVAILABLE: raise ImportError("pip install pymongo")
    client = pymongo.MongoClient(_sanitize_mongo_uri(uri))
    try:
        proj = {f: 1 for f in fields} if fields else {}
        proj["_id"] = 0
        cursor = client[database][collection].find({}, proj)
        if limit > 0: cursor = cursor.limit(limit)
        return pd.DataFrame(list(cursor))
    finally: client.close()

def _list_mongodb_collections(uri, database):
    if not MONGO_AVAILABLE: return []
    client = pymongo.MongoClient(_sanitize_mongo_uri(uri))
    try: return client[database].list_collection_names()
    finally: client.close()

def _fetch_mssql(host, port, user, password, database, query):
    if not MSSQL_AVAILABLE: raise ImportError("pip install pymssql")
    conn = pymssql.connect(server=host, port=int(port), user=user, password=password, database=database)
    try: return pd.read_sql_query(query, conn)
    finally: conn.close()

def _list_mssql_tables(host, port, user, password, database, schema):
    if not MSSQL_AVAILABLE: return []
    df = _fetch_mssql(host, port, user, password, database,
        f"SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE='BASE TABLE' AND TABLE_SCHEMA='{schema}'")
    return df["TABLE_NAME"].tolist() if "TABLE_NAME" in df.columns else []

def _fetch_bigquery(project, query, credentials_json=None):
    if not BIGQUERY_AVAILABLE: raise ImportError("pip install google-cloud-bigquery db-dtypes")
    if credentials_json:
        import google.oauth2.service_account as sa
        creds = sa.Credentials.from_service_account_info(json.loads(credentials_json))
        client = bq_lib.Client(project=project, credentials=creds)
    else: client = bq_lib.Client(project=project)
    return client.query(query).to_dataframe()

def _list_bigquery_datasets(project, credentials_json=None):
    if not BIGQUERY_AVAILABLE: return []
    if credentials_json:
        import google.oauth2.service_account as sa
        creds = sa.Credentials.from_service_account_info(json.loads(credentials_json))
        client = bq_lib.Client(project=project, credentials=creds)
    else: client = bq_lib.Client(project=project)
    return [d.dataset_id for d in client.list_datasets()]

def _fetch_ga4(property_id, credentials_json, start_date="30daysAgo", end_date="today"):
    if not GA_AVAILABLE: raise ImportError("pip install google-analytics-data")
    import google.oauth2.service_account as sa
    creds = sa.Credentials.from_service_account_info(
        json.loads(credentials_json),
        scopes=["https://www.googleapis.com/auth/analytics.readonly"])
    client = BetaAnalyticsDataClient(credentials=creds)
    req = RunReportRequest(
        property=f"properties/{property_id}",
        date_ranges=[DateRange(start_date=start_date, end_date=end_date)],
        dimensions=[Dimension(name="pageTitle"), Dimension(name="pagePath"), Dimension(name="country")],
        metrics=[Metric(name="sessions"), Metric(name="bounceRate"), Metric(name="averageSessionDuration")],
        limit=500)
    resp = client.run_report(req)
    rows = []
    for row in resp.rows:
        rows.append({d.name: row.dimension_values[i].value for i, d in enumerate(resp.dimension_headers)}
                  | {m.name: row.metric_values[j].value for j, m in enumerate(resp.metric_headers)})
    return pd.DataFrame(rows)

def _fetch_databricks(host, http_path, token, query):
    if not DATABRICKS_AVAILABLE: raise ImportError("pip install databricks-sql-connector")
    with dbsql.connect(server_hostname=host, http_path=http_path, access_token=token) as conn:
        with conn.cursor() as cur:
            cur.execute(query)
            cols = [d[0] for d in cur.description]
            return pd.DataFrame(cur.fetchall(), columns=cols)

def _list_databricks_tables(host, http_path, token, catalog="main", schema="default"):
    if not DATABRICKS_AVAILABLE: return []
    df = _fetch_databricks(host, http_path, token, f"SHOW TABLES IN {catalog}.{schema}")
    return df["tableName"].tolist() if "tableName" in df.columns else []

def fetch_all_rows(db_type, conn_params, table_name, tmp_path=""):
    """Fetch ALL rows from any DB type, no limit."""
    query = f"SELECT * FROM {table_name}"

    if db_type == "SQLite (file)":
        return _fetch_sqlite(tmp_path, query)
    elif db_type == "MySQL / MariaDB":
        return _fetch_mysql(conn_params['host'], conn_params['port'],
                            conn_params['user'], conn_params['password'],
                            conn_params['database'], query)
    elif db_type == "PostgreSQL":
        return _fetch_postgres(conn_params['host'], conn_params['port'],
                               conn_params['user'], conn_params['password'],
                               conn_params['database'], query)
    elif db_type == "SQL Server (MSSQL)":
        query = f"SELECT * FROM {conn_params['schema']}.{table_name}"
        return _fetch_mssql(conn_params['host'], conn_params['port'],
                            conn_params['user'], conn_params['password'],
                            conn_params['database'], query)
    elif db_type == "Snowflake":
        return _fetch_snowflake(conn_params['account'], conn_params['user'],
                                conn_params['password'], conn_params['warehouse'],
                                conn_params['database'], conn_params['schema'], query)
    elif db_type == "MongoDB":
        client = pymongo.MongoClient(_sanitize_mongo_uri(conn_params['uri']))
        try:
            docs = list(client[conn_params['database']][table_name].find({}, {"_id": 0}))
            return pd.DataFrame(docs)
        finally: client.close()
    elif db_type == "Google BigQuery":
        return _fetch_bigquery(conn_params['project'], query,
                               conn_params.get('credentials_json'))
    elif db_type == "Google Analytics 4":
        return _fetch_ga4(conn_params['property_id'],
                          conn_params.get('credentials_json', ''),
                          conn_params.get('start_date', '30daysAgo'),
                          conn_params.get('end_date', 'today'))
    elif db_type == "Databricks SQL":
        query = f"SELECT * FROM {conn_params['catalog']}.{conn_params['schema']}.{table_name}"
        return _fetch_databricks(conn_params['host'], conn_params['http_path'],
                                 conn_params['token'], query)
    return pd.DataFrame()

# ══════════════════════════════════════════════════════════════
# METADATA EXTRACTOR  — column-level stats for any DB type
# ══════════════════════════════════════════════════════════════
def get_table_metadata(db_type, conn_params, table_name, tmp_path="", raw_df=None):
    """
    Returns a list of dicts:
    { col_name, data_type, null_count, null_pct, unique_count,
      avg_len, min_len, max_len, sample_values }
    Works from raw_df if already fetched, otherwise queries DB for schema first.
    """
    meta = []

    # Use already-fetched df if available
    if raw_df is not None and not raw_df.empty:
        total = len(raw_df)
        for col in raw_df.columns:
            series = raw_df[col]
            null_count  = int(series.isna().sum())
            null_pct    = round(null_count / total * 100, 1) if total else 0
            unique_count = int(series.nunique())
            is_text = series.dtype == object
            lengths = series.dropna().astype(str).str.len() if is_text else None
            avg_len  = round(float(lengths.mean()), 1)  if is_text and len(lengths) else 0
            min_len  = int(lengths.min())               if is_text and len(lengths) else 0
            max_len  = int(lengths.max())               if is_text and len(lengths) else 0
            samples  = series.dropna().astype(str).head(3).tolist()
            meta.append({
                "col_name":     col,
                "data_type":    str(series.dtype),
                "null_count":   null_count,
                "null_pct":     null_pct,
                "unique_count": unique_count,
                "avg_len":      avg_len,
                "min_len":      min_len,
                "max_len":      max_len,
                "sample_values": samples,
            })
        return meta

    # Fallback: schema-only query (before full fetch)
    try:
        if db_type == "SQLite (file)":
            conn = sqlite3.connect(tmp_path)
            info = pd.read_sql_query(f"PRAGMA table_info({table_name})", conn)
            row_count = conn.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
            conn.close()
            for _, r in info.iterrows():
                meta.append({"col_name": r["name"], "data_type": r["type"],
                             "null_count": "?", "null_pct": "?",
                             "unique_count": "?", "avg_len": "?",
                             "min_len": "?", "max_len": "?",
                             "sample_values": [], "_row_count": row_count})

        elif db_type == "MySQL / MariaDB":
            info = _fetch_mysql(conn_params['host'], conn_params['port'],
                                conn_params['user'], conn_params['password'],
                                conn_params['database'], f"DESCRIBE {table_name}")
            for _, r in info.iterrows():
                meta.append({"col_name": r["Field"], "data_type": r["Type"],
                             "null_count":"?","null_pct":"?","unique_count":"?",
                             "avg_len":"?","min_len":"?","max_len":"?","sample_values":[]})

        elif db_type == "PostgreSQL":
            info = _fetch_postgres(conn_params['host'], conn_params['port'],
                                   conn_params['user'], conn_params['password'],
                                   conn_params['database'],
                                   f"SELECT column_name,data_type FROM information_schema.columns WHERE table_name='{table_name}'")
            for _, r in info.iterrows():
                meta.append({"col_name": r["column_name"], "data_type": r["data_type"],
                             "null_count":"?","null_pct":"?","unique_count":"?",
                             "avg_len":"?","min_len":"?","max_len":"?","sample_values":[]})

        elif db_type == "SQL Server (MSSQL)":
            info = _fetch_mssql(conn_params['host'], conn_params['port'],
                                conn_params['user'], conn_params['password'],
                                conn_params['database'],
                                f"SELECT COLUMN_NAME,DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME='{table_name}'")
            for _, r in info.iterrows():
                meta.append({"col_name": r["COLUMN_NAME"], "data_type": r["DATA_TYPE"],
                             "null_count":"?","null_pct":"?","unique_count":"?",
                             "avg_len":"?","min_len":"?","max_len":"?","sample_values":[]})

        elif db_type == "Snowflake":
            info = _fetch_snowflake(conn_params['account'], conn_params['user'],
                                    conn_params['password'], conn_params['warehouse'],
                                    conn_params['database'], conn_params['schema'],
                                    f"DESCRIBE TABLE {table_name}")
            for _, r in info.iterrows():
                meta.append({"col_name": r.get("name","?"), "data_type": r.get("type","?"),
                             "null_count":"?","null_pct":"?","unique_count":"?",
                             "avg_len":"?","min_len":"?","max_len":"?","sample_values":[]})

        elif db_type == "Databricks SQL":
            info = _fetch_databricks(conn_params['host'], conn_params['http_path'],
                                     conn_params['token'], f"DESCRIBE {conn_params['catalog']}.{conn_params['schema']}.{table_name}")
            for _, r in info.iterrows():
                meta.append({"col_name": r.get("col_name","?"), "data_type": r.get("data_type","?"),
                             "null_count":"?","null_pct":"?","unique_count":"?",
                             "avg_len":"?","min_len":"?","max_len":"?","sample_values":[]})

        elif db_type == "MongoDB":
            # Sample 1 doc to infer field types
            client = pymongo.MongoClient(_sanitize_mongo_uri(conn_params['uri']))
            try:
                sample = client[conn_params['database']][table_name].find_one({}, {"_id": 0})
                if sample:
                    for k, v in sample.items():
                        meta.append({"col_name": k, "data_type": type(v).__name__,
                                     "null_count":"?","null_pct":"?","unique_count":"?",
                                     "avg_len":"?","min_len":"?","max_len":"?","sample_values":[str(v)]})
            finally: client.close()

    except Exception as ex:
        meta = [{"col_name": f"Error: {ex}", "data_type":"","null_count":"","null_pct":"",
                 "unique_count":"","avg_len":"","min_len":"","max_len":"","sample_values":[]}]

    return meta


def render_metadata_panel(meta, detected_fb_col=""):
    """Render a styled metadata table with column-level stats."""
    if not meta:
        st.info("No metadata available.")
        return

    st.markdown(
        '<div style="font-family:Space Mono,monospace;font-size:0.7rem;color:#888;'
        'letter-spacing:2px;text-transform:uppercase;margin-bottom:0.6rem;">'
        '🗂️ Table Metadata — Column Inspector</div>',
        unsafe_allow_html=True
    )

    header_html = (
        '<div class="meta-row" style="border-bottom:2px solid rgba(249,168,37,0.4);'
        'padding-bottom:6px;margin-bottom:4px;">'
        '<span class="meta-col-name">Column</span>'
        '<span class="meta-col-type">Type</span>'
        '<span class="meta-col-null">Nulls</span>'
        '<span style="min-width:70px;font-family:Space Mono,monospace;font-size:0.68rem;color:#f9a825;">Unique</span>'
        '<span style="min-width:80px;font-family:Space Mono,monospace;font-size:0.68rem;color:#f9a825;">Avg Len</span>'
        '<span class="meta-col-sample">Sample Values</span>'
        '</div>'
    )

    rows_html = ""
    for m in meta:
        is_fb = m["col_name"] == detected_fb_col
        highlight = "background:rgba(249,168,37,0.07);border-radius:4px;" if is_fb else ""
        fb_badge  = ' <span style="background:#f9a825;color:#000;border-radius:3px;padding:0 5px;font-size:0.6rem;font-weight:700;">FEEDBACK</span>' if is_fb else ""
        null_pct  = f"{m['null_pct']}%" if isinstance(m['null_pct'], (int, float)) else str(m['null_pct'])
        avg_len   = f"{m['avg_len']}" if m['avg_len'] != "?" else "?"
        samples   = " · ".join([str(s)[:40] for s in m['sample_values'][:2]]) if m['sample_values'] else "—"

        rows_html += (
            f'<div class="meta-row" style="{highlight}">'
            f'<span class="meta-col-name">{m["col_name"]}{fb_badge}</span>'
            f'<span class="meta-col-type">{m["data_type"]}</span>'
            f'<span class="meta-col-null">{null_pct}</span>'
            f'<span style="min-width:70px;font-family:Space Mono,monospace;font-size:0.68rem;color:#aaa;">{m["unique_count"]}</span>'
            f'<span style="min-width:80px;font-family:Space Mono,monospace;font-size:0.68rem;color:#aaa;">{avg_len}</span>'
            f'<span class="meta-col-sample">{samples}</span>'
            f'</div>'
        )

    st.markdown(
        f'<div style="background:rgba(0,0,0,0.2);border:1px solid rgba(255,255,255,0.07);'
        f'border-radius:10px;padding:0.8rem 1rem;">'
        f'{header_html}{rows_html}</div>',
        unsafe_allow_html=True
    )

def smart_detect_columns(df):
    cols_lower = {c.lower().strip(): c for c in df.columns}

    def match_exact(name_list):
        return next((cols_lower[n] for n in name_list if n in cols_lower), None)

    def match_partial(keyword_list):
        return next((orig for lower, orig in cols_lower.items()
                     if any(kw in lower for kw in keyword_list)), None)

    def match_longest_text():
        skip_re = re.compile(
            r'\b(id|uuid|key|hash|token|flag|count|score|rating|status|type|'
            r'category|label|tag|code|email|phone|url|link|image|photo|avatar|ip)\b', re.I)
        text_cols = [c for c in df.select_dtypes(include='object').columns
                     if not skip_re.search(c)]
        if not text_cols: return None
        return max(text_cols,
                   key=lambda c: df[c].dropna().astype(str).str.len().mean())

    # ── Feedback column ───────────────────────────────────────
    layer_used = ""
    fb_col = match_exact(FEEDBACK_COL_EXACT)
    if fb_col: layer_used = "Layer 1 — Exact name match"

    if not fb_col:
        fb_col = match_partial(FEEDBACK_COL_PARTIAL)
        if fb_col: layer_used = "Layer 2 — Partial name match"

    if not fb_col:
        fb_col = match_longest_text()
        if fb_col: layer_used = "Layer 3 — Longest text column (data-driven)"

    if not fb_col:
        try:
            ai_res = ai_detect_feedback_column(df)
            fb_col = ai_res.get("feedback_col")
            layer_used = "Layer 4 — AI fallback"
        except: pass

    # ── Name column ───────────────────────────────────────────
    nm_col = match_exact(NAME_COL_EXACT)
    if not nm_col:
        nm_col = match_partial(["name", "author", "reviewer", "user", "customer"])

    # ── Date column ───────────────────────────────────────────
    dt_col = match_exact(DATE_COL_EXACT)
    if not dt_col:
        dt_col = match_partial(["date", "time", "created", "posted", "submitted"])

    # ── Per-column confidence scores ──────────────────────────
    col_scores = {}
    for lower, orig in cols_lower.items():
        if orig == fb_col:
            score = 95
        elif lower in FEEDBACK_COL_EXACT:
            score = 80
        elif any(kw in lower for kw in FEEDBACK_COL_PARTIAL):
            score = 50
        else:
            avg = df[orig].dropna().astype(str).str.len().mean() if df[orig].dtype == object else 0
            score = min(40, int(avg / 10))
        col_scores[orig] = score

    return {
        "feedback_col":     fb_col,
        "name_col":         nm_col,
        "date_col":         dt_col,
        "layer_used":       layer_used,
        "confidence_score": col_scores.get(fb_col, 0) if fb_col else 0,
        "column_scores":    col_scores,
        "reason":           f"Detected via {layer_used}" if layer_used else "No feedback column found",
    }

def ai_detect_feedback_column(df, source_label=""):
    hfb, hnm, hdt = _heuristic_detect(df)
    if not client:
        return {"feedback_col": hfb, "name_col": hnm, "date_col": hdt,
                "confidence_score": 50, "reason": "No API key.", "column_scores": {}}
    schema_lines = []
    for col in df.columns:
        sample  = df[col].dropna().astype(str).head(3).tolist()
        avg_len = int(df[col].dropna().astype(str).str.len().mean()) if len(df[col].dropna()) else 0
        schema_lines.append(f'  - "{col}" | dtype:{df[col].dtype} | avg_len:{avg_len} | samples:{sample}')
    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role":"system","content":_STRICT_DETECT_SYSTEM},
                      {"role":"user","content":f"{_FEEDBACK_DEFINITION}\n\n" + "\n".join(schema_lines)}],
            max_tokens=500, temperature=0.0)
        raw = re.sub(r"^```(?:json)?|```$", "", resp["choices"][0]["message"]["content"].strip(), flags=re.M).strip()
        result = json.loads(raw)
        result.setdefault("feedback_col", hfb)
        result.setdefault("name_col", hnm)
        result.setdefault("date_col", hdt)
        result.setdefault("confidence_score", 0)
        result.setdefault("reason", "")
        result.setdefault("column_scores", {})
        return result
    except:
        return {"feedback_col": hfb, "name_col": hnm, "date_col": hdt,
                "confidence_score": 50, "reason": "AI error; heuristic used.", "column_scores": {}}

def _heuristic_detect(df):
    text_cols  = [c for c in df.columns if df[c].dtype == object]
    skip_re    = re.compile(r'\b(id|uuid|key|hash|token|flag|count|score|rating|status|type|category|label|tag|code|email|phone|url|link|image|photo|avatar)\b', re.I)
    candidates = [c for c in text_cols if not skip_re.search(c)]
    fb_col     = next((c for c in candidates if any(k in c.lower() for k in ['feedback','review','comment','text','body','description','message','opinion','remark','note'])), None)
    if not fb_col and candidates:
        fb_col = max(candidates, key=lambda c: df[c].dropna().astype(str).str.len().mean())
    if not fb_col and text_cols: fb_col = text_cols[0]
    nm_col = next((c for c in df.columns if any(k in c.lower() for k in ['name','reviewer','author','user','customer','username'])), None)
    dt_col = next((c for c in df.columns if any(k in c.lower() for k in ['date','time','created','posted','timestamp','when','at'])), None)
    return fb_col, nm_col, dt_col


def save_session(sid, uname, uid8, src, total, notes=""):
    now = datetime.now()
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        "INSERT OR REPLACE INTO sessions"
        "(session_id, user_name, short_uuid, created_date, created_time, created_at, source_type, total_entries, notes)"
        "VALUES(?,?,?,?,?,?,?,?,?)",
        (sid, uname, uid8, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"),
         now.strftime("%Y-%m-%d %H:%M:%S"), src, total, notes)
    )
    conn.commit()
    conn.close()

def gen_sid(uname):
    now = datetime.now()
    safe = re.sub(r'[^a-zA-Z0-9]', '', uname.upper())[:12] or "USER"
    uid8 = str(uuid.uuid4()).replace('-', '')[:8].upper()
    return f"{safe}-{now.strftime('%Y%m%d')}-{now.strftime('%H%M%S')}-{uid8}", uid8

def extract_exact_date(text):
    if not text or not isinstance(text,str): return ''
    m=re.search(r'(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})',text)
    if m:
        try:
            y,mo,d=int(m.group(1)),int(m.group(2)),int(m.group(3))
            if 1990<=y<=2100 and 1<=mo<=12 and 1<=d<=31:
                return datetime(y,mo,d).strftime("%b %d, %Y")
        except: pass
    m=re.search(r'(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\.?\s+(\d{1,2}),?\s+(\d{4})',text,re.I)
    if m:
        try:
            mo=MONTH_MAP[m.group(1).lower()[:3]]; d,y=int(m.group(2)),int(m.group(3))
            return datetime(y,mo,d).strftime("%b %d, %Y")
        except: pass
    m=re.search(r'(\d{1,2})\s+(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)\.?\s+(\d{4})',text,re.I)
    if m:
        try:
            d=int(m.group(1)); mo=MONTH_MAP[m.group(2).lower()[:3]]; y=int(m.group(3))
            return datetime(y,mo,d).strftime("%b %d, %Y")
        except: pass
    m=re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})',text)
    if m:
        try:
            a,b,y=int(m.group(1)),int(m.group(2)),int(m.group(3))
            mo,d=(a,b) if a<=12 else (b,a)
            if 1<=mo<=12 and 1<=d<=31: return datetime(y,mo,d).strftime("%b %d, %Y")
        except: pass
    return ''

def is_junk(text: str) -> bool:
    if not text:
        return True
    t = text.strip().lower()
    if len(t) < 50:
        return True
    if _JUNK_RE.search(t):
        return True
    for pat in _JUNK_PATTERNS:
        if pat.search(text):
            return True
    alpha = sum(1 for c in t if c.isalpha())
    if len(t) > 0 and alpha / len(t) < 0.4:
        return True
    return False

def preprocess(text):
    if not text or not isinstance(text, str):
        return ''
    text = text.lower()
    text = re.sub(r'http\S+|www\S+', '', text)
    text = re.sub(r'[^a-z0-9\s]', '', text)
    try:
        words = word_tokenize(text)
    except:
        words = text.split()
    sw = set(stopwords.words('english'))
    words = [w for w in words if w not in sw and len(w) > 2]
    try:
        lem = WordNetLemmatizer()
        words = [lem.lemmatize(w) for w in words]
    except:
        pass
    return ' '.join(words) if len(words) >= 2 else ''

def topic_model(texts, n=5):
    if len(texts) < 2:
        return [0] * len(texts)
    n = min(n, len(texts))
    try:
        vec = TfidfVectorizer(stop_words='english', max_features=500, min_df=1)
        X = vec.fit_transform(texts)
        W = NMF(n_components=n, random_state=42, max_iter=300).fit_transform(X)
        return W.argmax(axis=1).tolist()
    except:
        return [0] * len(texts)

def get_ai_with_names(feedback_list, existing_names=None):
    if existing_names is None:
        existing_names = [''] * len(feedback_list)

    try:
        prompt = URL_USER_PROMPT_TEMPLATE
        for i, (fb, nm) in enumerate(zip(feedback_list, existing_names), 1):
            hint = f" [Reviewer already known: {nm}]" if nm and nm.strip() else " [Extract reviewer name if present]"
            prompt += f"{i}. {fb}{hint}\n"

        resp = client.chat.completions.create(
            model="gpt-4",
            messages=[
                {"role": "system", "content": URL_SYSTEM_PROMPT},
                {"role": "user", "content": prompt}
            ],
            max_tokens=2500,
            temperature=0.85
        )

        text = resp["choices"][0]["message"]["content"]
        raw_lines = text.strip().split('\n')
        parsed = {}

        for line in raw_lines:
            line = line.strip()
            if not line: continue
            num_m = re.match(r'^(\d+)[\.\)]\s*', line)
            if not num_m: continue
            idx = int(num_m.group(1)) - 1
            line = line[num_m.end():]

            name = ""
            nm_m = re.search(r'\[NAME:\s*([^\]]+)\]', line, re.I)
            if nm_m:
                name = nm_m.group(1).strip()
                if name.lower() in ("anonymous", "unknown", "n/a"):
                    name = ""
                line = line.replace(nm_m.group(0), "").strip()

            sentiment = "Neutral"
            for s in ["Positive", "Negative", "Mixed", "Neutral"]:
                tag = f"[{s}]"
                if tag in line:
                    sentiment = s
                    line = line.replace(tag, "").strip()
                    break

            line = line.strip()
            if idx not in parsed:
                parsed[idx] = (name, line, sentiment)

        results = []
        for i in range(len(feedback_list)):
            if i in parsed:
                results.append(parsed[i])
            else:
                results.append(("", "", "Neutral"))

        return results

    except Exception as e:
        print("AI parsing error:", e)
        return [("", "", "Neutral")] * len(feedback_list)

def get_suggestions_openai(feedback_list):
    """
    Call OpenAI for one specific suggestion per feedback item.
    Returns a list of strings (same length as feedback_list).
    Empty string = AI failed for that item → caller uses generate_fallback_suggestion().
    """
    if not client:
        return [""] * len(feedback_list)
    try:
        user_prompt = CSV_USER_PROMPT_TEMPLATE
        for i, fb in enumerate(feedback_list, 1):
            user_prompt += f"{i}. {str(fb)[:400]}\n"

        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": CSV_SYSTEM_PROMPT},
                {"role": "user",   "content": user_prompt}
            ],
            max_tokens=3000,
            temperature=0.85
        )

        raw = resp.choices[0].message.content.strip()

        # Robust parser: split on numbered lines (handles 1. / 1) / 1:)
        segments = re.split(r'(?m)^(\d+)[\.):]\s+', raw)
        suggestions_map = {}
        i = 1
        while i < len(segments) - 1:
            try:
                num = int(segments[i])
                text = segments[i + 1].strip()
                text = re.sub(r'\s*\n\s*', ' ', text).strip()
                if len(text) > 10:
                    suggestions_map[num] = text
            except (ValueError, IndexError):
                pass
            i += 2

        result = []
        for idx in range(1, len(feedback_list) + 1):
            result.append(suggestions_map.get(idx, ""))
        return result

    except Exception as e:
        st.warning(f"⚠️ AI suggestion error: {e}. Using keyword-based fallback.")
        return [""] * len(feedback_list)

def rule_based_sentiment(text: str) -> str:
    if not text or len(text.strip()) < 10: return 'Neutral'
    t = text.lower().strip()
    pos = len(_POS.findall(t))
    neg = len(_NEG.findall(t))
    if pos >= 1 and neg >= 1 and _MIXED_IND.search(t): return 'Mixed'
    if pos > neg * 1.5: return 'Positive'
    if neg > pos * 1.5: return 'Negative'
    if pos > 0 and neg > 0: return 'Mixed'
    if pos > 0: return 'Positive'
    if neg > 0: return 'Negative'
    mild_pos = re.search(r'\b(nice|okay|ok|fine|decent|average|alright|not bad|acceptable)\b', t)
    mild_neg = re.search(r'\b(not great|could be better|needs improvement|lacking|mediocre|underwhelming)\b', t)
    if mild_pos and not mild_neg: return 'Positive'
    if mild_neg and not mild_pos: return 'Negative'
    if mild_pos and mild_neg: return 'Mixed'
    return 'Neutral'

def get_sentiment_batch_ai(feedback_list: list) -> list:
    if not client or not feedback_list:
        return [rule_based_sentiment(fb) for fb in feedback_list]
    try:
        prompt = ("Classify sentiment of each feedback. Reply ONLY with numbered lines:\n"
                  "1. Positive\n2. Negative\n...\nValid: Positive, Negative, Neutral, Mixed\n\nFeedbacks:\n")
        for i, fb in enumerate(feedback_list, 1):
            prompt += f"{i}. {str(fb)[:300]}\n"
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[{"role":"system","content":"Precise sentiment classifier. Output only numbered labels."},
                      {"role":"user","content":prompt}],
            max_tokens=len(feedback_list)*8+50, temperature=0.0)
        raw = resp["choices"][0]["message"]["content"].strip()
        results = {}
        for line in raw.split('\n'):
            m = re.match(r'^(\d+)\.\s*(Positive|Negative|Neutral|Mixed)', line.strip(), re.I)
            if m: results[int(m.group(1))-1] = m.group(2).capitalize()
        return [results.get(i, rule_based_sentiment(fb)) for i, fb in enumerate(feedback_list)]
    except Exception:
        return [rule_based_sentiment(fb) for fb in feedback_list]

def analyze_sentiments_all(feedback_list: list) -> list:
    rule_results = [rule_based_sentiment(fb) for fb in feedback_list]
    if not client: return rule_results
    uncertain_idx = [i for i, s in enumerate(rule_results) if s == 'Neutral']
    if not uncertain_idx: return rule_results
    uncertain_texts = [feedback_list[i] for i in uncertain_idx]
    ai_results = []
    for i in range(0, len(uncertain_texts), 20):
        ai_results.extend(get_sentiment_batch_ai(uncertain_texts[i:i+20]))
    for orig_idx, ai_sent in zip(uncertain_idx, ai_results):
        rule_results[orig_idx] = ai_sent
    return rule_results

# def get_advanced_analysis(feedback_list):
#     final_suggestions = []
#     final_sentiments = []

#     for fb in feedback_list:
#         raw = analyze_feedback_parts(fb)
#         sents, depts, sugs = parse_parts_output(raw)

#         final_sentiments.append(combine_sentiment(sents))
#         final_suggestions.append(combine_suggestions(depts, sugs))

#     return final_suggestions, final_sentiments

def get_advanced_analysis(feedback_list):
    final_suggestions = []
    final_sentiments = []

    for fb in feedback_list:
        try:
            # st.write(f"Analyzing feedback: {str(fb)[:200]}...")
            raw = analyze_feedback_parts(fb)
            # st.write("RAW AI OUTPUT:", raw)
            sents, depts, sugs = parse_parts_output(raw)
            # st.write("PARSED SENTIMENTS:", sents)
            # st.write("PARSED DEPARTMENTS:", depts)
            # st.write("PARSED SUGGESTIONS:", sugs)
            # 🔴 SAFETY: if parsing fails
            if not sents or not sugs:
                final_sentiments.append(rule_based_sentiment(fb))
                final_suggestions.append(generate_fallback_suggestion(fb, 0))
                continue

            final_sentiments.append(combine_sentiment(sents))
            final_suggestions.append(combine_suggestions(depts, sugs))

        except Exception as e:
            st.write("Advanced AI error:", e)
            final_sentiments.append(rule_based_sentiment(fb))
            final_suggestions.append(generate_fallback_suggestion(fb, 0))

    return final_suggestions, final_sentiments

def run_ai_analysis(fb_df, mode='full'):
    """
    mode='full'       : extract name + suggestion + sentiment (URL scraping)
    mode='suggestion' : suggestion only using CCX Officer prompt (CSV upload)
    """
    ai_names, sugs, sents = [], [], []
    existing = [
        fb_df['Reviewer_Name'].iloc[i]
        if 'Reviewer_Name' in fb_df.columns and i < len(fb_df) else ''
        for i in range(len(fb_df))
    ]
    bs = 10
    prog = st.progress(0)
    stat = st.empty()
    tb = max(1, (len(fb_df) + bs - 1) // bs)

    if client:
        for i in range(0, len(fb_df), bs):
            batch = fb_df['Feedback'].iloc[i:i + bs].tolist()
            # st.write(f"Processing batch {i // bs + 1}/{tb}...")
            # st.write(batch)
            bn = i // bs + 1
            stat.markdown(
                f'<p class="pulse" style="color:#f9a825;font-family:Space Mono,monospace;'
                f'font-size:0.8rem;">🤖 AI batch {bn}/{tb}...</p>',
                unsafe_allow_html=True,
            )
            if mode == 'full':
                for nm, sg, se in get_ai_with_names(batch, existing[i:i + bs]):
                    ai_names.append(nm)
                    sugs.append(sg)
                    sents.append(se)
            else:
                batch_sugs, batch_sents = get_advanced_analysis(batch)

                for sg, se in zip(batch_sugs, batch_sents):
                    ai_names.append('')
                    sugs.append(sg)
                    sents.append(se)
                    # sents.append('Neutral')
            prog.progress(min(1.0, (i + bs) / len(fb_df)))
            time.sleep(0.3)
    else:
        ai_names = [''] * len(fb_df)
        sugs = [''] * len(fb_df)
        # sents = ['Neutral'] * len(fb_df)
        # sents = analyze_sentiments_all(fb_df['Feedback'].tolist())
    if not sents:
        sents = analyze_sentiments_all(fb_df['Feedback'].tolist())
    stat.empty()
    prog.empty()
    return ai_names, sugs, sents

def generate_fallback_suggestion(feedback_text: str, topic_id: int) -> str:
    """
    Generate a non-repetitive, feedback-specific suggestion without AI.
    Priority order:
      1. Keyword match on complaint context → specific contextual action
      2. Hash-bucketed topic template  →  same feedback always gets same slot
         (deterministic, yet different rows almost always get different templates)
      3. Round-robin fallback  →  ensures variety even within the same hash bucket
    """
    text_lower = (feedback_text or '').lower()

    # Strategy 1: keyword match → specific contextual suggestion
    for pattern, suggestion in _FALLBACK_KEYWORD_ACTIONS:
        if re.search(pattern, text_lower):
            return suggestion

    # Strategy 2: hash-bucketed template selection
    # Use first 8 chars of MD5 of normalised text as a stable bucket index.
    # This means the same feedback always maps to the same template (reproducible),
    # but different feedbacks almost always land in different buckets.
    import hashlib as _hl
    bucket = int(_hl.md5(text_lower[:300].encode()).hexdigest()[:8], 16)
    templates = _FALLBACK_TOPIC_ACTIONS.get(topic_id, _FALLBACK_TOPIC_ACTIONS[0])
    suggestion = templates[bucket % len(templates)]

    # Strategy 3: if somehow the same suggestion was already used in this run,
    # advance the round-robin counter to pick the next one
    prev_idx = _fallback_topic_counters.get(topic_id, 0)
    candidate_idx = bucket % len(templates)
    # If this exact template was the last one used for this topic, shift by 1
    if candidate_idx == prev_idx and len(templates) > 1:
        candidate_idx = (candidate_idx + 1) % len(templates)
    suggestion = templates[candidate_idx]
    _fallback_topic_counters[topic_id] = candidate_idx
    return suggestion

def assign_ids(df):
    ctr = {}
    try:
        conn = sqlite3.connect(DB_PATH)
        ex = pd.read_sql_query("SELECT feedback_id FROM feedback_entries", conn)
        conn.close()
        for fid in ex['feedback_id'].dropna():
            m = re.match(r'^([A-Z]+)(\d+)$', str(fid).strip())
            if m:
                p, n = m.group(1), int(m.group(2))
                if ctr.get(p, 0) <= n:
                    ctr[p] = n + 1
    except:
        pass
    ids, local = [], {}
    for _, row in df.iterrows():
        p = get_prefix(row['Source'])
        if p not in local:
            local[p] = ctr.get(p, 1)
        n = local[p]
        ids.append(f"{p}{str(n).zfill(2) if n < 100 else n}")
        local[p] = n + 1
    df = df.copy()
    df.insert(0, 'Feedback_ID', ids)
    return df


def get_prefix(src):
    sl = src.lower().strip()
    for k, p in SOURCE_PREFIX_MAP.items():
        if k in sl:
            return p
    c = re.sub(r'[^a-zA-Z]', '', src).upper()
    return c[:2] if len(c) >= 2 else 'FB'


def build_results(fbs, src, names=None, dates=None, mode='full'):
    if names is None:
        names = [''] * len(fbs)
    if dates is None:
        dates = [''] * len(fbs)
    valid = [
        (fb, nm, dt) for fb, nm, dt in zip(fbs, names, dates)
        if fb and str(fb).strip() and not is_junk(str(fb))
    ]
    if not valid:
        return pd.DataFrame()
    fbs2, names2, dates2 = map(list, zip(*valid))
    df = pd.DataFrame({
        'Feedback': fbs2,
        'Source': src,
        'Reviewer_Name': names2,
        'Feedback_Date': dates2,
    })
    df['_c'] = df['Feedback'].apply(preprocess)
    df = df[df['_c'].str.strip() != ''].reset_index(drop=True)
    if df.empty:
        return df
    df['TopicID'] = topic_model(df['_c'].tolist())
    df['Topic'] = df['TopicID'].map(TOPIC_LABELS)
    ai_names, sugs, sents = run_ai_analysis(df, mode=mode)
    final_names = []
    for scraped, ai_nm in zip(df['Reviewer_Name'].tolist(), ai_names):
        if scraped and str(scraped).strip() and scraped not in ('nan', 'None', ''):
            final_names.append(str(scraped).strip())
        elif ai_nm and str(ai_nm).strip():
            final_names.append(str(ai_nm).strip())
        else:
            final_names.append('')
    df['Reviewer_Name'] = final_names

    def pick_suggestion(ai_sug, topic_id, feedback_text=''):
        if ai_sug and isinstance(ai_sug, str) and len(ai_sug.strip()) > 20:
            return ai_sug.strip()
        return generate_fallback_suggestion(feedback_text, topic_id)

    df['Suggestion'] = [
        pick_suggestion(s, t, f)
        for s, t, f in zip(sugs, df['TopicID'], df['Feedback'])
    ]
    df['Sentiment'] = sents
    df['Future_Action_Plan'] = generate_future_action_plan(df['Suggestion'].tolist())
    df['Future_Action_Plan'] = df['Future_Action_Plan'].apply(clean_action_plan)
    result = df[['Source', 'Reviewer_Name', 'Feedback_Date', 'Feedback', 'Topic', 'Sentiment', 'Suggestion', 'Future_Action_Plan']]
    return assign_ids(result)

def _ensure_cache_loaded():
    """Load all content_hashes from DB into memory once per process."""
    global _HASH_CACHE, _HASH_CACHE_LOADED
    if _HASH_CACHE_LOADED:
        return
    try:
        conn = sqlite3.connect(DB_PATH)
        rows = conn.execute(
            "SELECT content_hash FROM feedback_entries WHERE content_hash IS NOT NULL AND content_hash != ''"
        ).fetchall()
        conn.close()
        _HASH_CACHE.update(r[0] for r in rows)
    except Exception:
        pass
    _HASH_CACHE_LOADED = True

def make_content_hash(fb_text: str) -> str:
    """SHA-256 of the normalised feedback text."""
    return hashlib.sha256(_normalize_for_hash(fb_text).encode('utf-8')).hexdigest()

def _normalize_for_hash(text: str) -> str:
    """
    Produce a canonical string for dedup hashing.
    Lowercases, strips all punctuation/whitespace variations,
    and removes common filler so minor scraping differences
    (trailing ellipsis, extra spaces, HTML artefacts) don't
    create false new entries.
    """
    if not text or not isinstance(text, str):
        return ''
    t = text.lower()
    # remove URLs
    t = re.sub(r'http\S+|www\S+', '', t)
    # collapse punctuation runs and whitespace
    t = re.sub(r'[^\w\s]', ' ', t)
    t = re.sub(r'\s+', ' ', t).strip()
    # take first 500 chars of cleaned text — long enough to be unique,
    # short enough to be stable across minor scraper truncations
    return t[:500]

def render_results_tab(df, sid, uname_val):
    total = len(df); tc_data = df['Topic'].value_counts()
    st.markdown(f'<div class="session-box">🔑 <strong>{sid}</strong> &nbsp;·&nbsp; 👤 {uname_val} &nbsp;·&nbsp; 📅 {datetime.now().strftime("%d %b %Y, %H:%M")} &nbsp;·&nbsp; 📊 {total} entries</div>', unsafe_allow_html=True)
    c1,c2,c3,c4,c5,c6 = st.columns(6)
    c1.metric("Total Entries", f"{total:,}"); c2.metric("Topics", df['Topic'].nunique())
    c3.metric("Sources", df['Source'].nunique())
    named = df['Reviewer_Name'].astype(str).str.strip().ne('').sum() if 'Reviewer_Name' in df.columns else 0
    c4.metric("Named Reviewers", named)
    pos = df['Sentiment'].eq('Positive').sum() if 'Sentiment' in df.columns else 0
    neg = df['Sentiment'].eq('Negative').sum() if 'Sentiment' in df.columns else 0
    c5.metric("😊 Positive", pos); c6.metric("😞 Negative", neg)
    if 'Sentiment' in df.columns:
        st.divider(); st.markdown("### 🎯 Sentiment Distribution")
        sc = df['Sentiment'].value_counts(); sent_cols = st.columns(len(sc))
        SENT_S = {'Positive':('#4caf50','😊'),'Negative':('#e91e63','😞'),'Neutral':('#9e9e9e','😐'),'Mixed':('#ff9800','🤔')}
        for i,(sent,cnt) in enumerate(sc.items()):
            col_h, icon = SENT_S.get(sent,('#888','·'))
            with sent_cols[i]:
                st.markdown(f'<div style="background:rgba(255,255,255,0.04);border:1px solid {col_h}40;border-radius:12px;padding:1rem;text-align:center;"><div style="font-size:1.6rem;">{icon}</div><div style="font-size:1.4rem;font-weight:800;color:{col_h};font-family:Space Mono,monospace;">{cnt}</div><div style="font-size:0.72rem;color:#aaa;margin:2px 0;">{sent}</div><div style="font-size:0.75rem;color:#555;">{round(cnt/total*100)}%</div></div>', unsafe_allow_html=True)
    st.divider(); st.markdown("### 🗂️ Topic Distribution")
    cols = st.columns(min(len(tc_data),5))
    for i,(topic,count) in enumerate(tc_data.items()):
        with cols[i%5]:
            st.markdown(f'<div style="background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:12px;padding:1rem;text-align:center;"><div style="font-size:1.6rem;font-weight:800;color:#f9a825;font-family:Space Mono,monospace;">{count}</div><div style="font-size:0.68rem;color:#aaa;margin:4px 0;">{topic}</div><div style="font-size:0.75rem;color:#666;">{round(count/total*100)}%</div></div>', unsafe_allow_html=True)
    st.divider(); st.markdown("### 🔍 Browse & Filter")
    cf1,cf2,cf3,cf4 = st.columns(4)
    with cf1: tf = st.multiselect("Topic", df['Topic'].unique().tolist(), default=df['Topic'].unique().tolist(), key=f"tf_{sid[:8]}")
    with cf2: sf = st.multiselect("Source", df['Source'].unique().tolist(), default=df['Source'].unique().tolist(), key=f"sf_{sid[:8]}")
    with cf3:
        sent_opts = ['All'] + sorted(df['Sentiment'].unique().tolist()) if 'Sentiment' in df.columns else ['All']
        sent_filter = st.selectbox("Sentiment", sent_opts, key=f"sent_{sid[:8]}")
    with cf4: search_q = st.text_input("🔎 Search", placeholder="keyword...", key=f"sq_{sid[:8]}")
    filtered = df[df['Topic'].isin(tf) & df['Source'].isin(sf)]
    if sent_filter != 'All' and 'Sentiment' in filtered.columns:
        filtered = filtered[filtered['Sentiment']==sent_filter]
    if search_q.strip():
        filtered = filtered[filtered['Feedback'].str.contains(search_q.strip(), case=False, na=False)]
    st.markdown(f'<div style="font-family:Space Mono,monospace;font-size:0.75rem;color:#888;margin-bottom:1rem;">Showing {len(filtered):,} of {total:,}</div>', unsafe_allow_html=True)
    view = st.radio("View", ["📋 Table","🃏 Cards"], horizontal=True, key=f"view_{sid[:8]}")
    if view == "📋 Table":
        dcols = [c for c in ['Feedback_ID','Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion'] if c in filtered.columns]
        st.dataframe(filtered[dcols], use_container_width=True, height=420,
            column_config={"Feedback_ID":st.column_config.TextColumn("ID",width="small"),
                "Feedback":st.column_config.TextColumn("Feedback",width="large"),
                "Suggestion":st.column_config.TextColumn("AI Suggestion",width="large"),
                "Sentiment":st.column_config.TextColumn("Sentiment",width="small"),
                "Topic":st.column_config.TextColumn("Topic",width="medium"),
                "Reviewer_Name":st.column_config.TextColumn("Reviewer",width="small"),
                "Feedback_Date":st.column_config.TextColumn("Date",width="small")})
    else:
        for _, row in filtered.head(30).iterrows():
            tid = [k for k,v in TOPIC_LABELS.items() if v==row['Topic']]
            bc = TOPIC_COLORS.get(tid[0] if tid else 0,"badge-0")
            fid = row.get('Feedback_ID',''); name = row.get('Reviewer_Name','')
            date = row.get('Feedback_Date',''); sent = row.get('Sentiment','')
            SENT_C = {'Positive':'#4caf50','Negative':'#e91e63','Neutral':'#9e9e9e','Mixed':'#ff9800'}
            sc_hex = SENT_C.get(sent,'#9e9e9e')
            SENT_I = {'Positive':'😊','Negative':'😞','Neutral':'😐','Mixed':'🤔'}
            sent_icon = SENT_I.get(sent,'·')
            meta = []
            if name and str(name).strip(): meta.append(f"👤 {name}")
            if date and str(date).strip(): meta.append(f"📅 {date}")
            mh = "  &nbsp;·&nbsp;  ".join(meta)
            fb_full = row['Feedback']; fb_display = fb_full[:1200]
            fb_overflow = (f'<span style="color:#f9a825;font-family:Space Mono,monospace;font-size:0.7rem;"> …[{len(fb_full)-1200} more chars]</span>' if len(fb_full) > 1200 else '')
            st.markdown(f"""<div class="card"><div style="display:flex;align-items:center;gap:8px;margin-bottom:0.8rem;flex-wrap:wrap;">
    <span class="feedback-id">{fid}</span><span class="source-tag">{row['Source']}</span>
    <span class="badge {bc}">{row['Topic']}</span>
    {f'<span style="background:rgba(0,0,0,0.3);border:1px solid {sc_hex};border-radius:12px;padding:1px 10px;font-family:Space Mono,monospace;font-size:0.7rem;color:{sc_hex};font-weight:700;">{sent_icon} {sent}</span>' if sent else ''}
    {f'<span style="font-family:Space Mono,monospace;font-size:0.68rem;color:#888;">{mh}</span>' if mh else ''}
</div><div style="color:#ddd;font-size:0.9rem;line-height:1.7;margin-bottom:0.8rem;">{fb_display}{fb_overflow}</div>
<div style="border-top:1px solid rgba(255,255,255,0.06);padding-top:0.8rem;">
    <span style="font-family:Space Mono,monospace;font-size:0.65rem;color:#f9a825;letter-spacing:2px;text-transform:uppercase;">💡 AI Suggestion</span>
    <div style="color:#b0b0b0;font-size:0.85rem;margin-top:4px;">{row.get('Suggestion','')}</div>
</div></div>""", unsafe_allow_html=True)
        if len(filtered) > 30: st.info(f"Showing 30 cards. Switch to Table for all {len(filtered)}.")
    st.divider(); st.markdown("### 📥 Download")
    d1,d2,d3 = st.columns(3)
    with d1:
        ecols = [c for c in ['Feedback_ID','Source','Reviewer_Name','Feedback_Date','Feedback','Topic','Sentiment','Suggestion'] if c in filtered.columns]
        st.download_button("⬇️ CSV — Full Data", data=filtered[ecols].to_csv(index=False).encode('utf-8'),
            file_name=f"zeus_feedback_{sid}.csv", mime='text/csv', use_container_width=True, key=f"dl_csv_{sid[:8]}")
    with d2:
        txt = make_summary_txt(filtered, sid, uname_val)
        st.download_button("⬇️ TXT Summary Report", data=txt.encode('utf-8'),
            file_name=f"zeus_report_{sid}.txt", mime='text/plain', use_container_width=True, key=f"dl_txt_{sid[:8]}")
    with d3:
        try:
            xls = make_feedback_excel(filtered, sid, uname_val)
            if xls:
                st.download_button("⬇️ Excel Report (.xlsx)", data=xls,
                    file_name=f"zeus_feedback_{sid}.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    use_container_width=True, key=f"dl_xls_{sid[:8]}")
            else: st.info("`pip install openpyxl` for Excel")
        except: st.info("`pip install openpyxl` for Excel")
    st.divider()
    if st.button("🔄 Clear & Start Over", key=f"clear_{sid[:8]}"):
        for k in ['results_df','analyzed','session_id','short_uuid','db_step','db_connected',
                  'db_raw_df','db_detected_cols','db_metadata']:
            st.session_state[k] = None if k in ('results_df','session_id','short_uuid') \
                else (False if k in ('analyzed','db_connected') \
                else (1 if k=='db_step' else ({} if k=='db_detected_cols' else ([] if k=='db_metadata' else ''))))
        st.rerun()

def save_entries(session_id, df):
    """
    Save only new entries using content_hash dedup.
    Updates both DB and module-level cache atomically.
    Returns (saved_count, skipped_count).
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    saved = 0
    skipped = 0
    _ensure_cache_loaded()

    for _, row in df.iterrows():
        fb = row.get('Feedback', '')
        nm = row.get('Reviewer_Name', '')
        h = make_content_hash(fb)

        if h in _HASH_CACHE:
            skipped += 1
            continue

        try:
            c.execute(
                "INSERT INTO feedback_entries"
                "(session_id, entry_uuid, feedback_id, source, reviewer_name, feedback_date,"
                " feedback, topic, sentiment, suggestion, analyzed_at, content_hash)"
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                (session_id, str(uuid.uuid4()),
                 row.get('Feedback_ID', ''), row.get('Source', ''), nm,
                 row.get('Feedback_Date', ''), fb,
                 row.get('Topic', ''), row.get('Sentiment', 'Neutral'),
                 row.get('Suggestion', ''), at, h)
            )
            _HASH_CACHE.add(h)
            saved += 1
        except sqlite3.IntegrityError:
            _HASH_CACHE.add(h)
            skipped += 1

    conn.commit()
    conn.close()
    return saved, skipped

def make_summary_txt(df, session_id, user_name):
    tc=df['Topic'].value_counts(); total=len(df)
    lines=["╔═══════════════════════════════════════════════╗",
           "║      ZEUS FEEDBACK ANALYZER  —  REPORT        ║",
           "╚═══════════════════════════════════════════════╝",
           f"  Session ID   : {session_id}",f"  Analyst      : {user_name}",
           f"  Generated    : {datetime.now().strftime('%d %B %Y, %H:%M:%S')}",
           f"  Total Entries: {total}",
           "─────────────────────────────────────────────────","","TOPIC DISTRIBUTION","──────────────────"]
    for topic,count in tc.items():
        bar="█"*int(count/total*30)
        lines.append(f"  {topic:<35} {count:>4} ({round(count/total*100):>2}%)  {bar}")
    lines+=["","TOP AI SUGGESTION PER TOPIC","────────────────────────────"]
    for tid,label in TOPIC_LABELS.items():
        sub=df[df['Topic']==label]
        if not sub.empty: lines+=[f"\n  {label}",f"  → {sub.iloc[0]['Suggestion']}"]
    lines+=["","","ALL ENTRIES","───────────"]
    for i,row in df.iterrows():
        fid=row.get('Feedback_ID',f"#{i+1}"); name=row.get('Reviewer_Name',''); date=row.get('Feedback_Date','')
        lines+=[f"\n  [{fid}] [{row['Source']}]  {row['Topic']}",
                f"  Reviewer   : {name if name else 'Anonymous'}",
                f"  Date       : {date if date else 'Unknown'}",
                f"  Feedback   : {row['Feedback'][:300]}",
                f"  Suggestion : {row['Suggestion']}","  "+"·"*65]
    return '\n'.join(lines)

def make_feedback_excel(df, session_id, user_name):
    from openpyxl import Workbook as WB
    from openpyxl.styles import PatternFill as PF,Font as FT,Alignment as AL,Border as BD,Side as SD
    from openpyxl.utils import get_column_letter as gcl
    from openpyxl.chart import BarChart as BC,Reference as RF
    DARK='1A1A2E';NAV='0F3460';GOLD='F9A825';WHITE='FFFFFF';LGT='F5F5F5';GRY='E0E0E0';DGR='666666'
    TC={'⚡ Service Quality':('F9A825','1A1A2E'),'📦 Product Issues':('E53935','FFFFFF'),
        '💬 Communication':('0288D1','FFFFFF'),'⏱️ Speed & Delays':('F57C00','FFFFFF'),
        '💡 Innovation & Features':('7B1FA2','FFFFFF')}
    SC={'Positive':('2E7D32','FFFFFF'),'Negative':('C62828','FFFFFF'),'Neutral':('455A64','FFFFFF'),'Mixed':('E65100','FFFFFF')}
    def ft(sz=10,b=False,co='111111'): return FT(name='Arial',size=sz,bold=b,color=co)
    def fl(c): return PF('solid',fgColor=c)
    def bd():
        s=SD(style='thin',color=GRY); return BD(left=s,right=s,top=s,bottom=s)
    def al(h='left',v='center',w=False): return AL(horizontal=h,vertical=v,wrap_text=w)
    def gc(cols): return next((c for c in cols if c in df.columns),'')
    idc=gc(['Feedback_ID','feedback_id']); rvc=gc(['Reviewer_Name','reviewer_name'])
    dtc=gc(['Feedback_Date','feedback_date']); fbc=gc(['Feedback','feedback'])
    tpc=gc(['Topic','topic']); sgc=gc(['Suggestion','suggestion']); stc=gc(['Sentiment','sentiment'])
    src=gc(['Source','source'])
    df2=df.copy()
    if rvc:
        named=df2[df2[rvc].str.strip().ne('')].drop_duplicates(subset=[rvc],keep='first')
        unnamed=df2[df2[rvc].str.strip().eq('')]
        df2=pd.concat([named,unnamed]).sort_index().reset_index(drop=True)
    src_label=', '.join(df2[src].unique()[:3]) if src else 'N/A'
    wb=WB(); ws1=wb.active; ws1.title='Feedback Data'; ws1.sheet_view.showGridLines=False
    ws1.merge_cells(start_row=1,start_column=1,end_row=1,end_column=7)
    c=ws1.cell(row=1,column=1,value=f'  ⚡ ZEUS — {src_label}  |  {len(df2)} Entries  |  {datetime.now().strftime("%d %b %Y")}')
    c.font=ft(13,True,GOLD);c.fill=fl(DARK);c.alignment=al('left','center');ws1.row_dimensions[1].height=32
    ws1.merge_cells(start_row=2,start_column=1,end_row=2,end_column=7)
    c=ws1.cell(row=2,column=1,value=f'  Analyst: {user_name}  ·  Session: {session_id}  ·  Unique entries only')
    c.font=ft(9,False,WHITE);c.fill=fl(NAV);c.alignment=al('left','center');ws1.row_dimensions[2].height=16
    ws1.row_dimensions[3].height=4
    HDRS=['S.No','ID','Name of Feedbacker','Date','Feedback','Type','AI Assistant (Sentiment)']
    for i,w in enumerate([6,9,20,13,55,22,55],1): ws1.column_dimensions[gcl(i)].width=w
    for ci,lbl in enumerate(HDRS,1):
        c=ws1.cell(row=4,column=ci,value=lbl)
        c.font=ft(10,True,WHITE);c.fill=fl(DARK);c.border=bd();c.alignment=al('center','center')
    ws1.row_dimensions[4].height=22
    for ri,(_,row) in enumerate(df2.iterrows(),start=5):
        alt=(ri%2==0); rbg=LGT if alt else WHITE
        tp=str(row[tpc] if tpc else ''); sent=str(row[stc] if stc else 'Neutral')
        tbg,tfg=TC.get(tp,('888888','FFFFFF'))
        sug=str(row[sgc] if sgc else '')
        ai_cell=f"[{sent}] {sug}" if sent and sug else sug
        vals=[ri-4,row[idc] if idc else '',row[rvc] if rvc else '',row[dtc] if dtc else '',row[fbc] if fbc else '',tp,ai_cell]
        styles=[(rbg,DGR,9,False,'center'),(rbg,'0D47A1',9,True,'center'),(rbg,'1A237E',9,True,'center'),(rbg,DGR,9,False,'center'),(rbg,'212121',9,False,'left'),(tbg,tfg,8,True,'center'),('FFFDE7','4E342E',9,False,'left')]
        for ci,(v,(bg,fg,sz,b,h)) in enumerate(zip(vals,styles),1):
            c=ws1.cell(row=ri,column=ci,value=str(v) if v is not None else '')
            c.fill=fl(bg);c.font=ft(sz,b,fg);c.border=bd();c.alignment=al(h,'top' if h=='left' else 'center',w=(h=='left'))
        ws1.row_dimensions[ri].height=55
    ws1.freeze_panes='A5'; ws1.auto_filter.ref=f'A4:{gcl(7)}4'
    ws2=wb.create_sheet('Summary'); ws2.sheet_view.showGridLines=False
    ws2.merge_cells('A1:E1')
    c=ws2.cell(row=1,column=1,value='  📊 Topic & Sentiment Summary')
    c.font=ft(12,True,GOLD);c.fill=fl(DARK);c.alignment=al('left','center');ws2.row_dimensions[1].height=28
    for i,w in enumerate([5,28,10,12,60],1): ws2.column_dimensions[gcl(i)].width=w
    for ci,lbl in enumerate(['#','Topic','Count','Sentiment','Top Action'],1):
        c=ws2.cell(row=2,column=ci,value=lbl)
        c.font=ft(10,True,WHITE);c.fill=fl(DARK);c.border=bd();c.alignment=al('center','center')
    tc=df2[tpc].value_counts().reset_index() if tpc else pd.DataFrame(columns=['Topic','Count'])
    tc.columns=['Topic','Count']
    for ri,(_,row) in enumerate(tc.iterrows(),start=3):
        tp=row['Topic']; tbg,tfg=TC.get(tp,('888888','FFFFFF')); alt=(ri%2==0); rbg=LGT if alt else WHITE
        top_sent=''
        if stc and tpc:
            sub=df2[df2[tpc]==tp]
            top_sent=sub[stc].mode()[0] if not sub.empty else ''
        sug=df2[df2[tpc]==tp][sgc].iloc[0] if(tpc and sgc and len(df2[df2[tpc]==tp])>0) else ''
        for ci,(v,bg,fg) in enumerate(zip([ri-2,tp,int(row['Count']),top_sent,sug],[rbg,tbg,rbg,rbg,'FFFDE7'],[DGR,tfg,'1A1A2E',DGR,'4E342E']),1):
            c=ws2.cell(row=ri,column=ci,value=v)
            c.fill=fl(bg);c.font=ft(9,False,fg);c.border=bd();c.alignment=al('center','center',w=(ci==5))
        ws2.row_dimensions[ri].height=50
    if len(tc)>1:
        bar=BC(); bar.type='col'; bar.title='Feedback by Topic'; bar.style=10; bar.width=22; bar.height=12
        bar.add_data(RF(ws2,min_col=3,min_row=2,max_row=2+len(tc)),titles_from_data=True)
        bar.set_categories(RF(ws2,min_col=2,min_row=3,max_row=2+len(tc)))
        ws2.add_chart(bar,f'A{len(tc)+5}')
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

def detect_feedback_column(df):
    best_col = None
    max_length = 0

    for col in df.columns:
        if df[col].dtype == object:
            avg_len = df[col].astype(str).str.len().mean()

            # Ignore small values like time (00:00.0)
            if avg_len > max_length and avg_len > 30:
                max_length = avg_len
                best_col = col

    return best_col

def split_feedback(feedback):
    # split by sentence or logical breaks
    parts = re.split(r'\n+|\. ', feedback)
    return [p.strip() for p in parts if len(p.strip()) > 20]

def analyze_feedback_parts(feedback):
    # st.write("RAW AI OUTPUT:", raw)
    parts = split_feedback(feedback)
    # st.write(parts)

    prompt = """
Analyze each feedback part.

For each item:
1. Sentiment (Positive / Negative / Neutral)
2. Department responsible (choose best fit: Management, Operations, Kitchen, Service, HR, Training, Technology, Marketing, Facilities, etc.)
3. One actionable business recommendation

Format EXACTLY:
1. [Sentiment] [Department]:[Suggestion]
2. [Sentiment] [Department]:[Suggestion]

STRICT RULES (MANDATORY):

- Suggestion MUST be directly derived from the feedback text
- DO NOT add new ideas, strategies, or assumptions
- DO NOT generalize beyond what is explicitly mentioned
- If the issue is not mentioned → DO NOT create suggestion
- Use the SAME context and wording as feedback
- Keep suggestion under 15 words
- Maximum 1 short sentence per item

VALID EXAMPLE:
Feedback: "Staff was rude"
✔ Train staff to be polite

INVALID EXAMPLE:
Feedback: "Food was good"
❌ Introduce loyalty programs
❌ Expand menu variety
"""

    for i, p in enumerate(parts, 1):
        prompt += f"{i}. {p}\n"

    resp = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[
            {"role": "system", "content": "You are a CX expert."},
            {"role": "user", "content": prompt}

        ],
        temperature=0.2,
        max_tokens=300,
        )

    return resp.choices[0].message.content

def parse_parts_output(text):
    sentiments = []
    suggestions = []
    departments = []

    lines = [l.strip() for l in text.split("\n") if l.strip()]

    for line in lines:
        # Remove numbering like "1.", "2." etc.
        line = re.sub(r'^\d+\.\s*', '', line)

        # --- Pattern 1: [Negative] [Service]: Suggestion ---
        m1 = re.match(
            r'\[(Positive|Negative|Neutral)\]\s*\[(.*?)\]\s*:\s*(.*)',
            line,
            re.I
        )

        # --- Pattern 2: Negative Service: Suggestion ---
        m2 = re.match(
            r'(Positive|Negative|Neutral)\s+([A-Za-z]+)\s*:\s*(.*)',
            line,
            re.I
        )

        # --- Pattern 3: Negative Service Suggestion: Suggestion ---
        m3 = re.match(
            r'(Positive|Negative|Neutral)\s+([A-Za-z]+)\s+(?:Suggestion|Recommendation)\s*:\s*(.*)',
            line,
            re.I
        )

        if m1:
            sentiments.append(m1.group(1).capitalize())
            departments.append(m1.group(2).capitalize())
            suggestions.append(m1.group(3).strip())

        elif m2:
            sentiments.append(m2.group(1).capitalize())
            departments.append(m2.group(2).capitalize())
            suggestions.append(m2.group(3).strip())

        elif m3:
            sentiments.append(m3.group(1).capitalize())
            departments.append(m3.group(2).capitalize())
            suggestions.append(m3.group(3).strip())

    return sentiments, departments, suggestions

def combine_sentiment(sentiments):
    score_map = {"Positive": 1, "Neutral": 0, "Negative": -1}
    score = sum(score_map.get(s, 0) for s in sentiments)

    if score > 0:
        return "Positive"
    elif score < 0:
        return "Negative"
    else:
        return "Mixed"
    
from collections import defaultdict

def get_all_sessions():
    conn=sqlite3.connect(DB_PATH)
    df=pd.read_sql_query("SELECT session_id,user_name,created_date,created_time,source_type,total_entries,notes FROM sessions ORDER BY created_at DESC",conn)
    conn.close(); return df

def combine_suggestions(departments, suggestions):
    dept_map = defaultdict(list)

    for d, s in zip(departments, suggestions):
        dept_map[d].append(s)

    final = []
    for dept, sugs in dept_map.items():
        combined = f"{dept}: " + "; ".join(sugs)
        final.append(combined)

    return "\n".join(final)    
    
def make_db_excel():
    from openpyxl import Workbook as WB
    from openpyxl.styles import PatternFill as PF,Font as FT,Alignment as AL,Border as BD,Side as SD
    from openpyxl.utils import get_column_letter as gcl
    df=get_full_db_export()
    if df.empty: return None
    DARK='1A1A2E';NAV='0F3460';GOLD='F9A825';WHITE='FFFFFF';LGT='F5F5F5';GRY='E0E0E0';DGR='666666'
    TC={'⚡ Service Quality':('F9A825','1A1A2E'),'📦 Product Issues':('E53935','FFFFFF'),
        '💬 Communication':('0288D1','FFFFFF'),'⏱️ Speed & Delays':('F57C00','FFFFFF'),
        '💡 Innovation & Features':('7B1FA2','FFFFFF')}
    def ft(sz=10,b=False,co='111111'): return FT(name='Arial',size=sz,bold=b,color=co)
    def fl(c): return PF('solid',fgColor=c)
    def bd():
        s=SD(style='thin',color=GRY); return BD(left=s,right=s,top=s,bottom=s)
    def al(h='left',v='center',w=False): return AL(horizontal=h,vertical=v,wrap_text=w)
    wb=WB(); ws=wb.active; ws.title='All Feedback (Unique)'; ws.sheet_view.showGridLines=False
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=7)
    c=ws.cell(row=1,column=1,value=f'  🗄️ ZEUS Full Database  |  {len(df)} Entries  |  {datetime.now().strftime("%d %b %Y")}')
    c.font=ft(13,True,GOLD);c.fill=fl(DARK);c.alignment=al('left','center');ws.row_dimensions[1].height=32
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=7)
    c=ws.cell(row=2,column=1,value='  All sessions combined  ·  Unique IDs  ·  No repeated reviewers')
    c.font=ft(9,False,WHITE);c.fill=fl(NAV);c.alignment=al('left','center');ws.row_dimensions[2].height=16
    ws.row_dimensions[3].height=4
    HDRS=['S.No','ID','Name of Feedbacker','Date','Feedback','Type','AI Assistant (Sentiment)']
    for i,w in enumerate([6,9,20,13,55,22,55],1): ws.column_dimensions[gcl(i)].width=w
    for ci,lbl in enumerate(HDRS,1):
        c=ws.cell(row=4,column=ci,value=lbl)
        c.font=ft(10,True,WHITE);c.fill=fl(DARK);c.border=bd();c.alignment=al('center','center')
    ws.row_dimensions[4].height=22
    for ri,(_,row) in enumerate(df.iterrows(),start=5):
        alt=(ri%2==0); rbg=LGT if alt else WHITE
        tp=str(row.get('topic','')); tbg,tfg=TC.get(tp,('888888','FFFFFF'))
        sent=str(row.get('sentiment','Neutral')); sug=str(row.get('suggestion',''))
        ai_cell=f"[{sent}] {sug}" if sent and sug else sug
        vals=[ri-4,row.get('feedback_id',''),row.get('reviewer_name',''),row.get('feedback_date',''),row.get('feedback',''),tp,ai_cell]
        styles=[(rbg,DGR,9,False,'center'),(rbg,'0D47A1',9,True,'center'),(rbg,'1A237E',9,True,'center'),(rbg,DGR,9,False,'center'),(rbg,'212121',9,False,'left'),(tbg,tfg,8,True,'center'),('FFFDE7','4E342E',9,False,'left')]
        for ci,(v,(bg,fg,sz,b,h)) in enumerate(zip(vals,styles),1):
            c=ws.cell(row=ri,column=ci,value=str(v) if v is not None else '')
            c.fill=fl(bg);c.font=ft(sz,b,fg);c.border=bd();c.alignment=al(h,'top' if h=='left' else 'center',w=(h=='left'))
        ws.row_dimensions[ri].height=55
    ws.freeze_panes='A5'; ws.auto_filter.ref=f'A4:{gcl(7)}4'
    ws2=wb.create_sheet('By Source'); ws2.sheet_view.showGridLines=False
    ws2.merge_cells('A1:E1')
    c=ws2.cell(row=1,column=1,value='  📊 Entries by Source')
    c.font=ft(12,True,GOLD);c.fill=fl(DARK);c.alignment=al('left','center');ws2.row_dimensions[1].height=28
    for ci,lbl in enumerate(['#','Source','Entries','Named','Analyst'],1):
        c=ws2.cell(row=2,column=ci,value=lbl)
        c.font=ft(10,True,WHITE);c.fill=fl(DARK);c.border=bd();c.alignment=al('center','center')
    for i,w in enumerate([5,22,10,10,18],1): ws2.column_dimensions[gcl(i)].width=w
    grp=df.groupby('source').agg(entries=('feedback_id','count'),named=('reviewer_name',lambda x:(x.str.strip()!='').sum()),analyst=('user_name','first')).reset_index()
    for ri,(_,row) in enumerate(grp.iterrows(),start=3):
        alt=(ri%2==0); rbg=LGT if alt else WHITE
        for ci,v in enumerate([ri-2,row['source'],int(row['entries']),int(row['named']),row['analyst']],1):
            c=ws2.cell(row=ri,column=ci,value=v)
            c.fill=fl(rbg);c.font=ft(9,False,DGR);c.border=bd();c.alignment=al('center','center')
        ws2.row_dimensions[ri].height=20
    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()

def clean_action_plan(text):
    if not text:
        return text

    text = re.sub(r'\*\*', '', text)           # remove bold
    text = re.sub(r'\d+\.\s*', '', text)       # remove numbering
    text = re.sub(r'\n+', '\n', text).strip()  # clean spacing

    return text

def generate_future_action_plan(suggestions_list):
    plans = []

    for sug in suggestions_list:
        try:
            prompt = f"""
You are a CX strategist.

Based on the following operational suggestions, create a FUTURE CUSTOMER ACTION PLAN.
Department responsible (choose best fit: Management, Operations, Kitchen, Service, HR, Training, Technology, Marketing, Facilities, etc.)

Format:
Department Responsible [Suggestion]

STRICT RULES (MANDATORY):
- Maximum 5 lines ONLY
- Each line MUST be under 12 words
- One short sentence per line
- DO NOT repeat similar actions
- DO NOT expand or explain
- DO NOT add new ideas beyond input
- Merge similar points into one concise action
- Use simple, direct language

Suggestions:
{sug}

Output:
A short action plan only.
"""

            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are a CX transformation expert."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.4
            )

            plan = resp.choices[0].message.content.strip()

        except Exception as e:
            print("Future plan error:", e)
            plan = "Action plan could not be generated"

        plans.append(plan)

    return plans

# ══ SESSION STATE ══════════════════════════════════════════
for k, v in [
    ('results_df', None),
    ('analyzed', False),
    ('session_id', None),
    ('short_uuid', None),
    ('user_name', ''),
    ('source_type', ''),
    ('data', None),
    ('db_step', 1),
    ('csv_step', 1),
]:
    if k not in st.session_state:
        st.session_state[k] = v

# ══ HERO ═══════════════════════════════════════════════════
st.markdown('<div class="hero-title">⚡ ZEUS FEEDBACK ANALYZER</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="hero-sub">Customer Intelligence · Topic Modeling · AI Insights · Session History</div>',
    unsafe_allow_html=True,
)

# ── Analyst name input (inline, below hero) ────────────────
# _uname_col, _ = st.columns([2, 5])
# with _uname_col:
#     _typed_name = st.text_input(
#         "👤 Analyst Name",
#         value=st.session_state.user_name,
#         placeholder="e.g. Priya",
#         key="analyst_name_input",
#     )
#     if _typed_name:
#         st.session_state.user_name = _typed_name

# ══ TABS ═══════════════════════════════════════════════════
tab_csv, tab_url, tab_db, tab_results, tab_history = st.tabs([
    "📄 CSV Upload",
    "🔗 URL Scraper",
    "🗄️ Database Connection",
    "📊 Results",
    "🕘 History",
])

# ══ TAB 1 — CSV Upload ════════════════════════════════════
with tab_csv:
    st.markdown('<div class="card"><div class="card-title">📤 Upload Feedback CSV</div>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader("Drop CSV here", type=["csv"], label_visibility="collapsed")

    if uploaded_file:
        try:
            df_raw = pd.read_csv(uploaded_file)
            df_raw.columns = df_raw.columns.str.strip()
            df_original = df_raw.copy()

            # Detect and rename best feedback column
            feedback_col = detect_feedback_column(df_raw)
            if feedback_col:
                df_original.rename(columns={feedback_col: 'Feedback'}, inplace=True)

            st.success(f"✅ **{len(df_raw):,}** rows · Columns: `{'`, `'.join(df_raw.columns)}`")

            # Auto-select text columns, prioritising feedback-like names
            priority_keywords = ['feedback', 'review', 'comment', 'text', 'response', 'message']
            text_cols     = [c for c in df_raw.columns if df_raw[c].dtype == object]
            priority_cols = [c for c in text_cols if any(k in c.lower() for k in priority_keywords)]
            other_cols    = [c for c in text_cols if c not in priority_cols]
            selected_cols = priority_cols + other_cols[:3] or df_raw.columns[:3].tolist()

            df_combined = df_raw[selected_cols].fillna('').astype(str)
            fbs = df_combined.apply(
                lambda row: " | ".join(
                    f"{col}: {row[col]}" for col in selected_cols if row[col].strip()
                ),
                axis=1,
            ).tolist()

            ac = ['(None)'] + list(df_raw.columns)
            us, cn, cd = st.columns(3)
            with us:
                _typed_name = st.text_input(
                    "👤 Analyst Name",
                    value=st.session_state.user_name,
                    placeholder="e.g. Priya",
                    key="analyst_name_input",
                )
                if _typed_name:
                    st.session_state.user_name = _typed_name
            with cn:
                name_col = st.selectbox("Reviewer Name column (optional)", ac, index=0)
            with cd:
                date_col = st.selectbox("Date column (optional)", ac, index=0)
            

            st.dataframe(df_raw.head(5), use_container_width=True)
            notes = st.text_input("Session notes (optional)", placeholder="e.g. Q2 2025 product feedback")

            if st.button("🚀 Analyze CSV Feedback", key="csv_btn"):
                nm = st.session_state.user_name or "USER"
                sid, uid8 = gen_sid(nm)
                st.session_state.session_id  = sid
                st.session_state.short_uuid  = uid8
                revs = df_raw[name_col].astype(str).tolist() if name_col != '(None)' else [''] * len(fbs)
                dts  = df_raw[date_col].astype(str).tolist() if date_col != '(None)' else [''] * len(fbs)
                dts  = [extract_exact_date(d) or d for d in dts]
                with st.spinner("Analyzing with CCX AI..."):
                    results = build_results(fbs, "CSV Upload", revs, dts, mode='suggestion')
                if results.empty:
                    st.error("❌ No processable feedback.")
                else:
                    df_original = df_original.reset_index(drop=True)
                    results     = results.reset_index(drop=True)
                    results     = results.drop(columns=['Feedback'], errors='ignore')
                    final_df    = pd.concat([df_original, results], axis=1)

                    st.session_state.results_df = final_df
                    st.session_state.analyzed   = True

                    saved, skipped = save_entries(sid, results)
                    save_session(sid, nm, uid8, "CSV Upload", saved, notes)

                    msg = f"✅ Showing **{len(results):,}** results · **{saved:,}** new saved to DB"
                    if skipped > 0:
                        msg += f" · **{skipped}** skipped (already in DB)"
                    st.success(msg + " → switch to 📊 Results")

        except Exception as e:
            st.error(f"Error: {e}")
    st.markdown('</div>', unsafe_allow_html=True)

# ══ TAB 2 — URL Scraper (disabled) ════════════════════════
with tab_url:
    st.markdown("""
    <div style="text-align:center;padding:5rem 2rem;">
        <div style="font-size:3.5rem;margin-bottom:1rem;">🚧</div>
        <div style="font-family:'Space Mono',monospace;font-size:1rem;font-weight:700;
                    color:#f9a825;letter-spacing:3px;text-transform:uppercase;margin-bottom:0.6rem;">
            Coming Soon
        </div>
        <div style="color:#555;font-size:0.88rem;font-family:'Space Mono',monospace;
                    letter-spacing:1px;text-transform:uppercase;">
            URL Scraper is under construction
        </div>
        <div style="color:#444;font-size:0.8rem;margin-top:0.8rem;">
            Use <strong style="color:#f9a825;">CSV Upload</strong> or
            <strong style="color:#f9a825;">Database Connection</strong> in the meantime.
        </div>
    </div>
    """, unsafe_allow_html=True)

# ══ TAB 3 — Database Connection ════════════════════════════
with tab_db:
        # STEPS = ["Configure & Connect", "Select, Detect & Analyze"]
        # render_step_bar(st.session_state.db_step, STEPS)

        # ──────────────────────── STEP 1 ─────────────────────────────────────
        if st.session_state.db_step == 1:
            st.markdown('<div class="card"><div class="card-title">🔌 Step 1 — Configure Database Connection</div>', unsafe_allow_html=True)
            db_type = st.selectbox("Database Type", ALL_DB_TYPES, key="db_type_select")
            st.session_state.db_type = db_type
            pkg, avail = DB_DRIVERS[db_type]
            if not avail and pkg != "built-in":
                st.warning(f"⚠️ Driver missing — `pip install {pkg}`")

            conn_params = {}

            if db_type == "SQLite (file)":
                f = st.file_uploader("Upload .db / .sqlite file", type=["db","sqlite","sqlite3"], key="sqlite_upload")
                conn_params = {"file": f}

            elif db_type == "MySQL / MariaDB":
                c1,c2 = st.columns([3,1])
                with c1: h  = st.text_input("Host","localhost",key="my_h")
                with c2: p  = st.text_input("Port","3306",key="my_p")
                c3,c4 = st.columns(2)
                with c3: u  = st.text_input("Username",key="my_u")
                with c4: pw = st.text_input("Password",type="password",key="my_pw")
                db = st.text_input("Database name",key="my_db")
                conn_params = dict(host=h,port=p,user=u,password=pw,database=db)

            elif db_type == "PostgreSQL":
                c1,c2 = st.columns([3,1])
                with c1: h  = st.text_input("Host","localhost",key="pg_h")
                with c2: p  = st.text_input("Port","5432",key="pg_p")
                c3,c4 = st.columns(2)
                with c3: u  = st.text_input("Username",key="pg_u")
                with c4: pw = st.text_input("Password",type="password",key="pg_pw")
                db = st.text_input("Database name",key="pg_db")
                conn_params = dict(host=h,port=p,user=u,password=pw,database=db)

            elif db_type == "SQL Server (MSSQL)":
                c1,c2 = st.columns(2)
                with c1: h  = st.text_input("Host / Server",key="ms_h")
                with c2: p  = st.text_input("Port","1433",key="ms_p")
                c3,c4 = st.columns(2)
                with c3: u  = st.text_input("Username",key="ms_u")
                with c4: pw = st.text_input("Password",type="password",key="ms_pw")
                c5, c6 = st.columns(2)
                with c5: sc = st.text_input("Schema","dbo",key="ms_sc")
                with c6: db = st.text_input("Database name",key="ms_db")
                conn_params = dict(host=h,port=p,user=u,password=pw,database=db,schema=sc)

            elif db_type == "Snowflake":
                acct = st.text_input("Account (e.g. xy12345.us-east-1)",key="sf_acct")
                c1,c2 = st.columns(2)
                with c1: u  = st.text_input("Username",key="sf_u")
                with c2: pw = st.text_input("Password",type="password",key="sf_pw")
                c3,c4,c5 = st.columns(3)
                with c3: wh = st.text_input("Warehouse",key="sf_wh")
                with c4: db = st.text_input("Database",key="sf_db")
                with c5: sc = st.text_input("Schema","PUBLIC",key="sf_sc")
                conn_params = dict(account=acct,user=u,password=pw,warehouse=wh,database=db,schema=sc)

            elif db_type == "MongoDB":
                st.markdown('<div style="background:rgba(249,168,37,0.06);border:1px solid rgba(249,168,37,0.25);border-radius:8px;padding:0.7rem 1rem;margin-bottom:0.8rem;font-family:Space Mono,monospace;font-size:0.7rem;color:#f9a825;">⚡ Special characters auto-escaped — paste URI as-is.</div>', unsafe_allow_html=True)
                mongo_mode = st.radio("Input mode", ["📋 Full URI", "🔑 Host + Credentials"], horizontal=True, key="mg_mode")
                if mongo_mode == "📋 Full URI":
                    uri   = st.text_input("Connection URI", placeholder="mongodb+srv://user:pass@cluster0.xxxxx.mongodb.net/", key="mg_uri")
                    mg_db = st.text_input("Database name", key="mg_db")
                    conn_params = dict(uri=uri, database=mg_db)
                else:
                    from urllib.parse import quote_plus
                    mg_scheme = st.selectbox("Scheme", ["mongodb+srv","mongodb"], key="mg_scheme")
                    c1,c2 = st.columns(2)
                    with c1: mg_user = st.text_input("Username", key="mg_user_split")
                    with c2: mg_pass = st.text_input("Password", type="password", key="mg_pass_split")
                    mg_host = st.text_input("Host / Cluster", placeholder="cluster0.xxxxx.mongodb.net", key="mg_host_split")
                    mg_db   = st.text_input("Database name", key="mg_db_split")
                    enc_u   = quote_plus(mg_user) if mg_user else ""
                    enc_p   = quote_plus(mg_pass) if mg_pass else ""
                    cred    = f"{enc_u}:{enc_p}" if enc_p else enc_u
                    built_uri = f"{mg_scheme}://{cred}@{mg_host}/" if enc_u and mg_host else ""
                    conn_params = dict(uri=built_uri, database=mg_db)

            elif db_type == "Google BigQuery":
                proj      = st.text_input("GCP Project ID", key="bq_proj")
                cred_file = st.file_uploader("Service Account JSON (optional)", type=["json"], key="bq_cred")
                cred_json = cred_file.read().decode() if cred_file else None
                conn_params = dict(project=proj, credentials_json=cred_json)

            elif db_type == "Google Analytics 4":
                prop_id   = st.text_input("GA4 Property ID (numeric)", key="ga_prop")
                cred_file = st.file_uploader("Service Account JSON", type=["json"], key="ga_cred")
                cred_json = cred_file.read().decode() if cred_file else None
                c1,c2 = st.columns(2)
                with c1: ga_start = st.text_input("Start date","30daysAgo",key="ga_start")
                with c2: ga_end   = st.text_input("End date","today",key="ga_end")
                conn_params = dict(property_id=prop_id, credentials_json=cred_json,
                                   start_date=ga_start, end_date=ga_end)

            elif db_type == "Databricks SQL":
                db_host = st.text_input("Server Hostname", placeholder="adb-xxxx.azuredatabricks.net", key="dbr_host")
                db_http = st.text_input("HTTP Path", placeholder="/sql/1.0/warehouses/xxxx", key="dbr_http")
                db_tok  = st.text_input("Access Token", type="password", key="dbr_tok")
                c1,c2 = st.columns(2)
                with c1: db_cat = st.text_input("Catalog","main",key="dbr_cat")
                with c2: db_sc  = st.text_input("Schema","default",key="dbr_sc")
                conn_params = dict(host=db_host,http_path=db_http,token=db_tok,
                                   catalog=db_cat,schema=db_sc)

            st.session_state.db_conn_params = conn_params

            if st.button("🔌 Connect & Discover Tables →", key="db_connect_btn"):
                with st.spinner("Connecting…"):
                    try:
                        tables = []; src_label = db_type.split()[0]

                        if db_type == "SQLite (file)":
                            if not conn_params.get("file"): st.error("❌ Upload a SQLite file first."); st.stop()
                            tmp = os.path.join(tempfile.gettempdir(), f"zeus_{uuid.uuid4().hex[:8]}.db")
                            with open(tmp,"wb") as fh: fh.write(conn_params["file"].read())
                            st.session_state.db_tmp_path = tmp
                            tables = _list_sqlite_tables(tmp); src_label = "SQLite"

                        elif db_type == "MySQL / MariaDB":
                            tables = _list_mysql_tables(conn_params['host'],conn_params['port'],conn_params['user'],conn_params['password'],conn_params['database']); src_label = "MySQL"

                        elif db_type == "PostgreSQL":
                            tables = _list_postgres_tables(conn_params['host'],conn_params['port'],conn_params['user'],conn_params['password'],conn_params['database']); src_label = "PostgreSQL"

                        elif db_type == "SQL Server (MSSQL)":
                            tables = _list_mssql_tables(conn_params['host'],conn_params['port'],conn_params['user'],conn_params['password'],conn_params['database'],conn_params['schema']); src_label = "SQLServer"

                        elif db_type == "Snowflake":
                            tables = _list_snowflake_tables(conn_params['account'],conn_params['user'],conn_params['password'],conn_params['warehouse'],conn_params['database'],conn_params['schema']); src_label = "Snowflake"

                        elif db_type == "MongoDB":
                            tables = _list_mongodb_collections(conn_params['uri'],conn_params['database']); src_label = "MongoDB"

                        elif db_type == "Google BigQuery":
                            tables = _list_bigquery_datasets(conn_params['project'],conn_params.get('credentials_json')); src_label = "BigQuery"

                        elif db_type == "Google Analytics 4":
                            tables = ["(GA4 engagement report)"]; src_label = "GA4"

                        elif db_type == "Databricks SQL":
                            tables = _list_databricks_tables(conn_params['host'],conn_params['http_path'],conn_params['token'],conn_params.get('catalog','main'),conn_params.get('schema','default')); src_label = "Databricks"

                        st.session_state.db_available_tables = tables
                        st.session_state.db_src_label        = src_label
                        st.session_state.db_connected        = True
                        st.session_state.db_step             = 2
                        st.success(f"✅ Connected to **{src_label}** — found **{len(tables)}** object(s)")
                        st.rerun()

                    except ImportError as e: st.error(f"❌ Missing driver: {e}")
                    except Exception as e:   st.error(f"❌ Connection failed: {str(e)[:300]}")

            st.markdown('</div>', unsafe_allow_html=True)

        # ──────────────────────── STEP 2 ─────────────────────────────────────
        # Select table → Fetch ALL rows → Show metadata → Smart detect → Analyze
        # ─────────────────────────────────────────────────────────────────────
        elif st.session_state.db_step == 2:
            tables      = st.session_state.db_available_tables
            src_label   = st.session_state.db_src_label
            db_type     = st.session_state.db_type
            conn_params = st.session_state.db_conn_params
            tmp_path    = st.session_state.get("db_tmp_path","")

            st.markdown(
                f'<div class="session-box">✅ Connected to <strong>{src_label}</strong>'
                f' &nbsp;·&nbsp; <strong>{len(tables)}</strong> object(s) found</div>',
                unsafe_allow_html=True
            )

            st.markdown('<div class="card"><div class="card-title">📂 Step 2 — Select Table, Inspect Metadata & Analyze</div>', unsafe_allow_html=True)

            # ── Table selector ────────────────────────────────────────────────
            selected_table = st.selectbox(
                "Select table / collection", tables, key="db_table_select"
            )
            st.session_state.db_selected_table = selected_table

            # ── Fetch button ──────────────────────────────────────────────────
            if st.button("🔍 Fetch All Rows & Inspect", key="s2_fetch_all"):
                with st.spinner(f"Fetching all rows from **{selected_table}**…"):
                    try:
                        raw_df = fetch_all_rows(db_type, conn_params, selected_table, tmp_path)
                        if raw_df is None or raw_df.empty:
                            st.error("❌ Table is empty or returned no rows."); st.stop()

                        # Build metadata FIRST (uses raw_df for full stats)
                        with st.spinner("📊 Building column metadata…"):
                            meta = get_table_metadata(db_type, conn_params, selected_table,
                                                      tmp_path, raw_df=raw_df)

                        # Smart detection using pre-defined dictionary + data
                        with st.spinner("🤖 Detecting feedback column…"):
                            detected = smart_detect_columns(raw_df)

                        st.session_state.db_raw_df        = raw_df
                        st.session_state.db_detected_cols = detected
                        st.session_state.db_metadata      = meta
                        st.rerun()

                    except Exception as ex:
                        st.error(f"❌ {str(ex)[:400]}")

            # ── Show full panel once data loaded ──────────────────────────────
            raw_df   = st.session_state.get("db_raw_df")
            detected = st.session_state.get("db_detected_cols", {})
            meta     = st.session_state.get("db_metadata", [])

            if raw_df is not None and not raw_df.empty and detected:
                total_rows = len(raw_df); total_cols = len(raw_df.columns)

                # ── Stats banner ──────────────────────────────────────────────
                m1,m2,m3,m4 = st.columns(4)
                m1.metric("Total Rows",    f"{total_rows:,}")
                m2.metric("Total Columns", f"{total_cols}")
                m3.metric("Table",         selected_table)
                m4.metric("Source",        src_label)

                st.divider()

                # ── METADATA PANEL ────────────────────────────────────────────
                with st.expander("🗂️ Column Metadata Inspector", expanded=True):
                    render_metadata_panel(meta, detected_fb_col=detected.get("feedback_col",""))

                st.divider()

                # ── Detection result banner ───────────────────────────────────
                layer      = detected.get("layer_used", "Unknown")
                conf_score = detected.get("confidence_score", 0)
                reason     = detected.get("reason", "")
                layer_colors = {
                    "Layer 1": "#4caf50", "Layer 2": "#f9a825",
                    "Layer 3": "#00bcd4", "Layer 4": "#e91e63",
                }
                lc = next((v for k, v in layer_colors.items() if k in layer), "#888")

                st.markdown(
                    f'<div style="background:rgba(0,0,0,0.25);border:1px solid {lc}40;'
                    f'border-radius:10px;padding:0.9rem 1.2rem;margin-bottom:1rem;">'
                    f'<div style="display:flex;align-items:center;gap:1.5rem;flex-wrap:wrap;">'
                    f'<span style="font-family:Space Mono,monospace;font-size:0.72rem;'
                    f'color:{lc};font-weight:700;">🎯 {layer}</span>'
                    f'<span style="font-family:Space Mono,monospace;font-size:0.8rem;'
                    f'color:#f9a825;font-weight:800;">Confidence: {conf_score}/100</span>'
                    f'</div>'
                    f'<div style="font-size:0.8rem;color:#bbb;margin-top:4px;">{reason}</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )

                # ── Per-column confidence bars ────────────────────────────────
                scores = detected.get("column_scores", {})
                if scores:
                    st.markdown('<div style="font-family:Space Mono,monospace;font-size:0.7rem;color:#888;margin-bottom:0.5rem;letter-spacing:1px;text-transform:uppercase;">Feedback probability per column</div>', unsafe_allow_html=True)
                    sorted_scores = sorted(scores.items(), key=lambda x: x[1], reverse=True)[:6]
                    bar_cols = st.columns(len(sorted_scores))
                    for i, (col, score) in enumerate(sorted_scores):
                        is_top = i == 0
                        css = "conf-card-top" if is_top else "conf-card"
                        clr = "#f9a825" if is_top else "#aaa"
                        with bar_cols[i]:
                            st.markdown(
                                f'<div class="{css}">'
                                f'<div style="font-family:Space Mono,monospace;font-size:0.62rem;color:{clr};margin-bottom:3px;">{col[:16]}</div>'
                                f'<div style="font-size:1.1rem;font-weight:800;color:#f9a825;font-family:Space Mono,monospace;">{score}%</div>'
                                f'<div style="background:rgba(249,168,37,0.12);border-radius:3px;height:4px;margin-top:4px;">'
                                f'<div style="background:#f9a825;width:{score}%;height:4px;border-radius:3px;"></div></div></div>',
                                unsafe_allow_html=True
                            )

                st.divider()

                # ── Column mapping confirmation ────────────────────────────────
                st.markdown("**📌 Confirm column mapping** *(auto-detected — override if needed)*")
                all_cols = ["(none)"] + list(raw_df.columns)
                col_list = list(raw_df.columns)

                def _si(lst, val):
                    try: return lst.index(val) if val in lst else 0
                    except: return 0

                cm1,cm2,cm3 = st.columns(3)
                with cm1:
                    fb_col = st.selectbox("📝 Feedback column *", col_list,
                        index=_si(col_list, detected.get("feedback_col") or ""), key="final_fb_col")
                with cm2:
                    nm_col = st.selectbox("👤 Reviewer name", all_cols,
                        index=_si(all_cols, detected.get("name_col") or "(none)"), key="final_nm_col")
                with cm3:
                    dt_col = st.selectbox("📅 Date column", all_cols,
                        index=_si(all_cols, detected.get("date_col") or "(none)"), key="final_dt_col")

                # ── Sample preview ────────────────────────────────────────────
                if fb_col:
                    samples = raw_df[fb_col].dropna().astype(str).head(3).tolist()
                    sample_html = "".join(
                        f'<div style="font-size:0.8rem;color:#ccc;padding:4px 0;border-bottom:1px solid rgba(255,255,255,0.05);">"{s[:250]}"</div>'
                        for s in samples
                    )
                    st.markdown(
                        f'<div style="background:rgba(249,168,37,0.04);border:1px solid rgba(249,168,37,0.2);'
                        f'border-radius:8px;padding:0.8rem 1.2rem;margin:0.8rem 0;">'
                        f'<div style="font-family:Space Mono,monospace;font-size:0.68rem;color:#f9a825;margin-bottom:6px;">'
                        f'📋 SAMPLE VALUES — {fb_col}</div>{sample_html}</div>',
                        unsafe_allow_html=True
                    )

                # ── Summary confirmation line ──────────────────────────────────
                st.markdown(
                    f'<div style="background:rgba(249,168,37,0.07);border:1px solid rgba(249,168,37,0.3);'
                    f'border-radius:8px;padding:0.8rem 1.2rem;margin:0.8rem 0;">'
                    f'<div style="font-family:Space Mono,monospace;font-size:0.72rem;color:#f9a825;">'
                    f'Will analyze: <strong>{fb_col}</strong> as feedback'
                    f'{f" · <strong>{nm_col}</strong> as name" if nm_col != "(none)" else " · AI will extract names"}'
                    f'{f" · <strong>{dt_col}</strong> as date" if dt_col != "(none)" else ""}'
                    f' · <strong>{total_rows:,}</strong> total rows'
                    f'</div></div>',
                    unsafe_allow_html=True
                )

                db_notes = st.text_input("Session notes (optional)",
                    placeholder="e.g. Production DB — Q2 reviews", key="db_notes_s2")

                b1, b2 = st.columns([1,4])
                with b1:
                    if st.button("← Back", key="s2_back"):
                        st.session_state.db_step          = 1
                        st.session_state.db_connected     = False
                        st.session_state.db_raw_df        = None
                        st.session_state.db_detected_cols = {}
                        st.session_state.db_metadata      = []
                        st.rerun()
                with b2:
                    if st.button("🚀 Run Feedback Analysis →", key="s2_analyze"):
                        nm = st.session_state.user_name or "USER"
                        sid, uid8 = gen_sid(nm)
                        st.session_state.session_id = sid
                        st.session_state.short_uuid = uid8

                        fbs  = raw_df[fb_col].dropna().astype(str).tolist()
                        revs = raw_df[nm_col].astype(str).tolist() if nm_col != "(none)" else ['']*len(fbs)
                        dts  = raw_df[dt_col].astype(str).tolist() if dt_col != "(none)" else ['']*len(fbs)
                        dts  = [extract_exact_date(d) or d for d in dts]

                        with st.spinner("Analyzing feedback…"):
                            results = build_results(fbs, src_label, revs, dts, mode='suggestion')

                        if results.empty:
                            st.error("❌ No processable feedback found.")
                        else:
                            st.session_state.results_df = results
                            st.session_state.analyzed   = True
                            saved, skipped = save_entries(sid, results)
                            save_session(sid, nm, uid8, src_label, saved, db_notes)
                            msg = f"✅ **{len(results):,}** results · **{saved:,}** saved"
                            if skipped > 0: msg += f" · **{skipped}** duplicates skipped"
                            st.success(msg + " → switch to **📊 Results** tab")
                            st.session_state.db_step          = 1
                            st.session_state.db_connected     = False
                            st.session_state.db_raw_df        = None
                            st.session_state.db_detected_cols = {}
                            st.session_state.db_metadata      = []
                            st.rerun()

            st.markdown('</div>', unsafe_allow_html=True)
    # st.markdown('<div class="card"><div class="card-title">🔌 Database Connection</div>', unsafe_allow_html=True)

    # DB_OPTIONS = ["-- Select a Database --", "MySQL", "SQL Server", "Oracle", "MongoDB", "SQLite"]
    # selected_db = st.selectbox("Select Database Type", DB_OPTIONS, index=0)

    # if selected_db != "-- Select a Database --":
    #     st.markdown(f"### 🔌 Connect to {selected_db}")

    #     # ── MySQL ──────────────────────────────────────────────
    #     if selected_db == "MySQL":
    #         col1, col2 = st.columns(2)
    #         with col1:
    #             mysql_host     = st.text_input("Host",     placeholder="e.g. localhost or 192.168.1.1", key="mysql_host")
    #             mysql_port     = st.text_input("Port",     value="3306",                                key="mysql_port")
    #             mysql_database = st.text_input("Database", placeholder="e.g. my_database",             key="mysql_db")
    #         with col2:
    #             mysql_user     = st.text_input("Username", placeholder="e.g. root",                    key="mysql_user")
    #             mysql_password = st.text_input("Password", type="password",                             key="mysql_password")
    #         mysql_query = st.text_area("SQL Query (optional)", placeholder="SELECT * FROM feedback LIMIT 100", key="mysql_query")

    #         if st.button("🔗 Connect & Fetch", key="mysql_btn"):
    #             if mysql_host and mysql_user and mysql_database:
    #                 with st.spinner("Connecting to MySQL..."):
    #                     try:
    #                         import mysql.connector
    #                         conn = mysql.connector.connect(
    #                             host=mysql_host,
    #                             port=int(mysql_port or 3306),
    #                             database=mysql_database,
    #                             user=mysql_user,
    #                             password=mysql_password,
    #                         )
    #                         query = mysql_query.strip() or "SELECT * FROM feedback LIMIT 100"
    #                         st.session_state.data = pd.read_sql(query, conn)
    #                         conn.close()
    #                         st.success(f"✅ Connected to MySQL · {len(st.session_state.data):,} rows fetched")
    #                         st.dataframe(st.session_state.data.head(10), use_container_width=True)
    #                     except Exception as e:
    #                         st.error(f"❌ MySQL connection failed: {e}")
    #             else:
    #                 st.warning("⚠️ Please fill in Host, Username, and Database.")

    #     # ── SQL Server ─────────────────────────────────────────
    #     elif selected_db == "SQL Server":
    #         col1, col2 = st.columns(2)
    #         with col1:
    #             mssql_host     = st.text_input("Server / Host",  placeholder="e.g. localhost\\SQLEXPRESS", key="mssql_host")
    #             mssql_port     = st.text_input("Port",           value="1433",                            key="mssql_port")
    #             mssql_database = st.text_input("Database",       placeholder="e.g. FeedbackDB",           key="mssql_db")
    #         with col2:
    #             mssql_user     = st.text_input("Username",       placeholder="e.g. sa",                   key="mssql_user")
    #             mssql_password = st.text_input("Password",       type="password",                         key="mssql_password")
    #             mssql_driver   = st.text_input("ODBC Driver",    value="ODBC Driver 17 for SQL Server",   key="mssql_driver")
    #         mssql_query = st.text_area("SQL Query (optional)", placeholder="SELECT TOP 100 * FROM feedback", key="mssql_query")

    #         if st.button("🔗 Connect & Fetch", key="mssql_btn"):
    #             if mssql_host and mssql_user and mssql_database:
    #                 with st.spinner("Connecting to SQL Server..."):
    #                     try:
    #                         import pyodbc
    #                         conn_str = (
    #                             f"DRIVER={{{mssql_driver}}};"
    #                             f"SERVER={mssql_host},{mssql_port or 1433};"
    #                             f"DATABASE={mssql_database};"
    #                             f"UID={mssql_user};"
    #                             f"PWD={mssql_password}"
    #                         )
    #                         conn = pyodbc.connect(conn_str)
    #                         query = mssql_query.strip() or "SELECT TOP 100 * FROM feedback"
    #                         st.session_state.data = pd.read_sql(query, conn)
    #                         conn.close()
    #                         st.success(f"✅ Connected to SQL Server · {len(st.session_state.data):,} rows fetched")
    #                         st.dataframe(st.session_state.data.head(10), use_container_width=True)
    #                     except Exception as e:
    #                         st.error(f"❌ SQL Server connection failed: {e}")
    #             else:
    #                 st.warning("⚠️ Please fill in Server, Username, and Database.")

    #     # ── Oracle ─────────────────────────────────────────────
    #     elif selected_db == "Oracle":
    #         col1, col2 = st.columns(2)
    #         with col1:
    #             oracle_host    = st.text_input("Host",               placeholder="e.g. localhost",    key="oracle_host")
    #             oracle_port    = st.text_input("Port",               value="1521",                    key="oracle_port")
    #             oracle_service = st.text_input("Service Name / SID", placeholder="e.g. ORCL",         key="oracle_service")
    #         with col2:
    #             oracle_user     = st.text_input("Username", placeholder="e.g. system",  key="oracle_user")
    #             oracle_password = st.text_input("Password", type="password",            key="oracle_password")
    #         oracle_query = st.text_area(
    #             "SQL Query (optional)",
    #             placeholder="SELECT * FROM feedback WHERE ROWNUM <= 100",
    #             key="oracle_query",
    #         )

    #         if st.button("🔗 Connect & Fetch", key="oracle_btn"):
    #             if oracle_host and oracle_user and oracle_service:
    #                 with st.spinner("Connecting to Oracle..."):
    #                     try:
    #                         import cx_Oracle
    #                         dsn  = cx_Oracle.makedsn(oracle_host, int(oracle_port or 1521), service_name=oracle_service)
    #                         conn = cx_Oracle.connect(user=oracle_user, password=oracle_password, dsn=dsn)
    #                         query = oracle_query.strip() or "SELECT * FROM feedback WHERE ROWNUM <= 100"
    #                         st.session_state.data = pd.read_sql(query, conn)
    #                         conn.close()
    #                         st.success(f"✅ Connected to Oracle · {len(st.session_state.data):,} rows fetched")
    #                         st.dataframe(st.session_state.data.head(10), use_container_width=True)
    #                     except Exception as e:
    #                         st.error(f"❌ Oracle connection failed: {e}")
    #             else:
    #                 st.warning("⚠️ Please fill in Host, Username, and Service Name.")

    #     # ── MongoDB ────────────────────────────────────────────
    #     elif selected_db == "MongoDB":
    #         col1, col2 = st.columns(2)
    #         with col1:
    #             mongo_host       = st.text_input("Host",            placeholder="e.g. localhost or cluster.mongodb.net", key="mongo_host")
    #             mongo_port       = st.text_input("Port",            value="27017",                                        key="mongo_port")
    #             mongo_database   = st.text_input("Database",        placeholder="e.g. feedback_db",                      key="mongo_db")
    #         with col2:
    #             mongo_user       = st.text_input("Username",        placeholder="e.g. admin (leave blank if none)",      key="mongo_user")
    #             mongo_password   = st.text_input("Password",        type="password",                                     key="mongo_password")
    #             mongo_collection = st.text_input("Collection Name", placeholder="e.g. reviews",                         key="mongo_col")
    #         mongo_limit = st.number_input(
    #             "Max Documents to Fetch", min_value=1, max_value=10000, value=100, step=50, key="mongo_limit"
    #         )

    #         if st.button("🔗 Connect & Fetch", key="mongo_btn"):
    #             if mongo_host and mongo_database and mongo_collection:
    #                 with st.spinner("Connecting to MongoDB..."):
    #                     try:
    #                         from pymongo import MongoClient
    #                         if mongo_user and mongo_password:
    #                             uri = f"mongodb://{mongo_user}:{mongo_password}@{mongo_host}:{mongo_port or 27017}/"
    #                         else:
    #                             uri = f"mongodb://{mongo_host}:{mongo_port or 27017}/"
    #                         client = MongoClient(uri)
    #                         db_mongo = client[mongo_database]
    #                         col_mongo = db_mongo[mongo_collection]
    #                         docs = list(col_mongo.find({}, {"_id": 0}).limit(int(mongo_limit)))
    #                         if docs:
    #                             st.session_state.data = pd.DataFrame(docs)
    #                             st.success(f"✅ Connected to MongoDB · {len(st.session_state.data):,} documents fetched")
    #                             st.dataframe(st.session_state.data.head(10), use_container_width=True)
    #                         else:
    #                             st.warning("⚠️ Collection is empty or no documents match.")
    #                         client.close()
    #                     except Exception as e:
    #                         st.error(f"❌ MongoDB connection failed: {e}")
    #             else:
    #                 st.warning("⚠️ Please fill in Host, Database, and Collection Name.")

    #     # ── SQLite ─────────────────────────────────────────────
    #     elif selected_db == "SQLite":
    #         sqlite_file = st.file_uploader(
    #             "Upload SQLite Database File (.db / .sqlite)",
    #             type=["db", "sqlite", "sqlite3"],
    #             key="sqlite_file",
    #         )
    #         if sqlite_file:
    #             sqlite_query = st.text_area(
    #                 "SQL Query",
    #                 placeholder="SELECT * FROM feedback LIMIT 100",
    #                 key="sqlite_query",
    #             )
    #             if st.button("📋 List Tables", key="sqlite_tables_btn"):
    #                 try:
    #                     
    #                     with tempfile.NamedTemporaryFile(delete=False, suffix=".db") as tmp:
    #                         tmp.write(sqlite_file.read())
    #                         tmp_path = tmp.name
    #                     conn_tmp = sqlite3.connect(tmp_path)
    #                     tables_df = pd.read_sql_query(
    #                         "SELECT name FROM sqlite_master WHERE type='table'", conn_tmp
    #                     )
    #                     conn_tmp.close()
    #                     st.info(f"📌 Tables found: **{', '.join(tables_df['name'].tolist())}**")
    #                     sqlite_file.seek(0)
    #                 except Exception as e:
    #                     st.error(f"❌ Could not read tables: {e}")

    #             if st.button("🔗 Connect & Fetch", key="sqlite_btn"):
    #                 if sqlite_query.strip():
    #                     with st.spinner("Reading SQLite database..."):
    #                         try:
    #                             import tempfile
    #                             with tempfile.NamedTemporaryFile(delete=False, suffix=".db") as tmp:
    #                                 tmp.write(sqlite_file.read())
    #                                 tmp_path = tmp.name
    #                             conn_tmp = sqlite3.connect(tmp_path)
    #                             st.session_state.data = pd.read_sql_query(sqlite_query.strip(), conn_tmp)
    #                             conn_tmp.close()
    #                             st.success(f"✅ Data fetched from SQLite · {len(st.session_state.data):,} rows")
    #                             st.dataframe(st.session_state.data.head(10), use_container_width=True)
    #                         except Exception as e:
    #                             st.error(f"❌ SQLite query failed: {e}")
    #                 else:
    #                     st.warning("⚠️ Please enter a SQL query.")
    #         else:
    #             st.info("📂 Please upload a `.db` or `.sqlite` file to get started.")

    # st.markdown('</div>', unsafe_allow_html=True)

# ══ TAB 4 — Results ════════════════════════════════════════
# (alias so the results block below can reference tab_results cleanly)
tab1 = tab_results
tab2 = tab_history

# for tab in [tab1, tab2]:
with tab1:
    df = st.session_state.get("results_df")

    # -------------------------------
    # SAFETY CHECK
    # -------------------------------
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        st.markdown("""
        <div style="text-align:center;padding:4rem;">
            <div style="font-size:4rem;">📭</div>
            <div style="font-family:'Space Mono',monospace;font-size:0.9rem;color:#555;
                        letter-spacing:2px;text-transform:uppercase;margin-top:1rem;">
                No results yet
            </div>
            <div style="color:#444;font-size:0.85rem;margin-top:0.5rem;">
                Upload a CSV or scrape a URL first
            </div>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    # -------------------------------
    # SAFE TO USE DF NOW
    # -------------------------------
    total = len(df)
    sid   = st.session_state.session_id or "—"
    uname = st.session_state.user_name or "—"
    tc    = df['Topic'].value_counts()

    st.markdown(
        f'<div class="session-box">🔑 <strong>{sid}</strong> &nbsp;·&nbsp; '
        f'👤 {uname} &nbsp;·&nbsp; '
        f'📅 {datetime.now().strftime("%d %b %Y, %H:%M")} &nbsp;·&nbsp; '
        f'📊 {total} entries</div>',
        unsafe_allow_html=True,
    )

    c1, c3, c4, c5 = st.columns(4)
    c1.metric("Total Entries", f"{total:,}")
    # c2.metric("Topics", df['Topic'].nunique())
    c3.metric("Sources", df['Source'].nunique())
    named = (
        df['Reviewer_Name'].astype(str).str.strip().ne('').sum()
        if 'Reviewer_Name' in df.columns else 0
    )
    c4.metric("Named Reviewers", named)
    pos = df['Sentiment'].eq('Positive').sum() if 'Sentiment' in df.columns else 0
    c5.metric("Positive", pos)

    # st.divider()
    # st.markdown("### 🗂️ Topic Distribution")
    cols = st.columns(min(len(tc), 5))
    # for i, (topic, count) in enumerate(tc.items()):
    #     with cols[i % 5]:
    #         st.markdown(f"""
    #         <div style="background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);
    #                     border-radius:12px;padding:1rem;text-align:center;">
    #             <div style="font-size:1.6rem;font-weight:800;color:#f9a825;
    #                         font-family:'Space Mono',monospace;">{count}</div>
    #             <div style="font-size:0.68rem;color:#aaa;margin:4px 0;">{topic}</div>
    #             <div style="font-size:0.75rem;color:#666;">{round(count / total * 100)}%</div>
    #         </div>""", unsafe_allow_html=True)

    st.divider()
    st.markdown("### 🔍 Browse & Filter")
    cf1, cf2, cf3 = st.columns(3)
    with cf1:
        tf = st.multiselect("Topic",  df['Topic'].unique().tolist(),  default=df['Topic'].unique().tolist())
    with cf2:
        sf = st.multiselect("Source", df['Source'].unique().tolist(), default=df['Source'].unique().tolist())
    with cf3:
        search_q = st.text_input("🔎 Search feedback", placeholder="keyword...")

    filtered = df[df['Topic'].isin(tf) & df['Source'].isin(sf)]
    if search_q.strip():
        filtered = filtered[filtered['Feedback'].str.contains(search_q.strip(), case=False, na=False)]

    st.markdown(
        f'<div style="font-family:Space Mono,monospace;font-size:0.75rem;color:#888;'
        f'margin-bottom:1rem;">Showing {len(filtered):,} of {total:,}</div>',
        unsafe_allow_html=True,
    )

    view = st.radio("View", ["📋 Table", "🃏 Cards"], horizontal=True)

    if view == "📋 Table":
        dcols = [
            c for c in
            ['Feedback_ID', 'Source', 'Reviewer_Name', 'Feedback_Date', 'Feedback', 'Sentiment', 'Suggestion','Future_Action_Plan']
            if c in filtered.columns
        ]
        st.dataframe(
            filtered[dcols],
            use_container_width=True,
            height=420,
            column_config={
                "Feedback_ID": st.column_config.TextColumn("ID",          width="small"),
                "Feedback":    st.column_config.TextColumn("Feedback",    width="large"),
                "Suggestion":  st.column_config.TextColumn("AI Suggestion", width="large"),
                "Sentiment":   st.column_config.TextColumn("Sentiment",  width="small"),
                # "Topic":       st.column_config.TextColumn("Topic",      width="medium"),
                "Reviewer_Name": st.column_config.TextColumn("Reviewer", width="small"),
                "Feedback_Date": st.column_config.TextColumn("Date",     width="small"),
                "Future_Action_Plan": st.column_config.TextColumn("Future Action Plan", width="large"),
            },
        )
    else:
        for _, row in filtered.head(30).iterrows():
            tid  = [k for k, v in TOPIC_LABELS.items() if v == row['Topic']]
            bc   = TOPIC_COLORS.get(tid[0] if tid else 0, "badge-0")
            fid  = row.get('Feedback_ID', '')
            name = row.get('Reviewer_Name', '')
            date = row.get('Feedback_Date', '')
            sent = row.get('Sentiment', '')
            sc   = {'Positive': '#4caf50', 'Negative': '#e91e63', 'Neutral': '#aaa', 'Mixed': '#ff9800'}.get(sent, '#aaa')
            meta = []
            if name and str(name).strip():
                meta.append(f"👤 {name}")
            if date and str(date).strip():
                meta.append(f"📅 {date}")
            mh = "  &nbsp;·&nbsp;  ".join(meta)
            st.markdown(f"""
            <div class="card">
                <div style="display:flex;align-items:center;gap:8px;margin-bottom:0.8rem;flex-wrap:wrap;">
                    <span class="feedback-id">{fid}</span>
                    <span class="source-tag">{row['Source']}</span>
                    <span class="badge {bc}">{row['Topic']}</span>
                    {f'<span style="background:rgba(0,0,0,0.3);border:1px solid {sc};border-radius:12px;'
                      f'padding:1px 8px;font-family:Space Mono,monospace;font-size:0.68rem;color:{sc};">{sent}</span>'
                      if sent else ''}
                    {f'<span style="font-family:Space Mono,monospace;font-size:0.68rem;color:#888;">{mh}</span>'
                      if mh else ''}
                </div>
                <div style="color:#ddd;font-size:0.9rem;line-height:1.6;margin-bottom:0.8rem;">
                    {row['Feedback'][:400]}{'...' if len(row['Feedback']) > 400 else ''}
                </div>
                <div style="border-top:1px solid rgba(255,255,255,0.06);padding-top:0.8rem;">
                    <span style="font-family:'Space Mono',monospace;font-size:0.65rem;color:#f9a825;
                                 letter-spacing:2px;text-transform:uppercase;">💡 AI Suggestion</span>
                    <div style="color:#b0b0b0;font-size:0.85rem;margin-top:4px;">{row.get('Suggestion', '')}</div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        if len(filtered) > 30:
            st.info(f"Showing 30 cards. Switch to Table for all {len(filtered)}.")

    st.divider()
    st.markdown("### 📥 Download")
    d1, d2, d3 = st.columns(3)
    with d1:
        ecols = [
            c for c in
            ['Feedback_ID', 'Source', 'Reviewer_Name', 'Feedback_Date', 'Feedback', 'Topic', 'Sentiment', 'Suggestion', 'Future_Action_Plan']
            if c in filtered.columns
        ]
        st.download_button(
            "⬇️ CSV — Feedback + AI",
            data=filtered[ecols].to_csv(index=False).encode('utf-8'),
            file_name=f"zeus_feedback_{sid}.csv",
            mime='text/csv',
            use_container_width=True,
        )
    with d2:
        txt = make_summary_txt(filtered, sid, uname)
        st.download_button(
            "⬇️ TXT Summary Report",
            data=txt.encode('utf-8'),
            file_name=f"zeus_report_{sid}.txt",
            mime='text/plain',
            use_container_width=True,
        )
    with d3:
        try:
            xls = make_feedback_excel(filtered, sid, uname)
            st.download_button(
                "⬇️ Excel — Feedback + AI (.xlsx)",
                data=xls,
                file_name=f"zeus_feedback_{sid}.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )
        except Exception as e:
            st.info(f"`pip install openpyxl` for Excel · {e}")

    st.divider()
    if st.button("🔄 Clear & Start Over"):
        st.session_state.results_df  = None
        st.session_state.analyzed    = False
        st.session_state.session_id  = None
        st.session_state.short_uuid  = None
        st.rerun()

# ══ TAB 5 — History ════════════════════════════════════════
with tab2:
    st.markdown(
        '<div class="card-title" style="font-family:Space Mono,monospace;font-size:0.75rem;'
        'letter-spacing:3px;text-transform:uppercase;color:#f9a825;">🗄️ Session History</div>',
        unsafe_allow_html=True,
    )
    try:
        sessions_df = get_all_sessions()
    except Exception:
        sessions_df = pd.DataFrame()

    if sessions_df.empty:
        st.info("No sessions yet. Run an analysis to start building history.")
    else:
        st.markdown(f"**{len(sessions_df)} session(s)** in local database")
        st.dataframe(
            sessions_df,
            use_container_width=True,
            column_config={
                "session_id":    st.column_config.TextColumn("Session ID",  width="large"),
                "user_name":     st.column_config.TextColumn("Analyst"),
                "created_date":  st.column_config.TextColumn("Date"),
                "created_time":  st.column_config.TextColumn("Time"),
                "source_type":   st.column_config.TextColumn("Source"),
                "total_entries": st.column_config.NumberColumn("Entries"),
                "notes":         st.column_config.TextColumn("Notes",       width="medium"),
            },
        )
        st.divider()
        st.markdown("### 🔍 Load & Export Session")
        sel_sid = st.selectbox(
            "Select Session",
            sessions_df['session_id'].tolist(),
            format_func=lambda x: (
                f"{x}  ·  "
                f"{sessions_df[sessions_df['session_id']==x]['created_date'].values[0]}  "
                f"{sessions_df[sessions_df['session_id']==x]['created_time'].values[0]}"
            ),
        )
        if sel_sid:
            try:
                entries = get_session_entries(sel_sid)
            except Exception:
                entries = pd.DataFrame()

            if not entries.empty:
                meta = sessions_df[sessions_df['session_id'] == sel_sid].iloc[0]
                st.markdown(
                    f'<div class="session-box">🔑 <strong>{sel_sid}</strong><br>'
                    f'👤 {meta["user_name"]} &nbsp;·&nbsp; '
                    f'📅 {meta["created_date"]} {meta["created_time"]} &nbsp;·&nbsp; '
                    f'📊 {len(entries)} entries &nbsp;·&nbsp; '
                    f'🌐 {meta["source_type"]}</div>',
                    unsafe_allow_html=True,
                )
                st.dataframe(
                    entries[['feedback_id', 'source', 'reviewer_name', 'feedback_date',
                              'feedback', 'topic', 'sentiment', 'suggestion', 'analyzed_at']],
                    use_container_width=True,
                    height=320,
                )
                hist_df = entries.rename(columns={
                    'feedback_id': 'Feedback_ID', 'source': 'Source',
                    'reviewer_name': 'Reviewer_Name', 'feedback_date': 'Feedback_Date',
                    'feedback': 'Feedback', 'topic': 'Topic',
                    'suggestion': 'Suggestion', 'sentiment': 'Sentiment',
                })
                h1, h2, h3 = st.columns(3)
                with h1:
                    ecols = [c for c in ['Feedback_ID', 'Source', 'Reviewer_Name', 'Feedback_Date',
                                         'Feedback', 'Topic', 'Sentiment', 'Suggestion']
                             if c in hist_df.columns]
                    st.download_button(
                        "⬇️ CSV",
                        data=hist_df[ecols].to_csv(index=False).encode('utf-8'),
                        file_name=f"zeus_{sel_sid}.csv",
                        mime='text/csv',
                        use_container_width=True,
                    )
                with h2:
                    txt = make_summary_txt(hist_df, sel_sid, meta['user_name'])
                    st.download_button(
                        "⬇️ TXT Report",
                        data=txt.encode('utf-8'),
                        file_name=f"zeus_{sel_sid}_report.txt",
                        mime='text/plain',
                        use_container_width=True,
                    )
                with h3:
                    try:
                        xls = make_feedback_excel(hist_df, sel_sid, meta['user_name'])
                        st.download_button(
                            "⬇️ Excel — Feedback + AI",
                            data=xls,
                            file_name=f"zeus_{sel_sid}.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            use_container_width=True,
                        )
                    except Exception:
                        st.info("`pip install openpyxl`")

                st.markdown("<br>", unsafe_allow_html=True)
                try:
                    db_xls = make_db_excel()
                    if db_xls:
                        st.download_button(
                            "⬇️ Excel — Full Database (.xlsx)",
                            data=db_xls,
                            file_name="zeus_full_database.xlsx",
                            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            use_container_width=True,
                        )
                except Exception:
                    pass

                try:
                    full_db = get_full_db_export()
                    st.download_button(
                        "⬇️ CSV — Full Database",
                        data=full_db.to_csv(index=False).encode('utf-8'),
                        file_name="zeus_full_database.csv",
                        mime='text/csv',
                        use_container_width=True,
                    )
                except Exception:
                    pass

                st.divider()
                if st.button(f"🗑️ Delete Session `{sel_sid}`"):
                    try:
                        delete_session(sel_sid)
                        st.success("Session deleted.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Could not delete: {e}")
            else:
                st.warning("No entries found for this session.")

        st.divider()
        with st.expander("⚠️ Danger Zone — Clear All History"):
            st.warning("Permanently deletes ALL sessions and entries from the local database.")
            if st.button("🔥 Wipe All History"):
                conn = sqlite3.connect(DB_PATH)
                conn.execute("DELETE FROM feedback_entries")
                conn.execute("DELETE FROM sessions")
                conn.commit()
                conn.close()
                st.success("All history cleared.")
                st.rerun()